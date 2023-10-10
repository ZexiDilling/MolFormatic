import os
from pathlib import Path

from natsort import natsort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from gui_functions import import_ms_data, get_peak_information, name_changer, purity_ops, add_start_end_time
from gui_popup import ms_raw_name_guard


def _grab_excel_qc_data(file):

    wb = load_workbook(file)

    # Grab EU data
    eu_sample_data = {}
    ws = wb["EU"]
    for row_index, row in enumerate(ws):
        if row_index > 0:
            rack_id = str(ws.cell(row=row_index + 1, column=1).value)
            well_id = ws.cell(row=row_index + 1, column=2).value
            tube_id = ws.cell(row=row_index + 1, column=3).value
            sample_name = ws.cell(row=row_index + 1, column=4).value
            # box = ws.cell(row=row_index + 1, column=5).value
            plate_pos = ws.cell(row=row_index + 1, column=5).value
            weight = ws.cell(row=row_index + 1, column=6).value
            eu_sample_data[sample_name] = {"rack_id": rack_id, "well_id": well_id, "weight": weight}
    ws = wb["DK"]
    sample_data = {}
    samples = []
    wrong_name = []
    for row_index, row in enumerate(ws):
        if row_index > 0:
            rack_id_dk = str(ws.cell(row=row_index + 1, column=1).value)
            well_id_dk = ws.cell(row=row_index + 1, column=2).value
            tube_id = ws.cell(row=row_index + 1, column=3).value
            sample_name = ws.cell(row=row_index + 1, column=4).value
            box = ws.cell(row=row_index + 1, column=5).value
            box_pos = ws.cell(row=row_index + 1, column=6).value
            weight_dk = ws.cell(row=row_index + 1, column=8).value
            temp_name = f"B{rack_id_dk.removeprefix('3000')}_{well_id_dk}"

            try:
                eu_sample_data[sample_name]
            except KeyError:
                wrong_name.append(sample_name)
                rack_id = None
                well_id = None
                weight = None
            else:
                rack_id = eu_sample_data[sample_name]["rack_id"]
                well_id = eu_sample_data[sample_name]["well_id"]
                weight = eu_sample_data[sample_name]["weight"]

            samples.append(sample_name)
            sample_data[temp_name] = {"sample_name": sample_name, "box_dk": box,
                                      "box_pos_dk": box_pos, "rack_id": rack_id, "well_id": well_id, "weight": weight,
                                      "rack_id_dk": rack_id_dk, "well_id_dk": well_id_dk, "weight_dk": weight_dk}

    wb.close()

    return sample_data


def _grab_smiles(folder):
    smiles_data = {}
    files = list(folder.iterdir())
    for file in files:
        if file.suffix == ".xlsx":
            continue
            # wb = load_workbook(file)
            # ws = wb.active

        elif file.suffix == ".csv":
            with open(file) as f:
                for row_index, line in enumerate(f):
                    values = line.split(";")
                    id = values[0]
                    chemical_formular = values[1]
                    mass = values[2]
                    smiles = values[3]
                    sample_name = values[4]
                    box = file.name.split("-")[2]
                    box_place = row_index

                    smiles_data[sample_name] = {"smiles": smiles, "mass": mass, "formel": chemical_formular,
                                                "box": box, "box_place": box_place}

    return smiles_data


def _grab_tube_barcode(folder):
    tube_list = ["MI64000005883", "MI64000005887"]
    tube_codes = {}
    all_files = folder.glob("**/*.txt")
    for files in list(all_files):
        with files.open() as file:
            lines = file.readlines()

            rack_id = files.name.removesuffix(".txt")
            tube_codes[rack_id] = {}
            for line_index, line in enumerate(lines):
                if line_index != 0:
                    line = line.split(";")
                    tube_id = line[-1].strip()

                    temp_pos = list(line[0])
                    tube_pos = f"{temp_pos[0]}{temp_pos[-1]}"
                    tube_codes[rack_id][tube_pos] = tube_id

                    if tube_id in tube_list:
                        print("--------------")
                        print(tube_id)
                        print(rack_id)
                        print("--------------")

    return tube_codes


def _reformat_tubes(old_tube_codes, new_tube_codes):
    formatting_tube_codes = {}
    rack_reformat = {}
    for new_racks in new_tube_codes:
        rack_reformat[new_racks] = {"tubes": [],
                                    }
        for tube_pos in new_tube_codes[new_racks]:
            tube_id = new_tube_codes[new_racks][tube_pos]
            rack_reformat[new_racks]["tubes"].append(tube_id)
            rack_reformat[new_racks][tube_id] = tube_pos

    for old_racks in old_tube_codes:
        for tube_pos in old_tube_codes[old_racks]:
            tube_id = old_tube_codes[old_racks][tube_pos]
            for new_racks in rack_reformat:
                if tube_id in rack_reformat[new_racks]["tubes"]:
                    new_tube_pos = rack_reformat[new_racks][tube_id]
                    try:
                        formatting_tube_codes[tube_id]
                    except KeyError:
                        formatting_tube_codes[tube_id] = {"rack": new_racks, "pos": new_tube_pos}
                    else:
                        formatting_tube_codes[tube_id]["rack"] = new_racks
                        formatting_tube_codes[tube_id]["pos"] = new_tube_pos
    return formatting_tube_codes


def _add_barcodes(sample_data, old_tube_codes, formatting_tube_codes):
    samples_for_deletion = []
    barcode_list = []
    for samples in sample_data:
        temp_rack = sample_data[samples]["rack_id"]
        temp_well = sample_data[samples]["well_id"]
        try:
            old_tube_codes[temp_rack]
        except KeyError:
            print(temp_rack)
            samples_for_deletion.append(samples)
        else:
            if temp_rack:
                temp_tube_id =  old_tube_codes[temp_rack][temp_well]
                try:
                    formatting_tube_codes[temp_tube_id]
                except KeyError:
                    samples_for_deletion.append(samples)
                    continue
                else:
                    new_rack = formatting_tube_codes[temp_tube_id]["rack"]
                    new_pos = formatting_tube_codes[temp_tube_id]["pos"]

                sample_data[samples]["tube_id"] = temp_tube_id
                sample_data[samples]["rack_id"] = new_rack
                sample_data[samples]["well_id"] = new_pos
                barcode_list.append(samples)
            else:
                sample_data[samples]["tube_id"] = "Missing"
                samples_for_deletion.append(samples)

    for samples in samples_for_deletion:
        sample_data.pop(samples)

    return barcode_list


def _cleaner(sample_data):
    temp_samples = []
    only_samples = []
    for samples in sample_data:
        if "blank" in samples:
            continue
        try:
            sample_data[samples]["mass"]
        except KeyError:
            temp_samples.append(samples)
        else:
            only_samples.append(samples)
    for samples in temp_samples:
        sample_data.pop(samples)

    return only_samples


def write_to_excel(file_name, file_location, sample_data, analysed_sample, missing_samples, row_count, include_headlines, counter, barcode_list):
    output_file = f"{file_location}\{file_name}.xlsx"

    cons_1 = 25
    cons_2 = 0.01
    cons_3 = 0.01
    cons_min = 0.5
    cons_max = 1.5


    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    red = "00FF0000"
    green = "0000FF00"
    pink = "00FF00FF"
    yellow = "00FFFF00"
    initial_clm = clm_counter = 1

    row_counter = row_count

    if include_headlines:
        headlines = ["Box", "Well", "Tube_ID", "Compound_ID", "formel", "box", "place_dk", "place_running",
                     "mol_weight", "min_amount", "max_amount",
                     "current_amount (mg)", "Amount for DK (g)", "mass_found", "ion", "ion_mass", "purity", "smiles"]

        for headline in headlines:
            ws.cell(column=clm_counter, row=row_counter).value = headline
            clm_counter += 1
        row_counter += 1
        clm_counter = initial_clm
    for samples in analysed_sample:
        if samples not in barcode_list:
            continue
        if "blank" in samples.casefold():
            continue
        if samples in missing_samples:
            print("MISSING SAMPLES!!!")
            print(samples)
            continue
        try:
            sample_data[samples]["formel"]
        except KeyError:
            print("NO FORMEL!!!")
            print(samples)
            continue

        try:
            sample_data[samples]["tube_id"]
        except KeyError:
            print("NO TUBE ID")
            print(samples)
            continue

        ws.cell(column=clm_counter + 0, row=row_counter).value = sample_data[samples]["rack_id"]
        ws.cell(column=clm_counter + 1, row=row_counter).value = sample_data[samples]["well_id"]
        ws.cell(column=clm_counter + 2, row=row_counter).value = sample_data[samples]["tube_id"]
        ws.cell(column=clm_counter + 3, row=row_counter).value = sample_data[samples]["sample_name"]
        ws.cell(column=clm_counter + 4, row=row_counter).value = sample_data[samples]["formel"]
        ws.cell(column=clm_counter + 5, row=row_counter).value = sample_data[samples]["box_dk"]
        ws.cell(column=clm_counter + 6, row=row_counter).value = sample_data[samples]["box_pos_dk"]
        ws.cell(column=clm_counter + 7, row=row_counter).value = sample_data[samples]["box_place_running"]

        mass = sample_data[samples]["mass"]
        ws.cell(column=clm_counter + 8, row=row_counter).value = mass
        weight_amount = sample_data[samples]["weight"]

        if weight_amount:
            if "mg" in weight_amount:
                weight_amount = weight_amount.split(" ")[0].strip()
            if "," in weight_amount:
                weight_amount = weight_amount.replace(",", ".")
        else:
            weight_amount = "None"

        try:
            float(weight_amount)
        except:
            weight_amount = f"{weight_amount}"
            min_amount = "None"
            max_amount = "None"
        else:
            try:
                mass = float(mass)
            except ValueError:

                mass = float(mass.replace(",", "."))

            weight_amount = float(weight_amount)
            min_amount = ((mass/cons_1)*cons_min/cons_2)*cons_3
            max_amount = ((mass/cons_1)*cons_max/cons_2)*cons_3

        ws.cell(column=clm_counter + 9, row=row_counter).value = min_amount
        ws.cell(column=clm_counter + 10, row=row_counter).value = max_amount
        ws.cell(column=clm_counter + 11, row=row_counter).value = weight_amount

        if type(weight_amount) != str:
            if min_amount < weight_amount < max_amount:
                ws.cell(column=clm_counter + 11, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=green)
            elif min_amount > weight_amount:
                ws.cell(column=clm_counter + 11, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=pink)
                counter["below"] += 1
            elif weight_amount > max_amount:
                ws.cell(column=clm_counter + 11, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=yellow)
                counter["above"] += 1
        else:
            ws.cell(column=clm_counter + 11, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=red)
            counter["missing_weight"] += 1
        ws.cell(column=clm_counter + 12, row=row_counter).value = sample_data[samples]["weight_dk"]
        if sample_data[samples]["max_hit"]:

            ws.cell(column=clm_counter + 13, row=row_counter).value = "Found"
            ws.cell(column=clm_counter + 13, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=green)
            ws.cell(column=clm_counter + 14, row=row_counter).value = sample_data[samples]["max_hit"]["ion"]
            ws.cell(column=clm_counter + 15, row=row_counter).value = sample_data[samples]["max_hit"]["ion_mass"]
            ws.cell(column=clm_counter + 16, row=row_counter).value = sample_data[samples]["max_hit"]["purity"][0]
            counter["hit"] += 1
        else:
            ws.cell(column=clm_counter + 13, row=row_counter).value = "No Hit"
            ws.cell(column=clm_counter + 13, row=row_counter).fill = PatternFill(fill_type="solid", fgColor=red)
            counter["no_hit"] += 1

        ws.cell(column=clm_counter + 17, row=row_counter).value = sample_data[samples]["smiles"]
        row_counter += 1
        counter["sample_amount"] += 1

    wb.save(output_file)
    return row_counter


def _combine(sample_data, smiles_data):
    missing_samples = {}
    for samples in sample_data:
        sample_name = sample_data[samples]["sample_name"]
        try:
            smiles_data[sample_name]
        except KeyError:
            rack_id = sample_data[samples]["rack_id_dk"]
            well_id = sample_data[samples]["well_id_dk"]
            box = sample_data[samples]["box_dk"]
            box_place = sample_data[samples]["box_pos_dk"]
            missing_samples[samples] = {"sample_name": sample_name, "rack_id": rack_id, "well_id": well_id,
                                        "box": box, "box_place": box_place}

        else:
            sample_data[samples]["smiles"] = smiles_data[sample_name]["smiles"]
            sample_data[samples]["mass"] = smiles_data[sample_name]["mass"]
            sample_data[samples]["formel"] = smiles_data[sample_name]["formel"]
            sample_data[samples]["box_running"] = smiles_data[sample_name]["box"]
            sample_data[samples]["box_place_running"] = smiles_data[sample_name]["box_place"]

    return sample_data, missing_samples


def missing_sample_report(missing_samples, location):

    wb = Workbook()
    ws = wb.active
    initial_clm = clm_counter = row_counter = 1

    headlines = ["Rack", "Well", "Compound_ID", "box", "placement"]

    for headline in headlines:
        ws.cell(column=clm_counter, row=row_counter).value = headline
        clm_counter += 1
    row_counter += 1
    clm_counter = initial_clm
    for samples in missing_samples:
        print(missing_samples[samples])
        ws.cell(column=clm_counter + 0, row=row_counter).value = missing_samples[samples]["rack_id"]
        ws.cell(column=clm_counter + 1, row=row_counter).value = missing_samples[samples]["well_id"]
        ws.cell(column=clm_counter + 2, row=row_counter).value = missing_samples[samples]["sample_name"]
        ws.cell(column=clm_counter + 3, row=row_counter).value = missing_samples[samples]["box"]
        ws.cell(column=clm_counter + 4, row=row_counter).value = missing_samples[samples]["box_place"]
        row_counter += 1

    file_name = "missing_samples"
    file = output_file = f"{location}\{file_name}.xlsx"
    wb.save(file)


def testing(sample_data):
    cons_1 = 25
    cons_2 = 0.01
    cons_3 = 0.01
    cons_min = 0.5
    cons_max = 1.5

    for samples in sample_data:
        print("------------------")
        print(samples)
        print(sample_data[samples])
        mass = sample_data[samples]["mass"]
        current_amount = sample_data[samples]["weight"]
        if not current_amount:
            continue

        if "mg" in current_amount:
            current_amount = current_amount.split(" ")[0].strip()

        if current_amount:
            weight_amount = current_amount.split(" ")[0]
            if "," in weight_amount:
                weight_amount = weight_amount.replace(",", ".")
        else:
            weight_amount = "None"

        try:
            float(weight_amount)
        except:
            weight_amount = f"{weight_amount}"
            min_amount = "None"
            max_amount = "None"
        else:
            try:
                mass = float(mass)
            except ValueError:

                mass = float(mass.replace(",", "."))

            weight_amount = float(weight_amount)
            min_amount = ((mass / cons_1) * cons_min / cons_2) * cons_3
            max_amount = ((mass / cons_1) * cons_max / cons_2) * cons_3

        print(f" weight_amount - {weight_amount}")
        print(f" min_amount - {min_amount}")
        print(f" max_amount - {max_amount}")
        print(f" mass - {mass}")
        print("------------------")


if __name__ == "__main__":

    import configparser
    config = configparser.ConfigParser()
    config.read("config.ini")
    all_folders = Path(r"C:\Users\phch\Desktop\test\QC")

    folders = [f.path for f in os.scandir(all_folders) if f.is_dir()]
    include_headline = True
    sample_data = None
    compound_data = False
    slope_threshold = 5000000
    uv_threshold = 200000
    rt_solvent_peak = 0.0

    wavelength_data = "all"
    db_data = False
    # sample_data_file = r"C:\Users\phch\Desktop\test\Academic_collection_shadi.xlsx"
    sample_data_file = r"C:\Users\phch\Desktop\test\Academic_collection.xlsx"
    smiles_folder = Path(r"O:\Organisk kemi\molecular library running plate\FullPlates\NEW FILES\CSV FILES")
    # smiles_folder = Path(r"C:\Users\phch\Desktop\test\excel")
    barcode_folder = Path(r"C:\Users\phch\Desktop\test\barcodes_tubes")
    new_barcode_folder = Path(r"C:\Users\phch\Desktop\test\new_barcodes")
    file_name = "drive_1_"
    # file_location = r"C:\Users\phch\Desktop\test"
    # print("--------------peak info got--------------")
    # sample_data = _grab_excel_qc_data(sample_data_file)
    # print("--------------sample_data--------------")
    # smiles_data = _grab_smiles(smiles_folder)
    # print("--------------smiles_data--------------")
    old_tube_codes = _grab_tube_barcode(barcode_folder)
    # print("--------------Old-Barcodes--------------")
    # new_tube_codes = _grab_tube_barcode(new_barcode_folder)
    # print("--------------New-Barcodes--------------")
    #
    # formatting_tube_codes = _reformat_tubes(old_tube_codes, new_tube_codes)
    # print("--------------reformat--------------")
    # barcode_list = _add_barcodes(sample_data, old_tube_codes, formatting_tube_codes)
    # print("--------------Added barcodes--------------")
    # sample_data, missing_samples = _combine(sample_data, smiles_data)
    #
    # print("--------------combine--------------")
    # only_samples = _cleaner(sample_data)
    # # done_folders = []
    # done_folders = ['C:\\Users\\phch\\Desktop\\test\\QC\\batch_1']
    # row_count = 48
    # include_headline = False
    # location = r"C:\Users\phch\Desktop\test"
    # missing_sample_report(missing_samples, location)
    # # testing(sample_data)
    # # counter = {'below': 0, 'above': 0, 'missing_weight': 0, 'hit': 0, 'no_hit': 0, 'sample_amount': 0}
    # counter = {'below': 19, 'above': 1, 'missing_weight': 0, 'hit': 46, 'no_hit': 0, 'sample_amount': 46}
    #
    # for ms_folder in folders:
    #     if ms_folder in done_folders:
    #         continue
    #     print("Starting")
    #     all_data = import_ms_data(ms_folder)
    #
    #     _, purity_samples, purity_data, missing_samples = all_data
    #     print("All data ")
    #
    #     peak_information, peak_table_data, sample_peak_dict = get_peak_information(
    #         purity_data, slope_threshold, uv_threshold, rt_solvent_peak, sample_data,
    #         wavelength_data)
    #
    #     print("excel data got")
    #     only_samples = natsort.natsorted([keys for keys in only_samples])
    #     excel_data_samples = natsort.natsorted([keys for keys in sample_data])
    #
    #     if only_samples != excel_data_samples:
    #         # If they are not the same, a popup will show, where you can set names for the data.
    #         new_names = ms_raw_name_guard(only_samples, excel_data_samples, db_data,
    #                                       config)
    #
    #         name_changer(new_names, purity_data, sample_data, peak_information,
    #                      sample_peak_dict)
    #     print("names approved")
    #     ms_mode = "ms_pos"
    #     delta_mass = 0.25
    #     mz_threshold = 100000
    #     peak_amounts = 10
    #
    #     purity_overview_table_data, purity_peak_list_table_data = purity_ops(sample_data, purity_data, peak_information,
    #                                              ms_mode, delta_mass, mz_threshold,
    #                                              peak_amounts)
    #
    #     analysed_sample = [samples for samples in purity_data]
    #
    #     # add_start_end_time(purity_peak_list_table_data, sample_peak_dict)
    #
    #     row_count = write_to_excel(file_name, file_location, sample_data, analysed_sample, missing_samples, row_count,
    #                                include_headline, counter, barcode_list)
    #     include_headline = False
    #     print(f"----------{ms_folder} - Done!!! -------------------")
    #     all_data.clear()
    #     purity_samples.clear()
    #     purity_data.clear()
    #     missing_samples.clear()
    #     peak_information.clear()
    #     peak_table_data.clear()
    #     sample_peak_dict.clear()
    #     purity_overview_table_data.clear()
    #     purity_peak_list_table_data.clear()
    #     analysed_sample.clear()
    #     done_folders.append(ms_folder)
    #     print("---------------")
    #     print(row_count)
    #     print(done_folders)
    #     print("---------------")
    #     print(counter)
