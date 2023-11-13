from copy import deepcopy

from openpyxl import Workbook

from chem_operators import ChemOperators
from database_handler import DataBaseFunctions


def fp_search(config):

    dbf = DataBaseFunctions(config)
    co = ChemOperators()
    all_compound_rows = dbf.return_table_data("compound_main", "")
    methode = "morgan"
    threshold = 50
    compound_list = [4036510319, 4036578345, 4036508932, 4036514865, 4036517980, 4036493625, 4036490682, 4036570011,
                     4036513098, 4036578803, 4036506769, 4036489674, 4036575267, 4036517734, 4036509234, 4036493212,
                     4036506855, 4036495373, 4036507798, 4036493200, 4036495038, 4036490694, 4036505166]

    # for compound_index, compound_id in enumerate(compound_list):
    #     print(all_compound_rows[compound_id])
    #
    wb = Workbook()
    ws = wb.active
    row_counter = clm_counter = 1

    headlines = ["compound_id", "smiles", "compared_id", "compared_smiles", "score"]

    for headline in headlines:
        ws.cell(column=clm_counter, row=row_counter).value = headline
        clm_counter += 1
    row_counter += 1
    clm_counter = 1

    compound_list_len = [_ for _ in compound_list]
    compound_list_len = len(compound_list_len)
    for compound_index, compound_id in enumerate(compound_list):
        current_smiles = all_compound_rows[compound_id]["smiles"]
        row_data = co.structure_search(methode, threshold, all_compound_rows, current_smiles, remove_data=False)
        print("row_data_got")
        for compound in row_data:
            if threshold < all_compound_rows[compound]["match_score"] < 100:

                ws.cell(column=clm_counter + 0, row=row_counter).value = compound_id
                ws.cell(column=clm_counter + 1, row=row_counter).value = current_smiles
                ws.cell(column=clm_counter + 2, row=row_counter).value = compound
                ws.cell(column=clm_counter + 3, row=row_counter).value = all_compound_rows[compound]["smiles"]
                ws.cell(column=clm_counter + 4, row=row_counter).value = all_compound_rows[compound]["match_score"]
                row_counter += 1

        print(f"{compound_index + 1}/{compound_list_len} Done")

    location = r"C:\Users\phch\Desktop\test"
    file_name = "smiles_seachs"
    output_file = f"{location}\{file_name}.xlsx"
    wb.save(output_file)
    #


def temp_change(config):

    dbf = DataBaseFunctions(config)

    table = "biological_compound_data"
    hit = 0
    data_value = 1
    headline = "hit"
    # data_value = "alpha_so84"
    # headline = "assay_plate"

    data = dbf.find_data_single_lookup(table, data_value, headline)
    for counter, rows in enumerate(data):
        # if counter == 0:
        #     print(rows)
        key_value = rows[1]
        key_headline = "bio_data_id"
        table_data = {"bio_data_id": rows[1],
                      "compound_id": rows[2],
                      "assay_plate": rows[3],
                      "assay_well": rows[4],
                      "score": rows[5],
                      "hit": 0,
                      "concentration": rows[7],
                      "raw_data": rows[8],
                      "approved": rows[9],
                      "note": rows[10],
                      "transferred": rows[11]}

        dbf.update_database_items(table, table_data, key_value, key_headline)


if __name__ == "__main__":

    import configparser
    config = configparser.ConfigParser()
    config.read("config.ini")

    #
    # fp_search(config)

    # compound_search =
    # smiles =
    # threshold = 0
    # sub_search_methode = "morgan"
    # table = "compound_main"
