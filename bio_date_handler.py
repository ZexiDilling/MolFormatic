import configparser
import pandas as pd
from json_handler import plate_dict_reader
import re
from statistics import mean, stdev
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from matplotlib import colors

from bio_data_functions import *


class BIOAnalyser:
    """
    :param config: the config file for the program
    :type config: configparser.ConfigParser
    :param bio_plate_report_setup: dict over what state wells should be in, to be printed on the report sheet.
    :type bio_plate_report_setup: dict
    """
    def __init__(self, config, bio_plate_report_setup):
        self.config = config
        self.cal_stuff = {"avg": mean, "stdev": stdev}

        self.well_states_report_method = bio_plate_report_setup["well_states_report_method"]
        self.well_states_report = bio_plate_report_setup["well_states_report"]
        self.plate_report_calc_dict = bio_plate_report_setup["calc_dict"]
        self.plate_calc_dict = bio_plate_report_setup["plate_calc_dict"]
        self.plate_analysis = bio_plate_report_setup["plate_analysis_dict"]
        self.z_prime_calc = bio_plate_report_setup["z_prime_calc"]
        self.heatmap_colours = bio_plate_report_setup["heatmap_colours"]
        self.pora_threshold = bio_plate_report_setup["pora_threshold"]


    def __str__(self):
        """
        A class that handles the data from a Tecan platereader, where the data is in an excel formate.
        It does calculations and analysis of the data, and makes a final report based on everything
        :return: the analysed data
        """

    def _plate_well_dict(self):
        """
        Makes a dict over the state of each well (empty, sample, blank...)

        :return: pw_dict: A dict over the wells and what state they are in.
        :rtype: dict
        """
        pw_dict = {}
        for layout in self.plate:
            for counter in self.plate[layout]:
                try:
                    pw_dict[self.plate[layout][counter]["well_id"]] = self.plate[layout][counter]["state"]
                except TypeError:
                    pass

        return pw_dict

    def _data_converter(self, all_data, well_type):
        """
        convert raw data in the analysed data

        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param well_type: A dict over what state/type each well/cell is in.
        :type well_type: dict
        :return:
            - all_data: A dict over all plate date. all the analysed data will be added to this dict
            - pw_dict: A dict over the wells and what state they are in.
        :rtype:
            - dict
            - dict
        """

        pw_dict = self._plate_well_dict()

        for methode in self.plate_analysis:
            if self.plate_analysis[methode]["use"]:
                self._well_calculations(well_type, all_data, methode)

        all_data["calculations"]["other"] = {}
        if self.z_prime_calc:
            all_data["calculations"]["other"]["z_prime"] = z_prime_calculator(all_data, "normalised")

        return all_data, pw_dict

    def _well_calculations(self, well_type, all_data, methode):
        """
        Calculate each analyse methode for each well

        :param well_type: A dict for each state (empty, sample, blank...) with a list of the wells in that state
        :type well_type: dict
        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param methode: what analyse method is being used
        :return: The calculations of avg and stdev added to all_data
        """
        if methode != "original":
            all_data["plates"][methode] = {}
            all_data["plates"][methode]["wells"] = {}
        for state in well_type:
            all_data["plates"][methode][state] = []
            for well in well_type[state]:
                all_data["plates"][methode]["wells"][well] = self.plate_analysis[methode]["methode"](all_data, well)
                all_data["plates"][methode][state].append(well)

        all_data["calculations"][methode] = {}

        for state in well_type:
            all_data["calculations"][methode][state] = {}
            for calc in self.cal_stuff:
                if self.plate_calc_dict[methode][calc]:
                    try:
                        all_data["calculations"][methode][state][calc] = self.cal_stuff[calc](
                            [all_data["plates"][methode]["wells"][well] for well in all_data["plates"][methode][state]])
                    except ValueError:
                        all_data["calculations"][methode][state][calc] = None

    def _cal_info(self, ws, init_col, counter_row, temp_dict, methode):
        """
        Writes in the calculation information.

        :param ws: The worksheet for the excel filere where the data is added
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param init_col: column to writing to
        :type init_col: int
        :param counter_row: a counter for what row to write to
        :type counter_row: int
        :param temp_dict: the dict with the data for each well
        :type temp_dict: dict
        :param methode: the analysed method
        :type methode: str
        :return: counter_row: the next row to write on.
        :rtype: int
        """
        temp_row = counter_row
        for state in temp_dict["plates"][methode]:
            temp_col = init_col
            if state != "wells":
                if self.plate_calc_dict[methode]["state"][state]:
                    for calc in temp_dict["calculations"][methode][state]:
                        if counter_row == temp_row:
                            ws[ex_cell(counter_row, temp_col + 1)] = calc
                            ws[ex_cell(counter_row, temp_col + 1)].font = Font(b=True)
                        if temp_col == init_col:
                            ws[ex_cell(counter_row + 1, temp_col)] = state
                            ws[ex_cell(counter_row + 1, temp_col)].font = Font(b=True)
                        ws[ex_cell(counter_row + 1, temp_col + 1)] = temp_dict["calculations"][methode][state][calc]
                        temp_col += 1
                    counter_row += 1
        return counter_row

    def _write_plate(self, ws, counter_row, temp_dict, methode, well_row_col, pw_dict):
        """
        Writes the data for each analyse into the excel file including the calculations

        :param ws: The worksheet for the excel filere where the data is added
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param counter_row: What row to write to
        :type counter_row: int
        :param temp_dict: The dict for the specific analysed method
        :type temp_dict: dict
        :param methode: What analysed method are being looked at
        :type methode: str
        :param well_row_col: All the headlines for each row and column
        :type well_row_col: dict
        :param pw_dict: a dict for each well and it's state (empty, sample, blank...)
        :type pw_dict:dict
        :return: counter_row: the next row to write on.
        :rtype: int
        """
        indent_col = 3
        indent_row = 3
        init_row = counter_row + indent_row
        init_col = indent_col
        translate_wells_to_cells = {}
        counter_row += indent_row
        for index_row, row in enumerate(well_row_col["well_row"]):

            # sets the headline and colour for the headline for row
            ws.cell(column=-1 + indent_col, row=counter_row, value=row).fill = \
                PatternFill("solid", fgColor="DDDDDD")
            for index_col, col in enumerate(well_row_col["well_col"]):
                if index_row == 0:
                    # Merge cell above tables, and writes the name of the method used for the plate
                    # ws.merged_cells(start_row=counter_row - 2, start_column=indent_col - 1,
                    #                 end_row=counter_row - 2, end_column=indent_col + 1)
                    ws.cell(column=indent_col - 1, row=counter_row - 2, value=methode).font = Font(b=True)

                    # sets the headline and colour for the headline for column
                    ws.cell(column=index_col + indent_col, row=counter_row - 1, value=int(col)).fill = \
                        PatternFill("solid", fgColor="DDDDDD")

                temp_well = f"{row}{col}"
                temp_cell = ex_cell(counter_row, index_col + indent_col)
                translate_wells_to_cells[temp_well] = temp_cell
                # Writes the data in for each well. ignore wells witch state == empty  - - - - -
                #
                #ADD TO SETTINGS!!!
                #
                #
                if temp_well not in temp_dict["plates"][methode]["empty"]:
                    ws.cell(column=index_col + indent_col, row=counter_row,
                            value=temp_dict["plates"][methode]["wells"][temp_well])
            counter_row += 1
        free_col = len(well_row_col["well_col"]) + indent_col

        # Writes the info for the calculation for each method

        if self.plate_calc_dict[methode]["use"]:
            counter_row = self._cal_info(ws, init_col, counter_row, temp_dict, methode)

        # colour wells depending on what state the wells are (sample, blank, min, max...) and add a reading guide.
        if self.plate_analysis[methode]["state_map"]:
            state_mapping(self.config, ws, translate_wells_to_cells, self.plate, init_row, free_col, temp_dict, methode)

        # colour in the heat map, if sets to active. Can set for each method
        if self.plate_analysis[methode]["heatmap"]:
            heatmap(self.config, ws, pw_dict, translate_wells_to_cells, self.heatmap_colours)

        if self.plate_analysis[methode]["hit_map"]:
            hit_mapping(ws, temp_dict, self.pora_threshold, methode, translate_wells_to_cells, free_col, init_row)

        counter_row += 1
        return counter_row

    def cal_writer(self, ws, all_data, init_row):
        """
        Writes all the calculations to its own sheet for an overview.

        :param ws: The worksheet for the excel filere where the data is added
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param init_row: The first row to write data to.
        :type init_row: int
        :return: All the calculations writen in the worksheet called: report
        """
        indent_col = 2
        row_counter = init_row
        for plate_analysed in all_data["calculations"]:
            if self.plate_report_calc_dict[plate_analysed]["use"]:
                ws.cell(column=-1 + indent_col, row=row_counter, value=plate_analysed).font = Font(b=True)
                row_counter += 1
                for state in all_data["calculations"][plate_analysed]:
                    try:
                        self.plate_report_calc_dict[plate_analysed]["state"][state]
                        temp_name = "state"
                    except KeyError:
                        temp_name = "calc"

                    if self.plate_report_calc_dict[plate_analysed][temp_name][state]:
                        ws.cell(column=indent_col, row=row_counter, value=state).font = Font(b=True)
                        if plate_analysed != "other":
                            for calc in all_data["calculations"][plate_analysed][state]:
                                ws.cell(column=indent_col + 1, row=row_counter, value=calc)
                                ws.cell(column=indent_col + 2, row=row_counter,
                                        value=all_data["calculations"][plate_analysed][state][calc])

                        # Writes other calculations that's not avg or stdev, only writes z-prime
                        else:
                            if self.plate_report_calc_dict[plate_analysed]["calc"]["z_prime"]:
                                ws.cell(column=indent_col + 1, row=row_counter,
                                        value=all_data["calculations"][plate_analysed][state])
                        row_counter += 1
                row_counter += 1

    def _well_writer(self, ws, all_data, init_row):
        """
        Writes Well data from the different analysis method into the report sheet on the excel ark

        :param ws: The worksheet for the excel filere where the data is added
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param init_row: The first row to write data to.
        :type init_row: int
        :return: All the wells writen in a list in the worksheet called: report
        """
        indent_col = 6
        row_counter = init_row
        added = False

        for plate_analysed in all_data["plates"]:
            if self.well_states_report_method[plate_analysed]:
                # Writes headline for data inserts to see where the data is coming from
                ws.cell(column=indent_col, row=row_counter, value=plate_analysed).font = Font(b=True)
                row_counter += 1
                for counter in self.plate["well_layout"]:
                    for _ in self.plate["well_layout"][counter]:

                        # looks through the plate layout, finds the state for each well and check if it needs to be added
                        # based on bool-statment from well_states_report
                        if self.well_states_report[self.plate["well_layout"][counter]["state"]] and not added:
                            well = self.plate["well_layout"][counter]["well_id"]
                            ws.cell(column=indent_col + 1, row=row_counter, value=well)
                            ws.cell(column=indent_col + 2, row=row_counter,
                                    value=all_data["plates"][plate_analysed]["wells"][well])
                            added = True
                            row_counter += 1
                    added = False
                indent_col += 4
                row_counter = init_row

    def _report_writer_controller(self, wb, all_data):
        """
        pass the data into different modules to write data in to an excel ark

        :param wb: the excel ark / workbook
        :type wb: openpyxl.workbook.workbook.Workbook
        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :return: Create a new sheet in the workbook, called Report, and writes in wells and calculations depending on
            the analysis.
        """

        init_row = 2
        ws_report = wb.create_sheet("Report")
        self.cal_writer(ws_report, all_data, init_row)
        self._well_writer(ws_report, all_data, init_row)

    def _excel_controller(self, all_data, well_row_col, pw_dict):
        """
        Controls the flow for the data, to write into an excel file

        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param well_row_col: All the headlines for each row and column (numbers and letters for the cell values)
        :type well_row_col: dict
        :param pw_dict: dict over each well and what state it is (empty, sample, blank....)
        :type pw_dict: dict
        :return: A modified excel file, with all the calculations and data added, depending on the analysis method used.
        """

        wb = load_workbook(self.ex_file)
        ws_data = wb.create_sheet("analysis")
        counter_row = 0

        # sends each plate-analysed-type into the excel file
        for methode in all_data["plates"]:
            counter_row = self._write_plate(ws_data, counter_row, all_data, methode, well_row_col, pw_dict)
        self._report_writer_controller(wb, all_data)

        wb.save(self.ex_file)

    def bio_data_controller(self, ex_file, plate_layout, all_data, well_row_col, well_type, analysis, write_to_excel):
        """
        The control modul for the bio analysing

        :param ex_file: The excel file
        :type ex_file: str
        :param plate_layout: The layout for the plate with values for each well, what state they are in
        :type plate_layout: dict
        :param all_data: A dict over all plate date. all the analysed data will be added to this dict
        :type all_data: dict
        :param well_row_col: All the headlines for each row and column (numbers and letters for the cell values)
        :type well_row_col: dict
        :param well_type: A dict over what state/type each well/cell is in.
        :type well_type: dict
        :param analysis: The analysis method
        :type analysis: str
        :return: A dict over all plate date. all the analysed data will be added to this dict
        :rtype: dict
        """
        self.ex_file = ex_file
        self.plate = plate_layout

        all_data, pw_dict = self._data_converter(all_data, well_type)
        if write_to_excel:
            self._excel_controller(all_data, well_row_col, pw_dict)

        return all_data


if __name__ == "__main__":
    ...