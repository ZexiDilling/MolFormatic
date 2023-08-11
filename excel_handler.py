from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook, load_workbook
from math import ceil
import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
from openpyxl.drawing.image import Image as XLImage
import tempfile

from extra_functions import unit_converter
from info import plate_384_row, plate_384_column


def export_plate_layout(plate_layout, well_row_col, name, folder):
    wb = Workbook()
    ws = wb.active

    ini_row = 2
    ini_col = 2

    for index_row, row in enumerate(well_row_col["well_row"]):
        for index_col, col in enumerate(well_row_col["well_col"]):
            well = f"{row}{col}"
            colour = plate_layout[well]["colour"].removeprefix("#")

            ws.cell(column=index_col + ini_col, row=index_row + ini_row, value=well).fill \
                = PatternFill(fill_type="solid", fgColor=colour)

    wb.save(f"{folder}/{name}.xlsx")


def plate_dilution_write_vol_well_amount(config, file, dw_amount, add_wells, source_well_amount=None, layout=None):
    wb = load_workbook(file)
    try:
        ws_sample_data = wb["Sample_data"]
    except KeyError:
        return "Wrong Excel Layout"
    ws_extra_data = wb["Extra_data"]
    well_counter = 0
    if layout == "Row":
        plate_layout = plate_384_row
    else:
        plate_layout = plate_384_column

    for index_row, row in enumerate(ws_sample_data):
        # skipping headline row
        if index_row > 0:
            solvent = ws_sample_data.cell(row=index_row + 1, column=4).value
            plates = config["plate_types"][solvent]

            min_vol = float(config["plate_types_values"][plates].split(",")[0])
            sample_combinations = ws_sample_data.cell(row=index_row+1, column=8).value.split(",")
            sample_combinations_amount = len(sample_combinations)
            sample_replicates = ws_sample_data.cell(row=index_row+1, column=7).value
            plate_sets = ws_extra_data.cell(row=5, column=2).value
            max_conc = ws_extra_data.cell(row=2, column=2).value
            sample_conc = ws_sample_data.cell(row=index_row+1, column=3).value
            vol_needed = (min_vol * 10 * min_vol) * sample_combinations_amount * sample_replicates * plate_sets * (max_conc / sample_conc)

            # Finding dead volume depending on plate type
            dead_vol = float(config["plate_types_values"][plates].split(",")[1])
            # vol needed / max well working vol - dead_vol
            well_amount = ceil(vol_needed / (float(config["plate_types_values"][plates].split(",")[2])-dead_vol))


            # 10 % extra volume is needed per dilution well that is being made.
            if dw_amount.isdigit():
                if well_amount > dw_amount:
                    return "To few dilution well to complete the transferee"
                extra_vol = 1 + (0.1 * dw_amount)
            else:
                extra_vol = 1 + (0.1 * well_amount)

            vol_needed_total = (vol_needed * extra_vol) + (dead_vol * well_amount)
            ws_sample_data.cell(row=index_row + 1, column=9, value=vol_needed_total)
            ws_sample_data.cell(row=index_row + 1, column=10, value=well_amount)

            if add_wells:
                temp_well_list = []
                if source_well_amount == "Minimum":
                    source_well_amount = well_amount

                for i in range(source_well_amount):
                    temp_well_list.append(plate_layout[well_counter])
                    well_counter += 1
                well_list = ",".join(temp_well_list)
                ws_sample_data.cell(row=index_row + 1, column=1, value=well_list)


    wb.save(file)
    return "complete"


def _write_off_set_wells(ws_sample_data, dw_amount, well_layout):

    well_offset = 0
    # Get amount of wells that are used.
    # This method is assuming that the wells are in order
    for index_row, row_data in enumerate(ws_sample_data):
        if index_row > 0:
            well_count = len(ws_sample_data.cell(row=index_row + 1, column=1).value.split(","))
            well_offset += well_count

    # Adds the extra wells to the data:
    for index_row, row_data in enumerate(ws_sample_data):
        if index_row > 0:
            if ws_sample_data.cell(row=index_row + 1, column=6).value == "sample":
                temp_well_list = []
                well_list = ws_sample_data.cell(row=index_row + 1, column=1).value

                # Writes amount of wells between
                well_amount = len(well_list.split(","))
                ws_sample_data.cell(row=index_row + 1, column=11, value=well_amount)

                # Gets amount of wells depending on if it takes how much vol that is needed, or how many wells are
                # provided
                if dw_amount == "Minimum":
                    well_amount = ws_sample_data.cell(row=index_row + 1, column=10).value
                elif dw_amount == "Copy":
                    well_amount = len(well_list.split(","))
                else:
                    well_amount = dw_amount

                # Gets the wells from list of all wells for a 384 plate depending on how compound is added to the plate
                for i in range(well_amount):
                    if well_layout == "Row":
                        temp_well_list.append(plate_384_row[well_offset + i])
                    elif well_layout == "Column":
                        temp_well_list.append(plate_384_column[well_offset + i])

                # if I am added the new well's to the old list and writing all the data in the same cell
                # for wells in temp_well_list:
                #     well_list += f",{wells}"

                # Generate a string from list
                well_list = ",".join(temp_well_list)


                ws_sample_data.cell(row=index_row + 1, column=11, value=well_list)
                # increase the offset with the well added to the excel file
                well_offset += well_amount


def _info_dict(ws_sample_data):
    sample_info_dict = {}
    replicate_samples_list = []

    for index_row, row_data in enumerate(ws_sample_data):
        if index_row > 0:
            compound = ws_sample_data.cell(row=index_row + 1, column=2).value
            state = ws_sample_data.cell(row=index_row + 1, column=6).value

            well_list_1 = []
            well_list_2 = []

            for wells in ws_sample_data.cell(row=index_row + 1, column=1).value.split(","):
                well_list_1.append(wells)

            if ws_sample_data.cell(row=index_row + 1, column=11).value:
                for wells in ws_sample_data.cell(row=index_row + 1, column=11).value.split(","):
                    well_list_2.append(wells)

            conc = ws_sample_data.cell(row=index_row + 1, column=3).value
            replicates = ws_sample_data.cell(row=index_row + 1, column=7).value
            replicate_samples_list.append(replicates)
            solvent = ws_sample_data.cell(row=index_row + 1, column=4).value
            volume = ws_sample_data.cell(row=index_row + 1, column=5).value

            if ws_sample_data.cell(row=index_row + 1, column=8).value != "single":
                mix = True
            else:
                mix = False

            combinations = []
            if mix:
                for compounds in ws_sample_data.cell(row=index_row + 1, column=8).value.split(","):
                    combinations.append(compounds)

            sample_info_dict[compound] = {"state": state,
                                          "well_1": well_list_1,
                                          "well_2": well_list_2,
                                          "mix": mix,
                                          "conc": conc,
                                          "combination": combinations,
                                          "replicates": replicates,
                                          "solvent": solvent,
                                          "volume": volume}

    return sample_info_dict, replicate_samples_list


def plate_dilution_excel(file, save_plates, dw_amount="max", well_layout="row"):
    wb = load_workbook(file)
    ws_sample_data = wb["Sample_data"]
    ws_extra_data = wb["Extra_data"]

    # MISSING PP PLATES / AQ DISOLVED COMPOUNDS

    if save_plates:
        _write_off_set_wells(ws_sample_data, dw_amount, well_layout)

        # file = file.removesuffix(".xlsx")
        # file = f"{file}_added_wells.xlsx"
        wb.save(file)

    sample_info_dict, replicate_samples_list = _info_dict(ws_sample_data)
    replicate_samples_max = max(replicate_samples_list)
    replicate_plate_sets = ws_extra_data.cell(row=5, column=2).value
    dilution_factor = ws_extra_data.cell(row=3, column=2).value
    concentration_counter = ws_extra_data.cell(row=4, column=2).value
    control_vol = ws_extra_data.cell(row=7, column=2).value
    control_conc = ws_extra_data.cell(row=6, column=2).value

    return sample_info_dict, replicate_samples_max, replicate_plate_sets, dilution_factor, concentration_counter, \
           control_vol, control_conc


def purity_sample_layout_export(config, table_data, headings):
    output_folder = config["folders"]["main_output_folder"]
    file_name = f"{output_folder}/purity_sample_layout.xlsx"
    temp_df = pd.DataFrame(table_data, columns=headings)
    temp_df.to_excel(file_name)
    return file_name


def purity_sample_layout_import(file, table_headings):
    temp_df = pd.read_excel(file)
    temp_df = temp_df.to_dict("index")
    table_data = []
    for rows in temp_df:
        temp_data = [temp_df[rows][table_headings[0]], temp_df[rows][table_headings[1]], temp_df[rows][table_headings[2]]]
        table_data.append(temp_data)
    return table_data


def well_compound_list(file):
    """
    Takes excel file with wells in clm 1, compound name in clm 2, volume in clm 3 in uL, barcode in clm 4,
    plate type in clm 5 and compound type in clm 6.
    :param file: a file with data for a sourceplate with information about where compounds are placed
    :type file: pathlib.WindowsPath
    :return: compound_data - Data for what compound is in each well, based on excel files data.
    :rtype: dict
    """

    compound_data = {}

    # plate_name = file.name
    # # compound_data_org = {}
    # plate_name = plate_name.replace("-", "_").removesuffix(".xlsx")
    # compound_data[plate_name] = {}
    wb = load_workbook(filename=file)
    ws = wb.active
    for row, data in enumerate(ws):
        if row != 0:

            for col, cells in enumerate(data):
                if cells.value == None:
                    continue
                elif col == 0:
                    temp_well = cells.value
                elif col == 1:
                    try:
                        temp_compound = cells.value.casefold()
                    except AttributeError:
                        temp_compound = cells.value
                elif col == 2:
                    try:
                        # calculates volume in nL
                        temp_volume = float(cells.value) * 1000
                    except (ValueError, TypeError):
                        # return "Value_Error on volume column"
                        pass
                elif col == 3:
                    temp_plate_name = cells.value.casefold()

                elif col == 4:
                    temp_plate_type = cells.value.casefold()

                elif col == 5:
                    temp_compound_type = cells.value.casefold()

                    try:
                        compound_data[temp_compound_type]

                    except KeyError:
                        compound_data[temp_compound_type] = {"well_vol": {temp_well: temp_volume},
                                                             "compound": temp_compound,
                                                             "barcode": temp_plate_name,
                                                             "plate_type": temp_plate_type}
                    else:
                        compound_data[temp_compound_type]["well_vol"][temp_well] = temp_volume

    return compound_data


def insert_structure_NOT_USED_ATM(worksheet):
    """
    Inserts a structure for each row in the excel sheet. The smiles needs to be in the sheet, with the headline "smiles"
    :param worksheet: The worksheet where the picture should be insertet
    :return:
    """
    ws = worksheet
    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0:
            # Sets the first free column, for where to place the image
            col_index = len(row) + 1

            # Finds the headline for the smiles code, and set the smiles index to it.
            for headline_index, headline in enumerate(row):
                if headline.casefold() == "smiles":
                    smiles_index = headline_index

        # Grabs the smiles from the table
        smiles = row[smiles_index]

        mol = Chem.MolFromSmiles(smiles)
        temp_image = Draw.MolToImage(mol)

        # Save the PIL image as a temporary file
        temp_filename = tempfile.NamedTemporaryFile(suffix=".png", delete=False).name
        temp_image.save(temp_filename)

        # Create an Image object from the temporary file
        img = XLImage(temp_filename)

        # Calculate the image height
        image_height = img.height

        # Get the cell coordinate for the picture placement
        cell = ws.cell(row=row_index, column=col_index)

        # Add the image to the worksheet
        ws.add_image(img, cell.coordinate)

        # Calculate the required row height to fit the image
        required_row_height = int(image_height)

        # Set the row height
        ws.row_dimensions[row_index].height = required_row_height

        # Clean up the temporary file
        temp_image.close()


def get_source_layout(ex_file, source_layout):
    wb = load_workbook(filename=ex_file)
    ws = wb.active
    excel_dict = {}
    vol_unit = "uL"
    compound_to_group = {}
    group_counter = 1
    for row, data in enumerate(ws):
        if row == 0:

            for col, cells in enumerate(data):
                excel_dict[col] = cells.value.casefold()

        else:
            for col, cells in enumerate(data):

                if excel_dict[col] == "state":
                   temp_state = cells.value

                if temp_state != "sample":
                    source_layout[temp_state]["use"] = True

                    if excel_dict[col] == "compound":
                        source_layout[temp_state]["compound"] = cells.value

                    elif excel_dict[col] == "source_wells":
                        well_counter = source_layout[temp_state]["well_counter"]

                        try:
                            source_layout[temp_state]["source_wells"][well_counter]
                        except KeyError:
                            source_layout[temp_state]["source_wells"][well_counter] = {}

                        source_layout[temp_state]["source_wells"][well_counter]["well_id"] = cells.value

                    elif excel_dict[col] == "source_plate":
                        source_layout[temp_state]["source_wells"][well_counter]["plate"] = cells.value

                    elif excel_dict[col] == "vol":
                        temp_vol = f"{cells.value}{vol_unit}"
                        temp_vol, _, _, _ = unit_converter(temp_vol, old_unit_out=False, new_unit_out=False, as_list=True)
                        source_layout[temp_state]["source_wells"][well_counter]["vol"] = temp_vol

                else:
                    if excel_dict[col] == "compound":
                        temp_compound_id = cells.value

                        try:
                            compound_to_group[temp_compound_id]
                        except KeyError:
                            compound_to_group[temp_compound_id] = group_counter
                            group_counter += 1

                        try:
                            source_layout[temp_state][compound_to_group[temp_compound_id]]
                        except KeyError:
                            source_layout[temp_state][compound_to_group[temp_compound_id]] = {"compound_id": temp_compound_id,
                                                                                              "conc": {}}
                    elif excel_dict[col] == "concentration":
                        temp_conc = cells.value
                        temp_conc, _, _, _ = unit_converter(temp_conc, old_unit_out=False, new_unit_out=False, as_list=True)

                        try:
                            source_layout[temp_state][compound_to_group[temp_compound_id]]["conc"][temp_conc]
                        except KeyError:
                            source_layout[temp_state][compound_to_group[temp_compound_id]]["conc"][temp_conc] = {}

                    elif excel_dict[col] == "source_wells":
                        source_layout[temp_state][compound_to_group[temp_compound_id]]["conc"][temp_conc]["well"] = cells.value

                    elif excel_dict[col] == "source_plate":
                        source_layout[temp_state][compound_to_group[temp_compound_id]]["conc"][temp_conc]["plate"] = cells.value

                    elif excel_dict[col] == "vol":
                        temp_vol = f"{cells.value}{vol_unit}"
                        temp_vol, _, _, _ = unit_converter(temp_vol, old_unit_out=False, new_unit_out=False, as_list=True)
                        source_layout[temp_state][compound_to_group[temp_compound_id]]["conc"][temp_conc]["vol"] = temp_vol

            if temp_state != "sample":
                source_layout[temp_state]["well_counter"] += 1

    for states in source_layout:
        if states != "sample":
            if source_layout[states]["use"]:
                source_layout[states]["well_counter"] = 0

    return source_layout

if __name__ == "__main__":

    # file = "import/plate_dilution/The_form_we_send_out_to_ppl_with_a_name_that_make_sense.xlsx"
    # output = "save_plates"
    # #
    # import configparser
    # config = configparser.ConfigParser()
    # config.read("config.ini")
    #
    # print(config["plate_types_values"]["pp"].split(",")[0])
    # # plate_dilution_write_vol_well_amount(config, file)
    # table_data = [[4036482914, 4036482914, "Found"], [4036482915, 4036482915, "Found"], ["test_1", "None", "Not in DB"],
    #               ["test_2", "None", "Not in DB"]]
    # headings = ["Old Name", "new Name", "DB info"]
    #
    # file = purity_sample_layout_export(config, table_data, headings)
    # # purity_sample_layout_import(file)
    from pathlib import Path

    file = Path(r"C:\Users\phch\Desktop\more_data_files\alpha_SO\testing_new_worklist\Worklist_layout.xlsx")
    print(type(file))
    test = well_compound_list(file)
    print(test)
