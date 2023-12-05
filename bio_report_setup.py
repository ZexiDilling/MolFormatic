import tempfile

from openpyxl import Workbook
import numpy as np
from openpyxl.styles import Font
from rdkit import Chem
from rdkit.Chem import Draw
from openpyxl.drawing.image import Image as XLImage
import tempfile

from config_dictionary import bio_final_report_setup_fetch
from database_handler import DataBaseFunctions
from excel_handler import insert_structure
from import_openpyxl_handler import *
from bio_data_functions import bar_chart, frequency_writer


def _cal_writer_final_report(barcode, ws, all_data, init_row, init_col, report_output):
    """
    Writes the calculations in the combined report for all the plates

    :param barcode: The barcode for the plate
    :type barcode: str
    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param all_data: A dict over all plate date. all the analysed data will be added to this dict
    :type all_data: dict
    :param init_row: The first row to write to in the excel file
    :type init_row: int
    :param init_col: The first row to write to in the excel file
    :type init_col: int
    :param report_output: Gate dict for what information to write. is set in the settings
    :type report_output: dict
    :return: The overview report page in the final report.
    row_counter: The last row writen in
    :rtype: int
    """

    row_counter = init_row

    merge_cells_single_row(barcode, ws, row_counter, init_col, init_col + 2)
    try:
        ws.cell(column=init_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
    except ValueError:
        print("To many rows - _cal_writer_final_report")
    else:
        row_counter += 1
        for plate_analysed in all_data["calculations"]:
            # Removing other calculations than avg and stdev
            if plate_analysed != "other":
                # Checks to see if the overview of avg and stv should be included
                if report_output[plate_analysed]["overview"]:
                    # Writes the analysed method in, if the overview is set to true
                    merge_cells_single_row(plate_analysed, ws, row_counter, init_col, init_col + 2, True, "red_line")
                    row_counter += 1
                    for state in all_data["calculations"][plate_analysed]:
                        if state != "other":
                            if report_output[plate_analysed][state]:

                                ws.cell(column=init_col, row=row_counter, value=state).font = Font(b=True)
                                for calc in all_data["calculations"][plate_analysed][state]:
                                    # Writes avg and stdev including values
                                    ws.cell(column=init_col + 1, row=row_counter, value=calc)
                                    ws.cell(column=init_col + 2, row=row_counter,
                                            value=all_data["calculations"][plate_analysed][state][calc])
                                    row_counter += 1
            else:
                if report_output["z_prime"]:
                    ws.cell(column=init_col, row=row_counter, value="z-Prime").font = Font(b=True)
                    try:
                        ws.cell(column=init_col + 2, row=row_counter,
                                value=all_data["calculations"][plate_analysed]["z_prime"])
                    except KeyError:
                        ws.cell(column=init_col + 2, row=row_counter,
                                value="Z-Prime is not calculated for the plates")
                    row_counter += 1
                row_counter += 1
    return row_counter


def _well_writer_final_report(ws, hits, final_report_setup, init_row):
    """
    Writes all the wells in a list, in the excel file, for the values that are within the predetermined values.


    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param hits: A dict over all the values that are "hits" and needs to be added
    :type hits: dict
    :param final_report_setup: Gate for what data is added to the report. Is setting in the settings
    :type final_report_setup: dict
    :param init_row: The first row to write to.
    :type init_row: int
    :return: A list of wells in the excel sheet.
    """
    indent_col = 1
    row_counter = init_row

    for barcode in hits:
        # Writes headline for data inserts to see where the data is coming from
        ws.cell(column=indent_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
        row_counter += 1

        for method in hits[barcode]:
            if final_report_setup["methods"][method]:
                # writes method
                ws.cell(column=indent_col, row=row_counter, value=method).font = Font(b=True)
                row_counter += 1
                for split in hits[barcode][method]:
                    ws.cell(column=indent_col, row=row_counter, value=split).font = Font(b=True)
                    ws.cell(column=indent_col + 1, row=row_counter,
                            value=final_report_setup["pora_threshold"][split]["min"]).font = \
                        Font(underline="single")
                    ws.cell(column=indent_col + 2, row=row_counter,
                            value=final_report_setup["pora_threshold"][split]["max"]).font = \
                        Font(underline="single")
                    row_counter += 1
                    for well in hits[barcode][method][split]:
                        ws.cell(column=indent_col + 1, row=row_counter, value=well)
                        ws.cell(column=indent_col + 2, row=row_counter,
                                value=hits[barcode][method][split][well])
                        row_counter += 1
        indent_col += 4
        row_counter = init_row


def _get_data(dbf, all_plate_data, final_report_setup, plate_to_layout, bio_sample_dict=None):
    """
    Grabs data that are needed for the different output sheets in the excel file.

    :param all_plate_data: The data for all the plates.
    :type all_plate_data: dict
    :param final_report_setup: The settings for the final report. is set in the settings.
    :type final_report_setup: dict
    :param plate_to_layout: a dicts for the plate_layout
    :type plate_to_layout: dict
    :return:
        - temp_hits: All the hits that are within the values
        - data_calc_dict: A dicts over all the calculations and the values
        - plate_counter: The amount of plates that are being analysed
        - all_states: The states that are being used for the analysis
        - all_methods: The methods that are being used for the analysis
        - hit_data: a dict of sample ID's with there values
        - empty_wells: a dict over wells that are empty according to the worklist
    :rtype:
        - dict
        - dict
        - int
        - dict
        - dict
        - dict
        - dict
    """
    data_calc_dict = {}
    temp_hits = {}
    empty_wells = {}

    plate_counter = 0
    all_states = []
    all_methods = []
    freq_data = {"all_data": []}

    hit_data = {}
    data_amount = len(all_plate_data)
    for plate_index, barcode in enumerate(all_plate_data):
        empty_wells[barcode] = []
        plate_counter += 1
        temp_hits[barcode] = {}
        data_calc_dict[barcode] = {}
        for method in all_plate_data[barcode]["plates"]:
            if method != "other":
                if method not in all_methods:
                    all_methods.append(method)
            if final_report_setup["methods"][method]:
                temp_hits[barcode][method] = {}
                freq_data[barcode] = []
                for thresholds in final_report_setup["pora_threshold"]:
                    temp_hits[barcode][method][thresholds] = {}
                temp_plate_layout = eval(dbf.find_data_single_lookup("plate_layout", plate_to_layout[barcode],
                                                                     "layout_name")[0][5])
                for rows in temp_plate_layout:
                    if temp_plate_layout[rows]["state"] == "sample":
                        well = temp_plate_layout[rows]["well_id"]
                        temp_well_value = all_plate_data[barcode]["plates"][method]["wells"][well]
                        freq_data[barcode].append(temp_well_value)
                        freq_data["all_data"].append(temp_well_value)

                        if bio_sample_dict:
                            try:
                                compound_id = bio_sample_dict[barcode][well]["compound_id"]
                            except KeyError:
                                empty_wells[barcode].append(well)
                                print(f"Well {well} on {barcode} have no transferee in the worklist.")
                                continue
                            else:
                                if not compound_id:
                                    table = "compound_mp"
                                    well_headline = "mp_well"
                                    temp_well_data = bio_sample_dict[barcode][well]["source_well"]
                                    plate_headline = "mp_barcode"
                                    temp_plate_data = bio_sample_dict[barcode][well]["source_plate"]
                                    temp_row_data = dbf.find_data_double_lookup(table, temp_well_data, temp_plate_data,
                                                                                well_headline, plate_headline)
                                    try:
                                        temp_row_data[0][3]
                                    except IndexError:
                                        print(f"Missing data for MP: {temp_plate_data} well: {temp_well_data}")
                                        print(bio_sample_dict[barcode][well])
                                    else:
                                        compound_id = temp_row_data[0][3]
                                        bio_sample_dict[barcode][well]["compound_id"] = compound_id

                                hit_data[compound_id] = temp_well_value

                        for split in final_report_setup["pora_threshold"]:

                            # Check if the specific threshold is included in the report
                            if final_report_setup["pora_threshold"][split]:
                                if float(final_report_setup["pora_threshold"][split]["min"]) < float(temp_well_value) < \
                                        float(final_report_setup["pora_threshold"][split]["max"]):
                                    temp_hits[barcode][method][split][well] = temp_well_value

        for method in all_plate_data[barcode]["calculations"]:
            data_calc_dict[barcode][method] = {}
            if method != "other":
                for state in all_plate_data[barcode]["calculations"][method]:
                    if state not in all_states:
                        all_states.append(state)

                    data_calc_dict[barcode][method][state] = {}
                    for calc in all_plate_data[barcode]["calculations"][method][state]:
                        data_calc_dict[barcode][method][state][calc] = \
                            all_plate_data[barcode]["calculations"][method][state][calc]

            else:
                for other_calc in all_plate_data[barcode]["calculations"][method]:
                    data_calc_dict[barcode][method][other_calc] = \
                        all_plate_data[barcode]["calculations"][method][other_calc]

        print(f"plate {plate_index + 1} / {data_amount} have been calculated - {barcode}")

    return temp_hits, data_calc_dict, plate_counter, all_states, all_methods, freq_data, hit_data, empty_wells


def _data_writer(ws_matrix, ws_list, data_calc_dict, state, plate_counter, all_methods, use_list, use_max_min):
    """
    Writes all the data, and handles the flow of the data to witch sheet different things are writen in.

    :param ws_matrix: The Worksheet for the data that goes in Matrix formate for calculations
    :type ws_matrix: openpyxl.worksheet.worksheet.Worksheet
    :param ws_list: The worksheet that list calculations and the min/max values for the different once
    :type ws_list: openpyxl.worksheet.worksheet.Worksheet
    :param data_calc_dict: All the data in a dict formate
    :type data_calc_dict: dict
    :param state: What state the data is for (samples, minimum, maximum, blank...)
    :type state: str
    :param plate_counter: The amount of plates that are being analysed
    :type plate_counter: int
    :param all_methods: A list of all the methods
    :type all_methods: list
    :param use_list: If the list data should be added to the report. Is set in the settings.
    :type use_list: bool
    :param use_max_min: If the min_max data should be added to the report. Is set in the settings.
    :type use_max_min: bool
    :return: Values written into the excel sheet for the Matrix, and the list and or min_max depending on settings
    """
    init_row = 4
    init_col = 3
    spacer = 4
    list_spacer_clm = 6

    col_stdev = init_col + plate_counter + spacer
    col_counter = init_col + 1
    row_counter = init_row + 1
    col_stdev_counter = col_stdev + 1
    row_offset = init_row

    list_clm = init_col - 1
    list_row = init_row

    list_row_minmax = init_row

    for index_m, method in enumerate(all_methods):
        temp_avg_dict = {}
        temp_stdev_dict = {}
        mw_col = col_counter
        mw_row = row_counter
        mw_col_stdev = col_stdev_counter

        for barcodes in data_calc_dict:
            # Writes Plate names in row and clm for avg
            ws_matrix.cell(column=init_col - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws_matrix.cell(column=col_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            # Writes Plate names in row and clm for stdev
            ws_matrix.cell(column=col_stdev - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws_matrix.cell(column=col_stdev_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            for index_method, _ in enumerate(data_calc_dict[barcodes]):

                if index_method == 0:
                    # Writes method for avg
                    ws_matrix.cell(column=init_col, row=row_offset - 1, value=method).font = Font(b=True)
                    # Writes method for stdev
                    ws_matrix.cell(column=col_stdev, row=row_offset - 1, value=method).font = Font(b=True)
                    if method != "other":
                        for calc in data_calc_dict[barcodes][method][state]:
                            temp_value = data_calc_dict[barcodes][method][state][calc]
                            # gets avg values
                            if calc == "avg":
                                ws_matrix.cell(column=init_col, row=row_offset, value=calc).font = Font(b=True)
                                ws_matrix.cell(column=init_col, row=row_counter, value=temp_value)
                                ws_matrix.cell(column=col_counter, row=row_offset, value=temp_value)
                                temp_avg_dict[barcodes] = temp_value
                            elif calc == "stdev":
                                ws_matrix.cell(column=col_stdev, row=row_offset, value=calc).font = Font(b=True)
                                ws_matrix.cell(column=col_stdev, row=row_counter, value=temp_value)
                                ws_matrix.cell(column=col_stdev_counter, row=row_offset, value=temp_value)
                                temp_stdev_dict[barcodes] = temp_value
            # Sets offset for next loop, for writing headlines the right place
            col_counter += 1
            row_counter += 1
            col_stdev_counter += 1

        # calculate the % difference between avg for each plate
        _matrix_calculator(ws_matrix, mw_row, mw_col, temp_avg_dict)
        # calculate the % difference between stdev for each plate
        _matrix_calculator(ws_matrix, mw_row, mw_col_stdev, temp_stdev_dict)

        # sortes the dict by size
        temp_avg_dict = _sort_dict(temp_avg_dict)
        temp_stdev_dict = _sort_dict(temp_stdev_dict)

        # writes list of all avg for the different methods
        if use_list:
            _writes_list_of_values(ws_list, list_row, list_clm, temp_avg_dict, "avg", method)
            _writes_list_of_values(ws_list, list_row, list_clm + 3, temp_stdev_dict, "stdev", method)

        # Calculate how much space the list takes
        list_clm = (list_spacer_clm * (len(all_methods))) + 2

        # writes list of all avg for the different methods
        if use_max_min:
            _write_min_max_values(ws_list, list_row_minmax, list_clm, temp_avg_dict, "avg", method)
            _write_min_max_values(ws_list, list_row_minmax, list_clm + 4, temp_stdev_dict, "stdev", method)

        # makes sure that next loop is writen below the first method. One method per row, with avg and stdev for each.
        col_stdev = init_col + plate_counter + spacer
        col_counter = init_col + 1
        row_counter += spacer
        col_stdev_counter = col_stdev + 1
        row_offset += (plate_counter + spacer)
        list_row_minmax += 5
        list_clm = init_col - 1 + (list_spacer_clm * (index_m + 1))


def _matrix_calculator(ws, row, col, temp_data_dict):
    """
    Calculates and writes the values for the Matrix

    :param ws: The Worksheet for the data that goes in Matrix formate for calculations
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row for the locations of the matrix
    :type row: int
    :param col: The col for the locations of the matrix
    :type col: int
    :param temp_data_dict: The data for the matrix
    :type temp_data_dict: dict
    :return: Written a matrix in the final report for the excel file
    """

    for index_x, value_x in enumerate(temp_data_dict):
        for index_y, value_y in enumerate(temp_data_dict):
            try:
                # Divide the value of `value_x` by the value of `value_y` and multiply by 100
                temp_value = (float(temp_data_dict[value_x]) / float(temp_data_dict[value_y])) * 100
            except (ZeroDivisionError, TypeError):
                # Handle division by zero and TypeError (when a value can't be converted to float)
                temp_value = None
            # Write the calculated value in the cell (column = `col + index_x`, row = `row + index_y`)
            ws.cell(column=col + index_x, row=row + index_y, value=temp_value)


def _sort_dict(temp_data_dict):
    """
    This sorts the dict from lowest values to highest

    :param temp_data_dict: The data for the hit wells.
    :type temp_data_dict: dict
    :return: a sorted dict
    :rtype: dict
    """
    try:
        return {key: value for key, value in sorted(temp_data_dict.items(), key=lambda item: item[1])}
    except TypeError:
        return temp_data_dict


def _writes_list_of_values(ws, row, col, temp_data_dict, item_name, method=None):
    """
    Writes the hit values in a list.
    NEEDS TO ADD COMPOUND DATA TO THIS AT SOME POINT!!!

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row where the data is start being written
    :type row: int
    :param col: The col where the data is start being written
    :type col: int
    :param temp_data_dict: The data that needs to be writen. A dict of wells and their values
    :type temp_data_dict: dict
    :param item_name: The name for the data type. avg or stdev
    :type item_name: str
    :param method: What methods the data is from
    :type method: str
    :return: Data written in the excel file.
    """

    # writes method
    if method:
        merge_cells_single_row(method, ws, row - 2, col, col + 1, True, "red_line")

    # writes headline for the list
    ws.cell(column=col, row=row - 1, value="Barcode").font = Font(b=True)
    ws.cell(column=col + 1, row=row - 1, value=item_name).font = Font(b=True)
    # writes list of values and barcode / plate name
    row_counter = 0
    for index, values in enumerate(temp_data_dict):
        ws.cell(column=col, row=row + index, value=values)
        ws.cell(column=col + 1, row=row + index, value=temp_data_dict[values])
        row_counter += 1

    table_name = f"{item_name}_{method}"
    start_row = row - 1
    end_row = start_row + row_counter
    ws.add_table(table_purple(ws, table_name, start_row, col, end_row, col + 1))


def _write_min_max_values(ws, row, col, data_dict, item_name, method=None):
    """
    Writes the min and max values in the list sheet

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param row: The row where the data is start being written
    :type row: int
    :param col: The col where the data is start being written
    :type col: int
    :param item_name: The name for the data type. avg or stdev
    :type item_name: str
    :param method: What methods the data is from
    :type method: str
    :return: Data written in the excel file.
    """

    # writes method
    if method:
        merge_cells_single_row(method, ws, row - 2, col, col + 2, True, "red_line")

    # writes headlines
    ws.cell(column=col + 1, row=row - 1, value="Barcode").font = Font(b=True)
    ws.cell(column=col + 2, row=row - 1, value=item_name).font = Font(b=True)
    ws.cell(column=col, row=row, value="Maximum").font = Font(b=True)
    ws.cell(column=col, row=row + 1, value="Minimum").font = Font(b=True)

    # removes None values:
    temp_data_dict = {}
    for keys in data_dict:
        if data_dict[keys]:
            temp_data_dict[keys] = data_dict[keys]

    # writes max and barcode / plate name
    temp_dict_max = max(temp_data_dict, key=temp_data_dict.get)

    ws.cell(column=col + 1, row=row, value=temp_dict_max)
    ws.cell(column=col + 2, row=row, value=temp_data_dict[temp_dict_max])

    # writes max and barcode / plate name
    temp_dict_min = min(temp_data_dict, key=temp_data_dict.get)
    ws.cell(column=col + 1, row=row + 1, value=temp_dict_min)
    ws.cell(column=col + 2, row=row + 1, value=temp_data_dict[temp_dict_min])


def _z_prime(ws, data_calc_dict, use_list, use_max_min):
    """
    Writes the Z-Prime report page in the final report

    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param data_calc_dict: A dict over all the Z-Prime values
    :type data_calc_dict: dict
    :param use_list: If the list data should be added to the report. Is set in the settings.
    :type use_list: bool
    :param use_max_min: If the min_max data should be added to the report. Is set in the settings.
    :type use_max_min: bool
    :return: All the data for the Z-Prime in its own sheet in the final report excel file.
    """

    init_row = 2
    init_col = 2
    spacer = 1

    matrix_col = init_col + 8
    matrix_row = init_row + spacer

    col_counter = matrix_col + spacer
    row_counter = matrix_row + spacer

    z_prime_dict = {}

    for barcodes in data_calc_dict:
        # Writes Plate names
        ws.cell(column=matrix_col - 1, row=row_counter, value=barcodes).font = Font(b=True)
        ws.cell(column=col_counter, row=matrix_row - 1, value=barcodes).font = Font(b=True)
        # Writes values for Z-Prime
        try:
            z_prime = data_calc_dict[barcodes]["other_data"]["z_prime"]
        except KeyError:
            z_prime = data_calc_dict[barcodes]["other"]["z_prime"]

        ws.cell(column=matrix_col, row=row_counter, value=z_prime)
        ws.cell(column=col_counter, row=matrix_row, value=z_prime)
        col_counter += 1
        row_counter += 1
        # z_prime_list.append(z_prime)
        z_prime_dict[barcodes] = z_prime

    col_counter = init_col + 1
    row_counter = init_row + 1

    # _matrix_calculator(ws, row_counter, col_counter, z_prime_list)
    _matrix_calculator(ws, matrix_row + 1, matrix_col + 1, z_prime_dict)

    z_prime_dict = _sort_dict(z_prime_dict)
    if use_list:
        _writes_list_of_values(ws, row_counter, init_col, z_prime_dict, "z_prime")
    if use_max_min:
        _write_min_max_values(ws, row_counter, col_counter + 2, z_prime_dict, "z_prime")


def _fetch_smiles_data(config, sample_id):
    """
    Get smiles data from the database based on the ID of each compound
    :param config: The config handler, with all the default information in the config file.
    :type config: configparser.ConfigParser
    :param sample_id: The compound id for the sample
    :type sample_id: int
    :return: Smiles
    :rtype: str
    """
    dbf = DataBaseFunctions(config)
    table_name = "compound_main"
    headline_name = "compound_id"
    row_data = dbf.find_data_single_lookup(table_name, sample_id, headline_name)
    smiles = row_data[0][2]

    return smiles


def _write_hits(config, ws, hit_data, threshold, hit_amount, include_hit, include_smiles, include_structure,
                bio_sample_dict):
    """
    Writes the hits into a worksheet in the workbook, based either on a threshold or amount of hits
    :param config: The config handler, with all the default information in the config file.
    :type config: configparser.ConfigParser
    :param ws: The Worksheet for the data, for the excel file
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param hit_data: The compound id, smiles and hit-amount (how much the have supressed stuff)
    :type hit_data: dict
    :param threshold: If the report should include hits, then this is the threshold for what is consideret as hits
        It can either use threshold or hit_amount
    :type threshold: int or None
    :param hit_amount: If the report should include hits, then how many hits should be included
        It can either use threshold or hit_amount
    :type hit_amount: int or None
    :param include_smiles: If there should be id and smiles on the hits
    :type include_smiles: bool
    :param bio_sample_dict: A dict for ID's and smiles for each well on each plate
    :type bio_sample_dict: dict or None
    :return:
    """
    # headlines for the sheet

    headline_dict = {
        "ID": {"use": include_hit, "clm_index": 0},
        "Score": {"use": True, "clm_index": 0},
        "Smiles": {"use": include_smiles, "clm_index": 0},
        "Plate": {"use": True, "clm_index": 0},
        "Well": {"use": True, "clm_index": 0},
        "Structure": {"use": include_structure, "clm_index": 0},
    }
    headlines = []
    headline_counter = 0
    for headline in headline_dict:
        if headline[headline]["use"]:
            headline_counter += 1
            headline.append(headline)
            headline[headline]["clm_index"] = headline_counter

    # Sorting the hit data:
    sorted_hits = sorted(hit_data.items(), key=lambda x: x[1])
    sorted_hits_dict = dict(sorted_hits)

    # Write list depending on amount of hits or the threshold
    hit_list = []
    for counter, compound in enumerate(sorted_hits_dict):
        if hit_amount:
            if hit_amount > counter:
                hit_list.append(compound)
        else:
            if sorted_hits_dict[compound] <= threshold:
                hit_list.append(compound)

    # Gets all the data for the execel sheet in one dict
    hit_data_dict = {}
    for plates in bio_sample_dict:
        for wells in bio_sample_dict[plates]:
            temp_compound = bio_sample_dict[plates][wells]["compound_id"]
            if temp_compound in hit_list:
                if include_smiles:
                    smiles = _fetch_smiles_data(config, temp_compound)
                    bio_sample_dict[plates][wells]["smiles"] = smiles
                else:
                    smiles = ""

                hit_data_dict[temp_compound] = {
                    "score": sorted_hits_dict[temp_compound],
                    "smiles": smiles,
                    "plate": plates,
                    "well": wells
                }

    # Initail excel set-up
    row = 1
    col = 1

    # Write the headlines:
    for counter, headline in enumerate(headlines):
        ws.cell(column=col + counter, row=row, value=headline)

    row += 1

    # Writes the data

    for row_index, compound_id in enumerate(hit_list):

        if include_hit:
            ws.cell(column=headline_dict["place_holder"]["clm_index"], row=row + row_index,
                    value=compound_id)

        ws.cell(column=headline_dict["place_holder"]["clm_index"], row=row + row_index,
                value=hit_data_dict[compound_id]["score"])

        if include_smiles:
            ws.cell(column=headline_dict["place_holder"]["clm_index"], row=row + row_index,
                    value=hit_data_dict[compound_id]["smiles"])

        ws.cell(column=headline_dict["place_holder"]["clm_index"], row=row + row_index,
                value=hit_data_dict[compound_id]["plate"])
        ws.cell(column=headline_dict["place_holder"]["clm_index"], row=row + row_index,
                value=hit_data_dict[compound_id]["well"])

        if include_structure:
            # Grabs the smiles from the table
            smiles = hit_data_dict[compound_id]["smiles"]

            mol = Chem.MolFromSmiles(smiles)
            temp_image = Draw.MolToImage(mol)

            # Save the PIL image as a temporary file
            temp_filename = tempfile.NamedTemporaryFile(suffix=".png", delete=False).name
            temp_image.save(temp_filename)

            # Create an Image object from the temporary file
            img = XLImage(temp_filename)

            # Calculate the image height
            image_height = img.height

            # Get the cell coordinate for the picture placement
            cell = ws.cell(row=row + row_index, column=headline_dict["Structure"]["clm_index"])

            # Add the image to the worksheet
            ws.add_image(img, cell.coordinate)

            # Calculate the required row height to fit the image
            required_row_height = int(image_height)

            # Set the row height
            ws.row_dimensions[row + row_index].height = required_row_height

            # Clean up the temporary file
            temp_image.close()


def bio_final_report_controller(dbf, config, analyse_method, all_plate_data, output_file, include_hits, threshold,
                                hit_amount, include_smiles, bio_sample_dict, plate_to_layout, include_structure):
    """
    The controller for the flow of data, that writes the final report in an excel file.
    :param config: The config handler, with all the default information in the config file.
    :type config: configparser.ConfigParser
    :param analyse_method: What analyse method is being used.
    :type analyse_method: str
    :param all_plate_data: All the data for all the plates, including calculations and well states
    :type all_plate_data: dict
    :param output_file: The name and path for the final report
    :type output_file: str
    :param include_hits: If the final reports should include hits
    :type include_hits: bool
    :param threshold: If the report should include hits, then this is the threshold for what is consideret as hits
        It can either use threshold or hit_amount
    :type threshold: int or None
    :param hit_amount: If the report should include hits, then how many hits should be included
        It can either use threshold or hit_amount
    :type hit_amount: int or None
    :param include_smiles: If there should be id and smiles on the hits
    :type include_smiles: bool
    :param bio_sample_dict: A dict for ID's and smiles for each well on each plate
    :type bio_sample_dict: dict or None
    :param plate_to_layout: a dicts for the plate_layout
    :type plate_to_layout: dict
    :param include_structure: boolen to determen if the file report should include png of the structure
    :type include_structure: bool
    :return: An excel file ready to be presented... or something..
    """
    wb = Workbook()
    ws_report = wb.active
    ws_report.title = "Full report"
    ws_well_info = wb.create_sheet("Well Info")
    seperator = 0
    init_row = 2
    init_col = 2
    row = init_row
    col = init_col
    final_report_setup = bio_final_report_setup_fetch(config)
    # calc overview:
    # for plates in all_plate_data:
    #     print(plates)

    for index, barcode in enumerate(all_plate_data):
        row_counter = _cal_writer_final_report(barcode, ws_report, all_plate_data[barcode], row, col,
                                               final_report_setup["calc"])
        # Writes 5 plates horizontal, before changing rows.
        col += 4
        if index % 5 == 0 and index > 0:
            # Sets the spacing between each set of data
            if seperator == 0:
                seperator = row_counter
                print(seperator)
            row += seperator
            col = init_col
    print("Start plate Calculations")

    print("Start getting data")
    # gets data:
    temp_hits, data_calc_dict, plate_counter, all_states, all_methods, all_freq_data, hit_data, empty_wells = \
        _get_data(dbf, all_plate_data, final_report_setup, plate_to_layout, bio_sample_dict)

    print("Start writing well-data to the file")
    # write well data
    _well_writer_final_report(ws_well_info, temp_hits, final_report_setup, init_row)

    # writes Matrix of data:
    print("Starts writing the Matrix in the file")
    for states in all_states:
        if states != "other":
            if final_report_setup["data"][states]["matrix"]:
                _data_writer(ws_creator(wb, states, "Matrix"), ws_creator(wb, states, "Lists"), data_calc_dict, states,
                             plate_counter, all_methods, final_report_setup["data"][states]["list"],
                             final_report_setup["data"][states]["max_min"])

    print("Starts writing Z-prime into the file")
    # writes Z-prime
    if final_report_setup["data"]["z_prime"]["matrix"]:
        ws_z_prime = wb.create_sheet("Z-Prime")
        _z_prime(ws_z_prime, data_calc_dict, final_report_setup["data"]["z_prime"]["list"],
                 final_report_setup["data"]["z_prime"]["max_min"])

    print("Starts making the histogram")
    # histograms
    bin_min = 0
    bin_max = 150
    bin_width = 5
    include_outliers = True
    free_col = 1
    initial_row = 1
    ws_histograms = wb.create_sheet("histogram")
    bar_row = initial_row + 34
    bar_col = 1

    for freq_data in all_freq_data:
        headline = freq_data
        data_set = all_freq_data[freq_data]
        free_col, data_location, category_location = \
            frequency_writer(ws_histograms, headline, data_set, free_col, initial_row, bin_min, bin_max, bin_width,
                             include_outliers)

        bar_chart(ws_histograms, headline, bar_col, bar_row, data_location, category_location)
        free_col += 1
        bar_col += 10
        # sets amount of charts per row. 1 chart is aprox 10 cells long
        if bar_col >= 30:
            bar_col = 1
            bar_row += 15

    # writes hits based on treshedhold, either amount or score.
    if include_hits or include_smiles:
        print("write hits")
        ws_hits = wb.create_sheet("Hit List")
        _write_hits(config, ws_hits, hit_data, threshold, hit_amount, include_hits, include_smiles, include_structure,
                    bio_sample_dict)

    if include_smiles:
        print("inserting png's")
        insert_structure(ws_hits)

    print(f"The final report is done, here is the output file: {output_file}")
    wb.save(output_file)


if __name__ == "__main__":
    ...
