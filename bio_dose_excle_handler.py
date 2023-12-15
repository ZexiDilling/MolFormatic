from math import log10

import openpyxl
from numpy import array
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from statistics import mean

from bio_report_setup import _fetch_smiles_data
from file_type_handler_excel import insert_structure


def _write_dose_readings(ws_readings, ws_initial_row, ws_reading_clm, temp_data, add_dose, add_state_data, replicates):
    all_rows = []
    ws_reading_row = ws_initial_row
    if replicates:
        temp_sample_counter = 1
        temp_rep_counter = 0
        data_placement = {"dose": {"clm": [], "row": []},
                          "sheet": ws_readings}
    else:
        data_placement = {"dose": {"clm": [], "row": []},
                          "samples": {"clm": [], "row": []},
                          "sheet": ws_readings}

    if add_dose:
        temp_dose_clm = ws_reading_clm
        temp_dose_row = ws_reading_clm
        ws_reading_clm += 1

    if add_state_data:
        for states in temp_data["state_data"]:
            ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=states)
            for well_value in temp_data["state_data"][states]:
                ws_reading_row += 1
                ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=well_value)
            ws_reading_clm += 1
            all_rows.append(ws_reading_row)
            ws_reading_row = 1

    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data":
            temp_sample_name = samples.split("_")
            temp_sample_name = f"{temp_sample_name[0]}_{temp_sample_name[1]}_{temp_sample_name[2]}_{temp_sample_name[3]}"
            ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=samples)
            ws_reading_row += 1
            for reading_data in temp_data[samples]["reading"]["raw"]:

                ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=reading_data)

                if replicates:

                    data_placement_name = f"samples_{temp_sample_name}"
                    try:
                        data_placement[data_placement_name]
                    except KeyError:
                        data_placement[data_placement_name] = {"clm": [], "row": []}

                else:
                    data_placement_name = "samples"

                data_placement[data_placement_name]["row"].append(ws_reading_row)
                data_placement[data_placement_name]["clm"].append(ws_reading_clm)

                ws_reading_row += 1

            if add_dose:
                temp_name = f"Dose({temp_data[samples]['dose']['unit']})"
                ws_readings.cell(column=temp_dose_clm, row=temp_dose_row, value=temp_name)
                temp_dose_row += 1

                for reading_data in temp_data[samples]["dose"]["raw"]:
                    ws_readings.cell(column=temp_dose_clm, row=temp_dose_row, value=reading_data)

                    data_placement["dose"]["clm"].append(1)
                    data_placement["dose"]["row"].append(temp_dose_row)
                    temp_dose_row += 1
                add_dose = False
            if replicates:
                temp_rep_counter += 1
                if temp_rep_counter == replicates:
                    temp_sample_counter += 1
                    temp_rep_counter = 0
            ws_reading_clm += 1
            all_rows.append(ws_reading_row)
            ws_reading_row = ws_initial_row

    max_row = max(all_rows)
    return data_placement, max_row
    # reading - raw
    # reading - normalized
    # reading - min
    # reading - max
    # reading - fitted_normalized
    # reading - fitted
    # dose - raw
    # dose - normalized
    # dose - min
    # dose - max
    # dose - fitted_normalized
    # dose - fitted


def _write_dose_log(ws_log, temp_data, replicates, ws_initial_row):
    starting_row = ws_initial_row + 2
    temp_dose_row = ws_log_row = starting_row
    temp_dose_clm = ws_log_clm = 1
    all_rows = []
    add_dose = True
    if replicates:
        temp_sample_counter = 1
        temp_rep_counter = 0
        data_placement = {"dose": {"clm": [], "row": []},
                          "sheet": ws_log}
    else:
        data_placement = {"dose": {"clm": [], "row": []},
                          "samples": {"clm": [], "row": []},
                          "sheet": ws_log}

    ws_log_clm += 1

    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data":
            temp_sample_name = samples.split("_")
            temp_sample_name = f"{temp_sample_name[0]}_{temp_sample_name[1]}_{temp_sample_name[2]}_{temp_sample_name[3]}"
            ws_log.cell(column=ws_log_clm, row=ws_log_row, value=samples)
            ws_log_row += 1
            for reading_data in temp_data[samples]["reading"]["raw"]:

                ws_log.cell(column=ws_log_clm, row=ws_log_row, value=reading_data)

                if replicates:

                    data_placement_name = f"samples_{temp_sample_name}"
                    try:
                        data_placement[data_placement_name]
                    except KeyError:
                        data_placement[data_placement_name] = {"clm": [], "row": []}

                else:
                    data_placement_name = "samples"

                data_placement[data_placement_name]["row"].append(ws_log_row)
                data_placement[data_placement_name]["clm"].append(ws_log_clm)

                ws_log_row += 1

            if add_dose:
                temp_name = f"Dose Log10"
                ws_log.cell(column=temp_dose_clm, row=temp_dose_row, value=temp_name)

                # Changing unite to avoid having negative values on the graph
                temp_dose_row += 1
                temp_counter = 1
                temp_check = temp_data[samples]["dose"]["min"]
                while temp_check < 1:
                    temp_check *= temp_counter
                    if temp_counter == 1:
                        temp_counter = 1000
                    else:
                        temp_counter += 1000

                for reading_data in temp_data[samples]["dose"]["raw"]:
                    temp_value = reading_data * temp_counter
                    value = log10(temp_value)
                    ws_log.cell(column=temp_dose_clm, row=temp_dose_row, value=value)

                    data_placement["dose"]["clm"].append(1)
                    data_placement["dose"]["row"].append(temp_dose_row)
                    temp_dose_row += 1
                add_dose = False
            if replicates:
                temp_rep_counter += 1
                if temp_rep_counter == replicates:
                    temp_sample_counter += 1
                    temp_rep_counter = 0
            ws_log_clm += 1
            all_rows.append(ws_log_row)
            ws_log_row = starting_row
    max_row = max(all_rows)
    return data_placement, max_row


def __grab_avg_data(temp_data, replicates):
    avg_data = {}
    sample_list = []
    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data":
            if sample_index % replicates == 0:
                temp_sample_name = samples.removesuffix("_1")

                avg_data[temp_sample_name] = {"avg": []}
                sample_list.append(temp_sample_name)

            for reading_counter, reading_data in enumerate(temp_data[samples]["reading"]["raw"]):
                try:
                    avg_data[temp_sample_name][reading_counter]
                except KeyError:
                    avg_data[temp_sample_name][reading_counter] = [reading_data]
                else:
                    avg_data[temp_sample_name][reading_counter].append(reading_data)

    for sample in avg_data:
        for counter in avg_data[sample]:
            if counter != "avg":
                temp_avg = mean(avg_data[sample][counter])
                avg_data[sample]["avg"].append(temp_avg)
    return avg_data


def _write_avg_data(ws_avg, temp_data, replicates, ws_initial_row):
    starting_row = ws_initial_row + 2

    temp_dose_row = ws_avg_row = starting_row

    all_row_value = []
    avg_data = __grab_avg_data(temp_data, replicates)

    data_placement = {"dose": {"clm": [], "row": []},
                      "sheet": ws_avg}

    temp_dose_clm = ws_avg_clm = 1

    temp_dose_row += 1
    for sample_index, samples in enumerate(temp_data):
        if sample_index == 0:
            temp_name = f"Dose Log10"
            ws_avg.cell(column=temp_dose_clm, row=temp_dose_row, value=temp_name)

            # Changing unite to avoid having negative values on the graph
            temp_dose_row += 1
            temp_counter = 1
            temp_check = temp_data[samples]["dose"]["min"]
            while temp_check < 1:
                temp_check *= temp_counter
                if temp_counter == 1:
                    temp_counter = 1000
                else:
                    temp_counter += 1000

            for reading_data in temp_data[samples]["dose"]["raw"]:
                temp_value = reading_data * temp_counter
                value = log10(temp_value)
                ws_avg.cell(column=temp_dose_clm, row=temp_dose_row, value=value)

                data_placement["dose"]["clm"].append(1)
                data_placement["dose"]["row"].append(temp_dose_row)
                temp_dose_row += 1
        all_row_value.append(temp_dose_row)

    ws_avg_clm += 1
    for sample_index, samples in enumerate(avg_data):
        ws_avg.cell(column=ws_avg_clm, row=ws_avg_row, value=samples)
        # data_placement[samples] = {"raw_clm": [], "raw_row": [], "avg_clm": [], "avg_row": [], "header": [ws_avg_clm, ws_avg_row]}
        data_placement[samples] = {"avg_clm": [], "avg_row": [], "raw": {}, "header": [ws_avg_clm, ws_avg_row]}

        ws_avg_row += 1
        for counter in avg_data[samples]:

            ws_avg.cell(column=ws_avg_clm, row=ws_avg_row, value=counter)
            ws_avg_row += 1

            for data in avg_data[samples][counter]:
                ws_avg.cell(column=ws_avg_clm, row=ws_avg_row, value=data)
                if counter == "avg":
                    data_placement[samples]["avg_clm"].append(ws_avg_clm)
                    data_placement[samples]["avg_row"].append(ws_avg_row)
                    ws_avg_row += 1

                else:
                    try:
                        data_placement[samples]["raw"][counter]
                    except KeyError:
                        data_placement[samples]["raw"][counter] = {"clm": [ws_avg_clm], "row": [ws_avg_row]}
                    else:
                        data_placement[samples]["raw"][counter]["clm"].append(ws_avg_clm)
                        data_placement[samples]["raw"][counter]["row"].append(ws_avg_row)
                    ws_avg_row += 1
            ws_avg_clm += 1
            all_row_value.append(ws_avg_row)
            ws_avg_row = starting_row + 1
        ws_avg_row = starting_row

    max_row = max(all_row_value)
    return data_placement, max_row


def _write_dose_data(ws_data, ws_data_row, ws_data_clm, temp_data):

    # write_headlines
    check_list = ["EC50", "rsquared", "hillslope", "n_lowdose_datapoints", "std_resp_lowdose_datapoints",
                  "n_highdose_datapoints", "std_resp_highdose_datapoints", "doseconc_stepsize_at_EC50", "saxe_lowdose",
                  "saxe_highdose"]
    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data" and sample_index == 0:
            for data in temp_data[samples]:
                if data in check_list:
                    ws_data.cell(column=ws_data_clm, row=ws_data_row + 1, value=data)
                    ws_data_row += 1

    ws_data_row = 1
    ws_data_clm += 1

    for sample_index, samples in enumerate(temp_data):

        if samples != "state_data":
            ws_data.cell(column=ws_data_clm, row=ws_data_row, value=samples)
            ws_data_row += 1
            for data in temp_data[samples]:
                if data in check_list:
                    temp_value = temp_data[samples][data]["value"]
                    ws_data.cell(column=ws_data_clm, row=ws_data_row, value=temp_value)
                    ws_data_row += 1
        ws_data_row = 1
        ws_data_clm += 1


def _draw_curves(ws_diagram, data_placement, x_title, y_title):

    x_values_min_clm = min(data_placement["dose"]["clm"])
    x_values_min_row = min(data_placement["dose"]["row"])
    x_values_max_row = max(data_placement["dose"]["row"])
    chart_data_sheet = data_placement["sheet"]
    x_values = Reference(chart_data_sheet, min_col=x_values_min_clm, min_row=x_values_min_row, max_row=x_values_max_row)
    chart_list = ["B2", "L2", "V2", "B17", "L17", "V17", "B32", "L32", "V32"]
    chart_counter = 0

    for samples in data_placement:
        if samples != "dose" and samples != "sheet":
            chart = ScatterChart()
            chart.title = f"{samples}"
            chart.style = 15
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            y_value_min_clm = min(data_placement[samples]["clm"])
            y_value_max_clm = max(data_placement[samples]["clm"])
            Y_values_min_row = min(data_placement[samples]["row"])
            Y_values_max_row = max(data_placement[samples]["row"])

            for clm in range(y_value_min_clm, y_value_max_clm + 1):
                values = Reference(chart_data_sheet, min_col=clm, min_row=Y_values_min_row - 1, max_row=Y_values_max_row)
                series = Series(values, x_values, title_from_data=True)
                series.marker = openpyxl.chart.marker.Marker('x')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
            ws_diagram.add_chart(chart, chart_list[chart_counter])
            chart_counter += 1


def _draw_avg_curve(ws, data_placement, x_title, y_title):

    x_values_min_clm = min(data_placement["dose"]["clm"])
    x_values_min_row = min(data_placement["dose"]["row"])
    x_values_max_row = max(data_placement["dose"]["row"])
    chart_data_sheet = data_placement["sheet"]
    x_values = Reference(chart_data_sheet, min_col=x_values_min_clm, min_row=x_values_min_row, max_row=x_values_max_row)
    letter_list = ["B", "L", "V"]
    initial_number = 2
    number_space = 15
    chart_list = []
    for counter in range(10):
        for letters in letter_list:
            temp_number = initial_number + (number_space * counter)
            chart_list.append(f"{letters}{temp_number}")

    chart_list = ["B2", "L2", "V2", "B17", "L17", "V17", "B32", "L32", "V32"]
    chart_counter = 0

    for samples in data_placement:
        if samples != "dose" and samples != "sheet":
            chart = ScatterChart()
            chart.title = f"{samples}"
            chart.style = 15
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            add_curve = True
            for data in data_placement[samples]:
                if add_curve:

                    y_value_min_clm = min(data_placement[samples]["avg_clm"])
                    y_value_max_clm = max(data_placement[samples]["avg_clm"])
                    Y_values_min_row = min(data_placement[samples]["avg_row"])
                    Y_values_max_row = max(data_placement[samples]["avg_row"])
                    for clm in range(y_value_min_clm, y_value_max_clm + 1):
                        values = Reference(chart_data_sheet, min_col=clm, min_row=Y_values_min_row - 1, max_row=Y_values_max_row)
                        series = Series(values, x_values, title_from_data=True)
                        # series.marker = openpyxl.chart.marker.Marker('x')
                        # series.graphicalProperties.line.noFill = True
                        chart.series.append(series)
                    add_curve = False
                if data == "raw":
                    for data_points in data_placement[samples][data]:
                        temp_x_values_min_clm = data_placement["dose"]["clm"][0]
                        temp_x_values_min_row = data_placement["dose"]["row"][data_points]
                        temp_x_values_max_row = data_placement["dose"]["row"][data_points]
                        temp_x_values = Reference(chart_data_sheet, min_col=temp_x_values_min_clm,
                                                  min_row=temp_x_values_min_row, max_row=temp_x_values_max_row)
                        y_value_min_clm = min(data_placement[samples][data][data_points]["clm"])
                        y_value_max_clm = max(data_placement[samples][data][data_points]["clm"])
                        Y_values_min_row = min(data_placement[samples][data][data_points]["row"])
                        Y_values_max_row = max(data_placement[samples][data][data_points]["row"])

                        for clm in range(y_value_min_clm, y_value_max_clm + 1):
                            values = Reference(chart_data_sheet, min_col=clm, min_row=Y_values_min_row - 1,
                                               max_row=Y_values_max_row)
                            series = Series(values, temp_x_values, title_from_data=True)
                            series.marker = openpyxl.chart.marker.Marker('x')
                            series.graphicalProperties.line.noFill = True
                            chart.series.append(series)

            ws.add_chart(chart, chart_list[chart_counter])
            chart_counter += 1
            # else:
            #     print(samples)


def _write_overview(config, all_dose_data, ws_overview, plate_group_to_compound_id, include_structure):
    temp_row = 1
    temp_clm = 1
    if include_structure:
        headlines = ["Temp_name", "Compound_id", "smiles", "structure"]
    else:
        headlines = ["Temp_name", "Compound_id", "smiles"]

    for headline in headlines:
        ws_overview.cell(column=temp_clm, row=temp_row, value=headline)
        temp_clm += 1

    temp_row += 1
    temp_clm = 1

    for plates in all_dose_data:
        for samples in all_dose_data[plates]:
            if samples != "state_data":
                temp_name = samples
                compound_id = plate_group_to_compound_id[samples]
                smiles = _fetch_smiles_data(config, compound_id)

                ws_overview.cell(column=temp_clm + 0, row=temp_row, value=temp_name)
                ws_overview.cell(column=temp_clm + 1, row=temp_row, value=compound_id)
                ws_overview.cell(column=temp_clm + 2, row=temp_row, value=smiles)

        temp_row += 1

    if include_structure:
        insert_structure(ws_overview)


def dose_excel_controller(config, plate_reader_files, all_dose_data, plate_group_to_compound_id, save_location,
                          include_id, include_structure):

    for plate_counter, plates in enumerate(all_dose_data):
        temp_file = plate_reader_files[plate_counter]
        wb = load_workbook(temp_file)
        ws_readings = wb.create_sheet("Raw_Data")
        ws_diagram_raw = wb.create_sheet("Diagram_Raw")
        ws_diagram_fitted = wb.create_sheet("Diagram_Log")
        ws_diagram_avg = wb.create_sheet("Diagram_Avg")
        ws_data = wb.create_sheet("Data")
        x_title_raw = "Dose"
        y_title_raw = "Signal"
        x_title_log = "Dose - Log"
        y_title_log = "Signal"
        # Added to make sure that dose is only added once
        add_dose = True
        add_state_data = True
        replicates = 3

        ws_reading_row = ws_reading_clm = ws_data_row = ws_data_clm = ws_diagram_row = ws_diagram_clm = 1

        temp_data = all_dose_data[plates]

        data_placement_raw, raw_row = _write_dose_readings(ws_readings, ws_reading_row, ws_reading_clm, temp_data,
                                                           add_dose, add_state_data, replicates)
        data_placement_log, max_row = _write_dose_log(ws_readings, temp_data, replicates, raw_row)
        data_placement_avg, max_row = _write_avg_data(ws_readings, temp_data, replicates, max_row)
        _write_dose_data(ws_data, ws_data_row, ws_data_clm, temp_data)
        _draw_curves(ws_diagram_raw, data_placement_raw, x_title_raw, y_title_raw)
        _draw_curves(ws_diagram_fitted, data_placement_log, x_title_log, y_title_log)
        _draw_avg_curve(ws_diagram_avg, data_placement_avg, x_title_log, y_title_log)
        if include_id:
            ws_overview = wb.create_sheet("Overview")
            _write_overview(config, all_dose_data, ws_overview, plate_group_to_compound_id, include_structure)

        name = save_location/f"{plates}_dose_response.xlsx"
        wb.save(name)


if __name__ == "__main__":

    pass
