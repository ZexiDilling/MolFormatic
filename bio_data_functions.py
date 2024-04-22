import PySimpleGUI
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
import re
from import_openpyxl_handler import ex_cell
import datetime
from numpy import histogram, arange
from openpyxl.chart import BarChart, Reference
import xlrd

from extra_functions import row_col_to_cell
from info import plate_384_row


def org(all_data, well):
    """
    Original reading data from the platereader, for bio data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: All the readings from the original readings from the platreader in the same formate as the future methodes
    :rtype: dict
    """
    return all_data["plates"]["original"]["wells"][well]


def norm(all_data, well):
    """
    Normalises the data based on the avg of the minimum for bio data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns a dict over all the values after the have been normalized.
    :rtype: dict
    """
    return all_data["plates"]["original"]["wells"][well] - all_data["calculations"]["original"]["minimum"]["avg"]


def pora(all_data, well):
    """
    Percentage of remaining activity calculation based on the avg of the max of normalised data

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns the Percentage of remaining activity based on the normalized data.
    :rtype: dict
    """
    return ((100 * all_data["plates"]["normalised"]["wells"][well]) /
                all_data["calculations"]["normalised"]["max"]["avg"])


def pora_internal(all_data, well):
    """
    Percentage of remaining activity calculation based on the avg of the max of normalised data
    This should properly be deleted as it is not needed...

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param well: Witch well is being calculated
    :type well: str
    :return: Returns the Percentage of remaining activity based on the original data.
    :rtype: dict
    """
    return ((100 * all_data["plates"]["normalised"]["wells"][well]) /
            all_data["calculations"]["original"]["max"]["avg"])


def z_prime_calculator(all_data, method):
    """
    Calculate Z-prime

    :param all_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type all_data: dict
    :param method: The heading for the calculations that is not avg and stdev
    :type method: str
    :return: Returns the Z-Prime value
    :rtype: int
    """
    # OLD set-up
    # return 1 - ((3 * (all_data["calculations"][method]["max"]["stdev"] +
    #             (all_data["calculations"][method]["minimum"]["stdev"]))) /
    #             abs(all_data["calculations"][method]["max"]["avg"] +
    #             (all_data["calculations"][method]["minimum"]["avg"])))

    # Get the max average and standard deviation, and min average and standard deviation from the all_data dictionary
    max_avg = all_data["calculations"][method]["max"]["avg"]
    max_stdev = all_data["calculations"][method]["max"]["stdev"]
    min_avg = all_data["calculations"][method]["minimum"]["avg"]
    min_stdev = all_data["calculations"][method]["minimum"]["stdev"]

    # Calculate the result
    result = 1 - ((3 * (max_stdev + min_stdev)) / abs(max_avg + min_avg))

    # Return the result
    return result


def state_mapping(config, ws, translate_wells_to_cells, plate, init_row, free_col, temp_dict, methode):
    """
    Colour in the state of the wells and write a guide in the site to translate
    Might need to re-writes this modul, to exclude temp_dict, as it should not be needed, as all the information should
    come from the plate layout

    :param config: The config file
    :type config: configparser.ConfigParser
    :param ws: Witch worksheet that mappings is conected too
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param translate_wells_to_cells: A dict containing what well in the plate belongs to what cell in the excel file
    :type translate_wells_to_cells: dict
    :param plate: Plate layout for what well have what state
    :type plate: dict
    :param init_row: Row to start writing to
    :type init_row: int
    :param free_col: What column should be free / the first column after the last column used for the plate data
    :type free_col: int
    :param temp_dict: The data that is being analysed for the state mapping
    :type temp_dict: dict
    :param methode: The Method that is being used
    :type methode: str
    :return: The colouring of the cells, and a reading guide for the colours, in the excel ark
    """
    # Iterate through the wells in the plate
    init_row_start = init_row  # save the initial starting row
    print(plate)
    for counter in plate:
        state = plate[counter]["state"]  # get the state of the current well
        cell_color = config["plate_colouring"][state]  # get the color for the current state
        cell_color = cell_color.replace("#", "")  # remove the "#" symbol from the color
        temp_cell = translate_wells_to_cells[
            plate[counter]["well_id"]]  # get the cell for the current well
        # fill the cell with the color for the current state
        ws[temp_cell].fill = PatternFill("solid", fgColor=cell_color)

    # write the color guide in the worksheet
    for state in temp_dict["plates"][methode]:
        if state != "wells":  # skip the "wells" state
            if init_row_start == init_row:  # write the header only once
                ws[ex_cell(init_row + 1, free_col)] = "well state"
                # ws[ex_cell(init_row + 1, free_col + 1)] = "colour coding"
                # bold the header
                ws[ex_cell(init_row + 1, free_col)].font = Font(b=True)
                # ws[ex_cell(init_row + 1, free_col + 1)].font = Font(b=True)
            # Writes the state and colour each state with the right colour
            temp_colour = config["plate_colouring"][state]
            temp_colour = temp_colour.replace("#", "")
            ws[ex_cell(init_row + 2, free_col)] = state
            ws[ex_cell(init_row + 2, free_col)].fill = PatternFill("solid", fgColor=temp_colour)

            # write the color for the state
            # ws[ex_cell(init_row + 2, free_col + 1)] = config["plate_colouring"][state]
            init_row += 1  # increment the row for the next state


def heatmap(config, ws, pw_dict, translate_wells_to_cells, heatmap_colours):
    """
    Colour code based on values.

    :param ws: The worksheet where the heat map is placed
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param pw_dict: Dict for each well and what state it is (sample, blank, empty....
    :type pw_dict: dict
    :param translate_wells_to_cells: Dict for each well cells value
    :type translate_wells_to_cells: dict
    :return: A heatmap coloured depending on the options, in the excel file.
    """
    # create a list of well ids where the sample is present
    temp_list = [well for well in pw_dict if pw_dict[well] == "sample"]

    # add conditional formatting to the selected wells to create a heatmap
    print(translate_wells_to_cells)
    print(temp_list)
    ws.conditional_formatting.add(
        f"{translate_wells_to_cells[temp_list[0]]}:{translate_wells_to_cells[temp_list[-1]]}",
        ColorScaleRule(
            # set the starting color based on the 10th percentile of the data
            start_type='percentile',
            start_value=10,
            start_color=config["Settings_bio"]["plate_report_heatmap_colours_low"].replace("#", ""),

            # set the middle color based on the 50th percentile of the data
            mid_type='percentile',
            mid_value=50,
            mid_color=config["Settings_bio"]["plate_report_heatmap_colours_mid"].replace("#", ""),

            # set the end color based on the 90th percentile of the data
            end_type='percentile',
            end_value=90,
            end_color=config["Settings_bio"]["plate_report_heatmap_colours_high"].replace("#", "")
        )
    )
    # 2 colours heat mao
    # ws.conditional_formatting.add(f"{translate_wells_to_cells[temp_list[0]]}:"
    #                               f"{translate_wells_to_cells[temp_list[-1]]}",
    #                               ColorScaleRule(start_type="min",
    #                                              start_color=config["colours to hex"][heatmap_colours["start"]],
    #                                              end_type="max",
    #                                              end_color=config["colours to hex"][heatmap_colours["end"]]))


def hit_mapping(ws, temp_dict, pora_threshold, methode, translate_wells_to_cells, free_col, init_row):
    """
    Colour coding a plate depending on hits (the values are selected before hand), and writes the bounderies for the
    hits.

    :param ws: The worksheet where the hit map is
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param temp_dict: A dicts from all_data, that is only containing the values for the plate. These well values are
        the basis for the hits
    :type temp_dict: dict
    :param pora_threshold: The hit map threshold / bounderies
    :type pora_threshold: dict
    :param methode: The method that is being used
    :type methode: str
    :param translate_wells_to_cells: A dict containing what well in the plate belongs to what cell in the excel file
    :type translate_wells_to_cells: dict
    :param free_col: The first free column after the plate.
    :type free_col: int
    :param init_row: The row where the plate can state to be writen in the excel sheet.
    :type init_row: int
    :return: Colouring the cells in the excel file, depending on the boundaries of the hit.
    """

    # Iterate over all wells and fill the cells based on their pora value
    for wells in translate_wells_to_cells:
        if wells not in temp_dict["plates"][methode]["empty"]:
            for split in pora_threshold:
                # Check if the current split is not "colour"
                if split != "colour":
                    # check if it should be added:
                    if pora_threshold[split]["use"]:

                        # Check if the current well's pora value falls within the range of the current split
                        if float(pora_threshold[split]["min"]) < temp_dict["plates"][methode]["wells"][wells] < float(
                                pora_threshold[split]["max"]):
                            # Get the colour for the current split
                            temp_colour = pora_threshold["colour"][split]
                            # Remove the "#" from the colour code
                            cell_color = temp_colour.replace("#", "")
                            temp_cell = translate_wells_to_cells[wells]
                            # Fill the cell with the calculated color
                            ws[temp_cell].fill = PatternFill("solid", fgColor=cell_color)

    # Initialize the row where the colour guide will be written
    write_row = init_row + 1
    # Write the colour guide
    for threshold in pora_threshold:
        if threshold != "colour":
            # Write the threshold
            ws[ex_cell(write_row, free_col)] = threshold
            # Bold and underline the threshold
            ws[ex_cell(write_row, free_col)].font = Font(b=True, underline="single")
            # colour the threshold
            temp_colour = pora_threshold["colour"][threshold]
            temp_colour = temp_colour.replace("#", "")
            # Write the colour for the level
            ws[ex_cell(write_row, free_col)].fill = PatternFill("solid", fgColor=temp_colour)
            for level in pora_threshold[threshold]:
                if level != "use":
                    # Write the level
                    ws[ex_cell(write_row, free_col + 1)] = level

                    # write the value
                    ws[ex_cell(write_row, free_col + 2)] = pora_threshold[threshold][level]
                    write_row += 1


def frequency_writer(ws, headline, data_set, free_col, initial_row, bin_min, bin_max, bin_width, include_outliers):
    """
    Writes the frequency of samples with a score within the borders of each bin

    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param freq_data:
    :param free_col: What colmn is free in the sheet to start writing on
    :type free_col: int
    :param initial_row: what is the top row to start writing in
    :type initial_row: int
    :param bin_min: Min value for bin. => Min value for frequency data set to look at.
        can either be a value, or be a string, to calculate the bin based on the data set
    type bin_min: int or str
    :param bin_max: Max value for bin. => Max value for frequency data set to look at.
        can either be a value, or be a string, to calculate the bin based on the data set
    type bin_max: int or str
    :param bin_width: How many values is the frequency data split up into.
    :type bin_width: int
    :param include_outliers: if outliers should be included in the data
    :type include_outliers: bool
    :return: frequency data writen up
    """
    col = free_col
    data_bin_min = None
    data_bin_max = None
    bin_min_seperator = 0
    bin_max_seperator = 0

    row = initial_row
    # write headline:
    ws.cell(column=col, row=row, value=headline).font = Font(b=True)
    ws.cell(column=col + 1, row=row, value="Bin Values").font = Font(b=True)
    row += 1

    # Sets bin range to be range of samples values
    if bin_min == "data_set":
        bin_min = min(data_set)
    else:
        data_bin_min = min(data_set)
    if bin_max == "data_set":
        bin_max = max(data_set)
    else:
        data_bin_max = max(data_set)

    # Setup a hist for values below bin_min
    if data_bin_min < bin_min:
        bin_min_width = bin_min - data_bin_min
        min_hist_list_1, min_hist_list_2 = histogram(data_set, bins=arange(data_bin_min, bin_min +
                                                                                bin_width, bin_min_width))
        # Writes in data
        ws.cell(column=col, row=row, value=min_hist_list_1[0])
        ws.cell(column=col + 1, row=row, value=min_hist_list_2[0])

        bin_min_seperator = 1

    temp_list_1, temp_list_2 = \
        histogram(data_set, bins=arange(bin_min, bin_max + bin_width, bin_width))

    # temp_list_1, temp_list_2 = \
    #     histogram(temp_data_set)
    # write data points:
    for data_point in range(len(temp_list_1)):
        ws.cell(column=col, row=row + data_point + bin_min_seperator, value=temp_list_1[data_point])
        ws.cell(column=col + 1, row=row + data_point + bin_min_seperator, value=temp_list_2[data_point])

    # Setup a hist for values above bin_max
    if data_bin_max > bin_max:
        bin_max_width = data_bin_max - bin_max
        max_hist_list_1, max_hist_list_2 = histogram(data_set, bins=arange(bin_max, data_bin_max +
                                                                 bin_width, bin_max_width))

        ws.cell(column=col, row=row + data_point + bin_min_seperator + bin_max_seperator
                , value=max_hist_list_1[0])
        ws.cell(column=col + 1, row=row + data_point + bin_min_seperator + bin_max_seperator
                , value=max_hist_list_2[1])

    col += 1
    max_row = row + data_point + bin_max_seperator + bin_min_seperator

    if include_outliers:
        data_location = {"min_col": free_col, "min_row": initial_row + 1,
                         "max_col": free_col, "max_row": max_row - 1}
        category_location = {"min_col": col, "min_row": initial_row + 1,
                             "max_col": col, "max_row": max_row - 1}
    else:
        data_location = {"min_col": free_col, "min_row": initial_row,
                         "max_col": free_col, "max_row": max_row}
        category_location = {"min_col": col, "min_row": initial_row,
                             "max_col": col, "max_row": max_row}


    return col, data_location, category_location


def bar_chart(ws, title, free_col, initial_row, data_location, category_location):
    """
    Generate a bar chart
    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param title: The headline for the data
    :type title: str
    :param free_col: What colmn is free in the sheet to start writing on
    :type free_col: int
    :param initial_row: what is the top row to start writing in
    :type initial_row: int
    :param data_location: A dict with location (row, col) in the excelsheet, where the data for the
        bar-chart is located
    :type data_location: dict
    :param category_location: A dict with location (row, col) in the excelsheet, where the category for the
        bar-chart is located
    :type category_location: dict
    :param bin_max_seperator: A counter if there are values outside max
    :type bin_max_seperator: int
    :param bin_min_seperator: A counter if there are values outside min
    :type bin_min_seperator : int
    :return:
    """
    col_placement = free_col + 2

    # set up the chart, and the look
    # ToDo move this to settings
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.x_axis.title = "Bins"
    chart.y_axis.title = "Sample amount"
    chart.shape = 4

    data_set = Reference(ws, min_col=data_location["min_col"], min_row=data_location["min_row"],
                         max_col=data_location["max_col"], max_row=data_location["max_row"])
    category_set = Reference(ws, min_col=category_location["min_col"], min_row=category_location["min_row"],
                         max_col=category_location["max_col"], max_row=category_location["max_row"])

    chart.add_data(data_set, titles_from_data=False)
    chart.set_categories(category_set)
    chart_cell = row_col_to_cell(initial_row, col_placement)
    ws.add_chart(chart, chart_cell)


def well_row_col_type(plate_layout):
    """
    Makes two dicts. one for what type/state each well/cell is and a dict that translate each well to a column and row
    values

    :param plate_layout: The plate-layout for what state each well is in (sample, min, max...)
    :type plate_layout: dict
    :return: well_col_row: a dict with a list of values for rows and for columns, well_type: a dict over the
        type/state of each well/cell
    :rtype: dict, dict
    """
    well_type = {}
    well_col_row = {"well_col": [], "well_row": []}
    temp_plate_layout = plate_layout.get("well_layout", plate_layout)

    # Extract the well information
    for counter in temp_plate_layout:
        for keys in temp_plate_layout[counter]:
            # Get well id and column, row information
            if keys == "well_id":
                temp_well_row = re.sub(r"\d+", "", temp_plate_layout[counter][keys])
                temp_well_col = re.sub(r"\D+", "", temp_plate_layout[counter][keys])
                if temp_well_row not in well_col_row["well_row"]:
                    well_col_row["well_row"].append(temp_well_row)
                if temp_well_col not in well_col_row["well_col"]:
                    well_col_row["well_col"].append(temp_well_col)

            # Get well type information
            if keys == "state":
                well_type.setdefault(temp_plate_layout[counter]["state"], []).append(
                    temp_plate_layout[counter]["well_id"])

    return well_col_row, well_type


def _init_data_reading(file, plate_layout):
    well_col_row, well_type = well_row_col_type(plate_layout)
    try:
        wb = load_workbook(file)
    except InvalidFileException:
        return "fail"
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    # Initialize variables for plate type detection
    plate_type_384 = False
    plate_type_1536 = False

    # Initialize dictionary for all data
    all_data = {"plates": {}, "calculations": {}}
    n_rows = len(well_col_row["well_row"])
    return ws, all_data, n_rows, sheet, well_col_row, well_type


def _get_reverse_well(well_name, plate_type=384):
    if plate_type == 384:
        index_of_C = plate_384_row.index(well_name)
        new_well_name_index = 384 - index_of_C - 1

        new_well_name = plate_384_row[new_well_name_index]

        return new_well_name


def csv_converter(file, reverse_plate):

    wb = Workbook()
    ws = wb.active
    counter = 1
    name_index = 0
    well_number_index = 0
    well_letter_index = 0
    with open(file) as f:
        for line_index, line in enumerate(f):
            values = line.split("\t")

            for value_index, value in enumerate(values):
                value = value.strip('"')
                ws.cell(column=value_index + 1, row=counter, value=value)

                if line_index == 0:
                    if value.casefold() == "hor.":
                        well_letter_index = value_index
                    elif value.casefold() == "ver.":
                        well_number_index = value_index
                    elif value.casefold() == "name":
                        name_index = value_index
                else:
                    if value_index == well_letter_index:
                        well_letter = value
                    elif value_index == well_number_index:
                        well_number = value
                    elif value_index == name_index:
                        if not reverse_plate:
                            well_name = f"{well_letter}{well_number}"
                        else:
                            well_name = f"{well_letter}{well_number}"
                            well_name = _get_reverse_well(well_name)
                        ws.cell(column=value_index + 1, row=counter, value=well_name)

            counter += 1

    excel_file = file.split(".")[0]
    excel_file = f"{excel_file}.xlsx"
    wb.save(excel_file)
    print(excel_file)
    return excel_file


def xls_converter(file):

    book_xls = xlrd.open_workbook(file)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
    dst_file_path = file.split(".")[0]
    dst_file_path = f"{dst_file_path}.xlsx"
    book_xlsx.save(dst_file_path)
    return dst_file_path



def multiscan_data_dict(file, plate_layout, mixed_up):
    """
       The original data from the plate reader, loaded in from the excel file

       :param file: the Excel file, with the platereaders data
       :type file: str
       :param plate_layout: The platelayout to tell witch cell/well is in witch state (sample, blank, empty....)
       :type plate_layout: dict
       :param mixed_up: If the plate is put in the machine wrong
       :type mixed_up: bool
       :return:
           - all_data: all_data: A dict over the original data, and what each cell/well is
           - well_col_row: a dict with a list of values for rows and for columns,
           - well_type: a dict over the type/state of each well/cell
           - barcode: The barcode for the plate
       :rtype:
           - dict
           - dict
           - dict
           - str
       """

    check = _init_data_reading(file, plate_layout)
    if isinstance(check, str):
        return
    else:
        ws, all_data, n_rows, sheet, well_col_row, well_type = check

    compound_id_location = 0
    time_slots_location = 0
    time_slot = []
    time_slot_counter = {}
    time_counter = 0
    all_data["plates"]["original"] = {}
    all_data["plates"]["original"]["wells"] = {}
    well_letter_location = 0
    well_number_location = 0

    for row_index, row in enumerate(ws.values):
        for value_index, value in enumerate(row):
            if row_index == 0:
                if isinstance(value, str):
                    if value.casefold() == "id":
                        id_location = value_index

                    if value.casefold() == "hor.":
                        well_letter_location = value_index
                    if value.casefold() == "ver.":
                        well_number_location = value_index

                    if value.casefold() == "name":
                        compound_id_location = value_index
                if 0 < compound_id_location < value_index:
                    if time_slots_location == 0:
                        time_slots_location = value_index
                    time_slot.append(value)
                    time_slot_counter[time_counter] = value
                    time_counter += 1
            else:
                if row_index == 1 and value_index == id_location:
                    if "+" in value:
                        style = value.split("+")[1].split("-")[0].strip()
                    barcode = value.split(" ")[0].split("-")[0]
                    date = value.split(" ")[-1].split(")")[0]
                well_id = f"{row[well_letter_location]}{row[well_number_location]}"
                compound_id = row[compound_id_location]
                if mixed_up:
                    all_data["plates"]["original"]["wells"][compound_id] = {"compound_id": compound_id, "well_data": [row[time_slots_location:len(row)]]}
                else:
                    all_data["plates"]["original"]["wells"][well_id] = {"compound_id": compound_id, "well_data": [
                        row[time_slots_location:len(row)]]}

    all_data["time_slots"] = time_slot_counter

    try:
        barcode
    except UnboundLocalError:
        barcode = file.split("/")[-1].split(" ")[-1]

    try:
        date
    except UnboundLocalError:
        date = datetime.today()

    try:
        style
    except UnboundLocalError:
        style = "org"

    return all_data, well_col_row, well_type, barcode, date, style

def original_data_dict(file, plate_layout):
    """
    The original data from the plate reader, loaded in from the excel file

    :param file: the Excel file, with the platereaders data
    :type file: str
    :param plate_layout: The platelayout to tell witch cell/well is in witch state (sample, blank, empty....)
    :type plate_layout: dict
    :return:
        - all_data: all_data: A dict over the original data, and what each cell/well is
        - well_col_row: a dict with a list of values for rows and for columns,
        - well_type: a dict over the type/state of each well/cell
        - barcode: The barcode for the plate
    :rtype:
        - dict
        - dict
        - dict
        - str
    """

    check = _init_data_reading(file, plate_layout)
    if isinstance(check, str):
        return
    else:
        ws, all_data, n_rows, sheet, well_col_row, well_type = check

    # Iterate through each row in the sheet
    for row_index, row in enumerate(ws.values):     #Todo Change this to be a dict, as there can be multiple readings per data. There can be multiple plate readings inside an excel sheet
        for index_row, value in enumerate(row):

            # Get the date
            if value == "Date:":
                date = row[1]

            # Get the barcode
            if value == "Barcode:" and row[1]:
                barcode = row[1]

            # Get the number of rows to skip
            if value == "<>":
                skipped_rows = row_index

            # Check if the plate is 384-well
            if value == "I":
                plate_type_384 = True

            # Check if the plate is 1536-well
            if value == "AA":
                plate_type_1536 = True

    # Load the data into a pandas dataframe, skipping the rows that were specified earlier
    try:
        df_plate = pd.read_excel(file, sheet_name=sheet, skiprows=skipped_rows, nrows=n_rows)
    except UnboundLocalError:
        print(f"Could not find the <> tag in the file: {file}")
        return

    # Convert the dataframe to a dictionary
    df_plate_dict = df_plate.to_dict()

    # Check the size of the plate and display an error message if the size does not match the expected size
    expected_plate_sizes = {
        8: not plate_type_384 or plate_type_1536,
        16: plate_type_384,
        32: plate_type_1536
    }

    if n_rows not in expected_plate_sizes or not expected_plate_sizes[n_rows]:
        PySimpleGUI.PopupError(f"Wrong plate size, data is not {expected_plate_sizes[n_rows]}-wells")
        return None, None, None

    # temp_reading_dict stores the data from the dataframe in a dictionary format
    temp_reading_dict = {}
    for counter, heading in enumerate(df_plate_dict):
        for index, values in enumerate(df_plate_dict[heading]):
            temp_reading_dict[f"{well_col_row['well_row'][index]}{counter}"] = df_plate_dict[heading][values]

    # The code below stores the data in the all_data dictionary
    all_data["plates"]["original"] = {}
    all_data["plates"]["original"]["wells"] = {}
    for state in well_type:
        for well in well_type[state]:
            try:
                all_data["plates"]["original"]["wells"][well] = temp_reading_dict[well]

            # Incase blanked wells are skipped for the reading
            except KeyError:
                all_data["plates"]["original"]["wells"][well] = 1
    try:
        barcode
    except UnboundLocalError:
        barcode = file.split("/")[-1]

    try:
        date
    except UnboundLocalError:
        date = datetime.today()

    return all_data, well_col_row, well_type, barcode, date


def txt_to_xlsx(file):

    # get file name from the "file-path".
    file_name = file.split(".")[0]
    # setup the excel sheet:
    wb = Workbook()
    ws = wb.active

    col = 1
    row = 1

    file = open(file, "r")
    lines = file.readlines()

    for row_index, line in enumerate(lines):
        line = line.removesuffix("\n")
        line = line.strip()
        line = " ".join(line.split())
        line = line.split(" ")
        col_index_1 = 0
        for values in line:
            values = values.split("\t")
            for col_index_2, value in enumerate(values):
                ws.cell(column=col + col_index_1 + col_index_2, row=row + row_index, value=value)
            col_index_1 += 1

    file_name = f"{file_name}.xlsx"
    wb.save(file_name)

    return file_name


if __name__ == "__main__":
    # file = r"C:\Users\phch\OneDrive - Danmarks Tekniske Universitet\Mapper\Python_data\procul\raw\procul_002 + cas-EPSON Perfection V19-FnSm-240322-1155-Avelina calibration (Emilie 21-03-24).xlsx"
    # plate_layout = {1: {'group': 0, 'well_id': 'A1', 'state': 'blank', 'colour': '#7ebfc2'}, 2: {'group': 0, 'well_id': 'B1', 'state': 'blank', 'colour': '#7ebfc2'}, 3: {'group': 0, 'well_id': 'C1', 'state': 'blank', 'colour': '#7ebfc2'}, 4: {'group': 0, 'well_id': 'D1', 'state': 'blank', 'colour': '#7ebfc2'}, 5: {'group': 0, 'well_id': 'E1', 'state': 'blank', 'colour': '#7ebfc2'}, 6: {'group': 0, 'well_id': 'F1', 'state': 'blank', 'colour': '#7ebfc2'}, 7: {'group': 0, 'well_id': 'G1', 'state': 'blank', 'colour': '#7ebfc2'}, 8: {'group': 0, 'well_id': 'H1', 'state': 'blank', 'colour': '#7ebfc2'}, 9: {'group': 0, 'well_id': 'I1', 'state': 'blank', 'colour': '#7ebfc2'}, 10: {'group': 0, 'well_id': 'J1', 'state': 'blank', 'colour': '#7ebfc2'}, 11: {'group': 0, 'well_id': 'K1', 'state': 'blank', 'colour': '#7ebfc2'}, 12: {'group': 0, 'well_id': 'L1', 'state': 'blank', 'colour': '#7ebfc2'}, 13: {'group': 0, 'well_id': 'M1', 'state': 'blank', 'colour': '#7ebfc2'}, 14: {'group': 0, 'well_id': 'N1', 'state': 'blank', 'colour': '#7ebfc2'}, 15: {'group': 0, 'well_id': 'O1', 'state': 'blank', 'colour': '#7ebfc2'}, 16: {'group': 0, 'well_id': 'P1', 'state': 'blank', 'colour': '#7ebfc2'}, 17: {'group': 0, 'well_id': 'A2', 'state': 'blank', 'colour': '#7ebfc2'}, 18: {'group': 0, 'well_id': 'B2', 'state': 'blank', 'colour': '#7ebfc2'}, 19: {'group': 0, 'well_id': 'C2', 'state': 'blank', 'colour': '#7ebfc2'}, 20: {'group': 0, 'well_id': 'D2', 'state': 'blank', 'colour': '#7ebfc2'}, 21: {'group': 0, 'well_id': 'E2', 'state': 'blank', 'colour': '#7ebfc2'}, 22: {'group': 0, 'well_id': 'F2', 'state': 'blank', 'colour': '#7ebfc2'}, 23: {'group': 0, 'well_id': 'G2', 'state': 'blank', 'colour': '#7ebfc2'}, 24: {'group': 0, 'well_id': 'H2', 'state': 'blank', 'colour': '#7ebfc2'}, 25: {'group': 0, 'well_id': 'I2', 'state': 'blank', 'colour': '#7ebfc2'}, 26: {'group': 0, 'well_id': 'J2', 'state': 'blank', 'colour': '#7ebfc2'}, 27: {'group': 0, 'well_id': 'K2', 'state': 'blank', 'colour': '#7ebfc2'}, 28: {'group': 0, 'well_id': 'L2', 'state': 'blank', 'colour': '#7ebfc2'}, 29: {'group': 0, 'well_id': 'M2', 'state': 'blank', 'colour': '#7ebfc2'}, 30: {'group': 0, 'well_id': 'N2', 'state': 'blank', 'colour': '#7ebfc2'}, 31: {'group': 0, 'well_id': 'O2', 'state': 'blank', 'colour': '#7ebfc2'}, 32: {'group': 0, 'well_id': 'P2', 'state': 'blank', 'colour': '#7ebfc2'}, 33: {'group': 0, 'well_id': 'A3', 'state': 'blank', 'colour': '#7ebfc2'}, 34: {'group': 0, 'well_id': 'B3', 'state': 'blank', 'colour': '#7ebfc2'}, 35: {'group': '1', 'well_id': 'C3', 'state': 'positive', 'colour': '#09ca0e'}, 36: {'group': 0, 'well_id': 'D3', 'state': 'blank', 'colour': '#7ebfc2'}, 37: {'group': 0, 'well_id': 'E3', 'state': 'blank', 'colour': '#7ebfc2'}, 38: {'group': 0, 'well_id': 'F3', 'state': 'blank', 'colour': '#7ebfc2'}, 39: {'group': 0, 'well_id': 'G3', 'state': 'blank', 'colour': '#7ebfc2'}, 40: {'group': 0, 'well_id': 'H3', 'state': 'blank', 'colour': '#7ebfc2'}, 41: {'group': 0, 'well_id': 'I3', 'state': 'blank', 'colour': '#7ebfc2'}, 42: {'group': 0, 'well_id': 'J3', 'state': 'blank', 'colour': '#7ebfc2'}, 43: {'group': 0, 'well_id': 'K3', 'state': 'blank', 'colour': '#7ebfc2'}, 44: {'group': 0, 'well_id': 'L3', 'state': 'blank', 'colour': '#7ebfc2'}, 45: {'group': '10', 'well_id': 'M3', 'state': 'positive', 'colour': '#09ca0e'}, 46: {'group': 0, 'well_id': 'N3', 'state': 'blank', 'colour': '#7ebfc2'}, 47: {'group': 0, 'well_id': 'O3', 'state': 'blank', 'colour': '#7ebfc2'}, 48: {'group': 0, 'well_id': 'P3', 'state': 'negative', 'colour': '#d20000'}, 49: {'group': 0, 'well_id': 'A4', 'state': 'blank', 'colour': '#7ebfc2'}, 50: {'group': 0, 'well_id': 'B4', 'state': 'blank', 'colour': '#7ebfc2'}, 51: {'group': 0, 'well_id': 'C4', 'state': 'blank', 'colour': '#7ebfc2'}, 52: {'group': 0, 'well_id': 'D4', 'state': 'blank', 'colour': '#7ebfc2'}, 53: {'group': 0, 'well_id': 'E4', 'state': 'blank', 'colour': '#7ebfc2'}, 54: {'group': 0, 'well_id': 'F4', 'state': 'blank', 'colour': '#7ebfc2'}, 55: {'group': '12', 'well_id': 'G4', 'state': 'positive', 'colour': '#09ca0e'}, 56: {'group': 0, 'well_id': 'H4', 'state': 'blank', 'colour': '#7ebfc2'}, 57: {'group': 0, 'well_id': 'I4', 'state': 'blank', 'colour': '#7ebfc2'}, 58: {'group': 0, 'well_id': 'J4', 'state': 'blank', 'colour': '#7ebfc2'}, 59: {'group': 0, 'well_id': 'K4', 'state': 'blank', 'colour': '#7ebfc2'}, 60: {'group': 0, 'well_id': 'L4', 'state': 'blank', 'colour': '#7ebfc2'}, 61: {'group': 0, 'well_id': 'M4', 'state': 'blank', 'colour': '#7ebfc2'}, 62: {'group': 0, 'well_id': 'N4', 'state': 'blank', 'colour': '#7ebfc2'}, 63: {'group': 0, 'well_id': 'O4', 'state': 'blank', 'colour': '#7ebfc2'}, 64: {'group': 0, 'well_id': 'P4', 'state': 'blank', 'colour': '#7ebfc2'}, 65: {'group': 0, 'well_id': 'A5', 'state': 'negative', 'colour': '#d20000'}, 66: {'group': 0, 'well_id': 'B5', 'state': 'blank', 'colour': '#7ebfc2'}, 67: {'group': 0, 'well_id': 'C5', 'state': 'blank', 'colour': '#7ebfc2'}, 68: {'group': 0, 'well_id': 'D5', 'state': 'blank', 'colour': '#7ebfc2'}, 69: {'group': 0, 'well_id': 'E5', 'state': 'blank', 'colour': '#7ebfc2'}, 70: {'group': 0, 'well_id': 'F5', 'state': 'blank', 'colour': '#7ebfc2'}, 71: {'group': 0, 'well_id': 'G5', 'state': 'blank', 'colour': '#7ebfc2'}, 72: {'group': 0, 'well_id': 'H5', 'state': 'blank', 'colour': '#7ebfc2'}, 73: {'group': 0, 'well_id': 'I5', 'state': 'blank', 'colour': '#7ebfc2'}, 74: {'group': 0, 'well_id': 'J5', 'state': 'blank', 'colour': '#7ebfc2'}, 75: {'group': 0, 'well_id': 'K5', 'state': 'blank', 'colour': '#7ebfc2'}, 76: {'group': 0, 'well_id': 'L5', 'state': 'blank', 'colour': '#7ebfc2'}, 77: {'group': 0, 'well_id': 'M5', 'state': 'blank', 'colour': '#7ebfc2'}, 78: {'group': 0, 'well_id': 'N5', 'state': 'blank', 'colour': '#7ebfc2'}, 79: {'group': 0, 'well_id': 'O5', 'state': 'blank', 'colour': '#7ebfc2'}, 80: {'group': 0, 'well_id': 'P5', 'state': 'blank', 'colour': '#7ebfc2'}, 81: {'group': 0, 'well_id': 'A6', 'state': 'blank', 'colour': '#7ebfc2'}, 82: {'group': 0, 'well_id': 'B6', 'state': 'blank', 'colour': '#7ebfc2'}, 83: {'group': 0, 'well_id': 'C6', 'state': 'blank', 'colour': '#7ebfc2'}, 84: {'group': 0, 'well_id': 'D6', 'state': 'blank', 'colour': '#7ebfc2'}, 85: {'group': 0, 'well_id': 'E6', 'state': 'blank', 'colour': '#7ebfc2'}, 86: {'group': 0, 'well_id': 'F6', 'state': 'blank', 'colour': '#7ebfc2'}, 87: {'group': 0, 'well_id': 'G6', 'state': 'blank', 'colour': '#7ebfc2'}, 88: {'group': 0, 'well_id': 'H6', 'state': 'blank', 'colour': '#7ebfc2'}, 89: {'group': 0, 'well_id': 'I6', 'state': 'blank', 'colour': '#7ebfc2'}, 90: {'group': 0, 'well_id': 'J6', 'state': 'blank', 'colour': '#7ebfc2'}, 91: {'group': 0, 'well_id': 'K6', 'state': 'negative', 'colour': '#d20000'}, 92: {'group': 0, 'well_id': 'L6', 'state': 'blank', 'colour': '#7ebfc2'}, 93: {'group': 0, 'well_id': 'M6', 'state': 'blank', 'colour': '#7ebfc2'}, 94: {'group': 0, 'well_id': 'N6', 'state': 'blank', 'colour': '#7ebfc2'}, 95: {'group': 0, 'well_id': 'O6', 'state': 'blank', 'colour': '#7ebfc2'}, 96: {'group': 0, 'well_id': 'P6', 'state': 'blank', 'colour': '#7ebfc2'}, 97: {'group': 0, 'well_id': 'A7', 'state': 'blank', 'colour': '#7ebfc2'}, 98: {'group': 0, 'well_id': 'B7', 'state': 'blank', 'colour': '#7ebfc2'}, 99: {'group': 0, 'well_id': 'C7', 'state': 'blank', 'colour': '#7ebfc2'}, 100: {'group': 0, 'well_id': 'D7', 'state': 'blank', 'colour': '#7ebfc2'}, 101: {'group': 0, 'well_id': 'E7', 'state': 'blank', 'colour': '#7ebfc2'}, 102: {'group': 0, 'well_id': 'F7', 'state': 'blank', 'colour': '#7ebfc2'}, 103: {'group': 0, 'well_id': 'G7', 'state': 'blank', 'colour': '#7ebfc2'}, 104: {'group': 0, 'well_id': 'H7', 'state': 'blank', 'colour': '#7ebfc2'}, 105: {'group': 0, 'well_id': 'I7', 'state': 'blank', 'colour': '#7ebfc2'}, 106: {'group': 0, 'well_id': 'J7', 'state': 'blank', 'colour': '#7ebfc2'}, 107: {'group': 0, 'well_id': 'K7', 'state': 'blank', 'colour': '#7ebfc2'}, 108: {'group': 0, 'well_id': 'L7', 'state': 'blank', 'colour': '#7ebfc2'}, 109: {'group': 0, 'well_id': 'M7', 'state': 'blank', 'colour': '#7ebfc2'}, 110: {'group': 0, 'well_id': 'N7', 'state': 'blank', 'colour': '#7ebfc2'}, 111: {'group': 0, 'well_id': 'O7', 'state': 'blank', 'colour': '#7ebfc2'}, 112: {'group': 0, 'well_id': 'P7', 'state': 'blank', 'colour': '#7ebfc2'}, 113: {'group': 0, 'well_id': 'A8', 'state': 'blank', 'colour': '#7ebfc2'}, 114: {'group': 0, 'well_id': 'B8', 'state': 'blank', 'colour': '#7ebfc2'}, 115: {'group': 0, 'well_id': 'C8', 'state': 'blank', 'colour': '#7ebfc2'}, 116: {'group': 0, 'well_id': 'D8', 'state': 'blank', 'colour': '#7ebfc2'}, 117: {'group': 0, 'well_id': 'E8', 'state': 'blank', 'colour': '#7ebfc2'}, 118: {'group': 0, 'well_id': 'F8', 'state': 'blank', 'colour': '#7ebfc2'}, 119: {'group': 0, 'well_id': 'G8', 'state': 'blank', 'colour': '#7ebfc2'}, 120: {'group': 0, 'well_id': 'H8', 'state': 'blank', 'colour': '#7ebfc2'}, 121: {'group': 0, 'well_id': 'I8', 'state': 'blank', 'colour': '#7ebfc2'}, 122: {'group': 0, 'well_id': 'J8', 'state': 'blank', 'colour': '#7ebfc2'}, 123: {'group': 0, 'well_id': 'K8', 'state': 'blank', 'colour': '#7ebfc2'}, 124: {'group': 0, 'well_id': 'L8', 'state': 'blank', 'colour': '#7ebfc2'}, 125: {'group': 0, 'well_id': 'M8', 'state': 'blank', 'colour': '#7ebfc2'}, 126: {'group': 0, 'well_id': 'N8', 'state': 'blank', 'colour': '#7ebfc2'}, 127: {'group': 0, 'well_id': 'O8', 'state': 'blank', 'colour': '#7ebfc2'}, 128: {'group': '5', 'well_id': 'P8', 'state': 'positive', 'colour': '#09ca0e'}, 129: {'group': 0, 'well_id': 'A9', 'state': 'blank', 'colour': '#7ebfc2'}, 130: {'group': 0, 'well_id': 'B9', 'state': 'blank', 'colour': '#7ebfc2'}, 131: {'group': 0, 'well_id': 'C9', 'state': 'blank', 'colour': '#7ebfc2'}, 132: {'group': 0, 'well_id': 'D9', 'state': 'blank', 'colour': '#7ebfc2'}, 133: {'group': 0, 'well_id': 'E9', 'state': 'blank', 'colour': '#7ebfc2'}, 134: {'group': 0, 'well_id': 'F9', 'state': 'blank', 'colour': '#7ebfc2'}, 135: {'group': 0, 'well_id': 'G9', 'state': 'blank', 'colour': '#7ebfc2'}, 136: {'group': 0, 'well_id': 'H9', 'state': 'blank', 'colour': '#7ebfc2'}, 137: {'group': 0, 'well_id': 'I9', 'state': 'blank', 'colour': '#7ebfc2'}, 138: {'group': 0, 'well_id': 'J9', 'state': 'blank', 'colour': '#7ebfc2'}, 139: {'group': 0, 'well_id': 'K9', 'state': 'blank', 'colour': '#7ebfc2'}, 140: {'group': 0, 'well_id': 'L9', 'state': 'blank', 'colour': '#7ebfc2'}, 141: {'group': 0, 'well_id': 'M9', 'state': 'blank', 'colour': '#7ebfc2'}, 142: {'group': 0, 'well_id': 'N9', 'state': 'blank', 'colour': '#7ebfc2'}, 143: {'group': 0, 'well_id': 'O9', 'state': 'blank', 'colour': '#7ebfc2'}, 144: {'group': 0, 'well_id': 'P9', 'state': 'blank', 'colour': '#7ebfc2'}, 145: {'group': 0, 'well_id': 'A10', 'state': 'blank', 'colour': '#7ebfc2'}, 146: {'group': '6', 'well_id': 'B10', 'state': 'positive', 'colour': '#09ca0e'}, 147: {'group': 0, 'well_id': 'C10', 'state': 'blank', 'colour': '#7ebfc2'}, 148: {'group': 0, 'well_id': 'D10', 'state': 'blank', 'colour': '#7ebfc2'}, 149: {'group': 0, 'well_id': 'E10', 'state': 'negative', 'colour': '#d20000'}, 150: {'group': 0, 'well_id': 'F10', 'state': 'blank', 'colour': '#7ebfc2'}, 151: {'group': 0, 'well_id': 'G10', 'state': 'blank', 'colour': '#7ebfc2'}, 152: {'group': 0, 'well_id': 'H10', 'state': 'blank', 'colour': '#7ebfc2'}, 153: {'group': '4', 'well_id': 'I10', 'state': 'positive', 'colour': '#09ca0e'}, 154: {'group': 0, 'well_id': 'J10', 'state': 'blank', 'colour': '#7ebfc2'}, 155: {'group': 0, 'well_id': 'K10', 'state': 'blank', 'colour': '#7ebfc2'}, 156: {'group': 0, 'well_id': 'L10', 'state': 'blank', 'colour': '#7ebfc2'}, 157: {'group': 0, 'well_id': 'M10', 'state': 'blank', 'colour': '#7ebfc2'}, 158: {'group': 0, 'well_id': 'N10', 'state': 'blank', 'colour': '#7ebfc2'}, 159: {'group': 0, 'well_id': 'O10', 'state': 'blank', 'colour': '#7ebfc2'}, 160: {'group': 0, 'well_id': 'P10', 'state': 'blank', 'colour': '#7ebfc2'}, 161: {'group': 0, 'well_id': 'A11', 'state': 'blank', 'colour': '#7ebfc2'}, 162: {'group': 0, 'well_id': 'B11', 'state': 'blank', 'colour': '#7ebfc2'}, 163: {'group': 0, 'well_id': 'C11', 'state': 'blank', 'colour': '#7ebfc2'}, 164: {'group': 0, 'well_id': 'D11', 'state': 'blank', 'colour': '#7ebfc2'}, 165: {'group': 0, 'well_id': 'E11', 'state': 'blank', 'colour': '#7ebfc2'}, 166: {'group': 0, 'well_id': 'F11', 'state': 'blank', 'colour': '#7ebfc2'}, 167: {'group': 0, 'well_id': 'G11', 'state': 'blank', 'colour': '#7ebfc2'}, 168: {'group': 0, 'well_id': 'H11', 'state': 'blank', 'colour': '#7ebfc2'}, 169: {'group': 0, 'well_id': 'I11', 'state': 'blank', 'colour': '#7ebfc2'}, 170: {'group': 0, 'well_id': 'J11', 'state': 'blank', 'colour': '#7ebfc2'}, 171: {'group': 0, 'well_id': 'K11', 'state': 'blank', 'colour': '#7ebfc2'}, 172: {'group': 0, 'well_id': 'L11', 'state': 'blank', 'colour': '#7ebfc2'}, 173: {'group': 0, 'well_id': 'M11', 'state': 'blank', 'colour': '#7ebfc2'}, 174: {'group': 0, 'well_id': 'N11', 'state': 'blank', 'colour': '#7ebfc2'}, 175: {'group': 0, 'well_id': 'O11', 'state': 'blank', 'colour': '#7ebfc2'}, 176: {'group': 0, 'well_id': 'P11', 'state': 'blank', 'colour': '#7ebfc2'}, 177: {'group': 0, 'well_id': 'A12', 'state': 'blank', 'colour': '#7ebfc2'}, 178: {'group': 0, 'well_id': 'B12', 'state': 'blank', 'colour': '#7ebfc2'}, 179: {'group': 0, 'well_id': 'C12', 'state': 'blank', 'colour': '#7ebfc2'}, 180: {'group': 0, 'well_id': 'D12', 'state': 'blank', 'colour': '#7ebfc2'}, 181: {'group': 0, 'well_id': 'E12', 'state': 'blank', 'colour': '#7ebfc2'}, 182: {'group': 0, 'well_id': 'F12', 'state': 'blank', 'colour': '#7ebfc2'}, 183: {'group': 0, 'well_id': 'G12', 'state': 'blank', 'colour': '#7ebfc2'}, 184: {'group': 0, 'well_id': 'H12', 'state': 'blank', 'colour': '#7ebfc2'}, 185: {'group': 0, 'well_id': 'I12', 'state': 'blank', 'colour': '#7ebfc2'}, 186: {'group': 0, 'well_id': 'J12', 'state': 'blank', 'colour': '#7ebfc2'}, 187: {'group': 0, 'well_id': 'K12', 'state': 'blank', 'colour': '#7ebfc2'}, 188: {'group': 0, 'well_id': 'L12', 'state': 'blank', 'colour': '#7ebfc2'}, 189: {'group': 0, 'well_id': 'M12', 'state': 'negative', 'colour': '#d20000'}, 190: {'group': 0, 'well_id': 'N12', 'state': 'blank', 'colour': '#7ebfc2'}, 191: {'group': 0, 'well_id': 'O12', 'state': 'blank', 'colour': '#7ebfc2'}, 192: {'group': 0, 'well_id': 'P12', 'state': 'blank', 'colour': '#7ebfc2'}, 193: {'group': 0, 'well_id': 'A13', 'state': 'blank', 'colour': '#7ebfc2'}, 194: {'group': 0, 'well_id': 'B13', 'state': 'blank', 'colour': '#7ebfc2'}, 195: {'group': 0, 'well_id': 'C13', 'state': 'blank', 'colour': '#7ebfc2'}, 196: {'group': 0, 'well_id': 'D13', 'state': 'blank', 'colour': '#7ebfc2'}, 197: {'group': 0, 'well_id': 'E13', 'state': 'blank', 'colour': '#7ebfc2'}, 198: {'group': 0, 'well_id': 'F13', 'state': 'blank', 'colour': '#7ebfc2'}, 199: {'group': 0, 'well_id': 'G13', 'state': 'blank', 'colour': '#7ebfc2'}, 200: {'group': 0, 'well_id': 'H13', 'state': 'blank', 'colour': '#7ebfc2'}, 201: {'group': 0, 'well_id': 'I13', 'state': 'blank', 'colour': '#7ebfc2'}, 202: {'group': 0, 'well_id': 'J13', 'state': 'blank', 'colour': '#7ebfc2'}, 203: {'group': 0, 'well_id': 'K13', 'state': 'blank', 'colour': '#7ebfc2'}, 204: {'group': 0, 'well_id': 'L13', 'state': 'blank', 'colour': '#7ebfc2'}, 205: {'group': 0, 'well_id': 'M13', 'state': 'blank', 'colour': '#7ebfc2'}, 206: {'group': 0, 'well_id': 'N13', 'state': 'blank', 'colour': '#7ebfc2'}, 207: {'group': 0, 'well_id': 'O13', 'state': 'blank', 'colour': '#7ebfc2'}, 208: {'group': 0, 'well_id': 'P13', 'state': 'blank', 'colour': '#7ebfc2'}, 209: {'group': 0, 'well_id': 'A14', 'state': 'blank', 'colour': '#7ebfc2'}, 210: {'group': 0, 'well_id': 'B14', 'state': 'blank', 'colour': '#7ebfc2'}, 211: {'group': 0, 'well_id': 'C14', 'state': 'blank', 'colour': '#7ebfc2'}, 212: {'group': 0, 'well_id': 'D14', 'state': 'blank', 'colour': '#7ebfc2'}, 213: {'group': 0, 'well_id': 'E14', 'state': 'blank', 'colour': '#7ebfc2'}, 214: {'group': 0, 'well_id': 'F14', 'state': 'blank', 'colour': '#7ebfc2'}, 215: {'group': 0, 'well_id': 'G14', 'state': 'blank', 'colour': '#7ebfc2'}, 216: {'group': 0, 'well_id': 'H14', 'state': 'blank', 'colour': '#7ebfc2'}, 217: {'group': 0, 'well_id': 'I14', 'state': 'blank', 'colour': '#7ebfc2'}, 218: {'group': 0, 'well_id': 'J14', 'state': 'blank', 'colour': '#7ebfc2'}, 219: {'group': 0, 'well_id': 'K14', 'state': 'blank', 'colour': '#7ebfc2'}, 220: {'group': 0, 'well_id': 'L14', 'state': 'blank', 'colour': '#7ebfc2'}, 221: {'group': 0, 'well_id': 'M14', 'state': 'blank', 'colour': '#7ebfc2'}, 222: {'group': 0, 'well_id': 'N14', 'state': 'blank', 'colour': '#7ebfc2'}, 223: {'group': 0, 'well_id': 'O14', 'state': 'blank', 'colour': '#7ebfc2'}, 224: {'group': 0, 'well_id': 'P14', 'state': 'blank', 'colour': '#7ebfc2'}, 225: {'group': 0, 'well_id': 'A15', 'state': 'blank', 'colour': '#7ebfc2'}, 226: {'group': 0, 'well_id': 'B15', 'state': 'blank', 'colour': '#7ebfc2'}, 227: {'group': 0, 'well_id': 'C15', 'state': 'blank', 'colour': '#7ebfc2'}, 228: {'group': 0, 'well_id': 'D15', 'state': 'blank', 'colour': '#7ebfc2'}, 229: {'group': '3', 'well_id': 'E15', 'state': 'positive', 'colour': '#09ca0e'}, 230: {'group': 0, 'well_id': 'F15', 'state': 'blank', 'colour': '#7ebfc2'}, 231: {'group': 0, 'well_id': 'G15', 'state': 'blank', 'colour': '#7ebfc2'}, 232: {'group': 0, 'well_id': 'H15', 'state': 'blank', 'colour': '#7ebfc2'}, 233: {'group': 0, 'well_id': 'I15', 'state': 'blank', 'colour': '#7ebfc2'}, 234: {'group': 0, 'well_id': 'J15', 'state': 'blank', 'colour': '#7ebfc2'}, 235: {'group': 0, 'well_id': 'K15', 'state': 'blank', 'colour': '#7ebfc2'}, 236: {'group': 0, 'well_id': 'L15', 'state': 'blank', 'colour': '#7ebfc2'}, 237: {'group': 0, 'well_id': 'M15', 'state': 'blank', 'colour': '#7ebfc2'}, 238: {'group': 0, 'well_id': 'N15', 'state': 'blank', 'colour': '#7ebfc2'}, 239: {'group': 0, 'well_id': 'O15', 'state': 'blank', 'colour': '#7ebfc2'}, 240: {'group': 0, 'well_id': 'P15', 'state': 'blank', 'colour': '#7ebfc2'}, 241: {'group': 0, 'well_id': 'A16', 'state': 'blank', 'colour': '#7ebfc2'}, 242: {'group': 0, 'well_id': 'B16', 'state': 'blank', 'colour': '#7ebfc2'}, 243: {'group': 0, 'well_id': 'C16', 'state': 'blank', 'colour': '#7ebfc2'}, 244: {'group': 0, 'well_id': 'D16', 'state': 'blank', 'colour': '#7ebfc2'}, 245: {'group': 0, 'well_id': 'E16', 'state': 'blank', 'colour': '#7ebfc2'}, 246: {'group': 0, 'well_id': 'F16', 'state': 'blank', 'colour': '#7ebfc2'}, 247: {'group': 0, 'well_id': 'G16', 'state': 'negative', 'colour': '#d20000'}, 248: {'group': 0, 'well_id': 'H16', 'state': 'blank', 'colour': '#7ebfc2'}, 249: {'group': 0, 'well_id': 'I16', 'state': 'blank', 'colour': '#7ebfc2'}, 250: {'group': '8', 'well_id': 'J16', 'state': 'positive', 'colour': '#09ca0e'}, 251: {'group': 0, 'well_id': 'K16', 'state': 'blank', 'colour': '#7ebfc2'}, 252: {'group': 0, 'well_id': 'L16', 'state': 'blank', 'colour': '#7ebfc2'}, 253: {'group': 0, 'well_id': 'M16', 'state': 'blank', 'colour': '#7ebfc2'}, 254: {'group': 0, 'well_id': 'N16', 'state': 'blank', 'colour': '#7ebfc2'}, 255: {'group': 0, 'well_id': 'O16', 'state': 'blank', 'colour': '#7ebfc2'}, 256: {'group': 0, 'well_id': 'P16', 'state': 'blank', 'colour': '#7ebfc2'}, 257: {'group': 0, 'well_id': 'A17', 'state': 'blank', 'colour': '#7ebfc2'}, 258: {'group': 0, 'well_id': 'B17', 'state': 'blank', 'colour': '#7ebfc2'}, 259: {'group': 0, 'well_id': 'C17', 'state': 'blank', 'colour': '#7ebfc2'}, 260: {'group': 0, 'well_id': 'D17', 'state': 'blank', 'colour': '#7ebfc2'}, 261: {'group': 0, 'well_id': 'E17', 'state': 'blank', 'colour': '#7ebfc2'}, 262: {'group': 0, 'well_id': 'F17', 'state': 'blank', 'colour': '#7ebfc2'}, 263: {'group': 0, 'well_id': 'G17', 'state': 'blank', 'colour': '#7ebfc2'}, 264: {'group': 0, 'well_id': 'H17', 'state': 'blank', 'colour': '#7ebfc2'}, 265: {'group': 0, 'well_id': 'I17', 'state': 'blank', 'colour': '#7ebfc2'}, 266: {'group': 0, 'well_id': 'J17', 'state': 'blank', 'colour': '#7ebfc2'}, 267: {'group': 0, 'well_id': 'K17', 'state': 'blank', 'colour': '#7ebfc2'}, 268: {'group': 0, 'well_id': 'L17', 'state': 'blank', 'colour': '#7ebfc2'}, 269: {'group': 0, 'well_id': 'M17', 'state': 'blank', 'colour': '#7ebfc2'}, 270: {'group': '7', 'well_id': 'N17', 'state': 'positive', 'colour': '#09ca0e'}, 271: {'group': 0, 'well_id': 'O17', 'state': 'blank', 'colour': '#7ebfc2'}, 272: {'group': 0, 'well_id': 'P17', 'state': 'blank', 'colour': '#7ebfc2'}, 273: {'group': 0, 'well_id': 'A18', 'state': 'blank', 'colour': '#7ebfc2'}, 274: {'group': 0, 'well_id': 'B18', 'state': 'blank', 'colour': '#7ebfc2'}, 275: {'group': '11', 'well_id': 'C18', 'state': 'positive', 'colour': '#09ca0e'}, 276: {'group': 0, 'well_id': 'D18', 'state': 'blank', 'colour': '#7ebfc2'}, 277: {'group': 0, 'well_id': 'E18', 'state': 'blank', 'colour': '#7ebfc2'}, 278: {'group': 0, 'well_id': 'F18', 'state': 'blank', 'colour': '#7ebfc2'}, 279: {'group': 0, 'well_id': 'G18', 'state': 'blank', 'colour': '#7ebfc2'}, 280: {'group': 0, 'well_id': 'H18', 'state': 'blank', 'colour': '#7ebfc2'}, 281: {'group': 0, 'well_id': 'I18', 'state': 'blank', 'colour': '#7ebfc2'}, 282: {'group': 0, 'well_id': 'J18', 'state': 'blank', 'colour': '#7ebfc2'}, 283: {'group': 0, 'well_id': 'K18', 'state': 'blank', 'colour': '#7ebfc2'}, 284: {'group': 0, 'well_id': 'L18', 'state': 'blank', 'colour': '#7ebfc2'}, 285: {'group': 0, 'well_id': 'M18', 'state': 'blank', 'colour': '#7ebfc2'}, 286: {'group': 0, 'well_id': 'N18', 'state': 'blank', 'colour': '#7ebfc2'}, 287: {'group': 0, 'well_id': 'O18', 'state': 'blank', 'colour': '#7ebfc2'}, 288: {'group': 0, 'well_id': 'P18', 'state': 'blank', 'colour': '#7ebfc2'}, 289: {'group': 0, 'well_id': 'A19', 'state': 'blank', 'colour': '#7ebfc2'}, 290: {'group': 0, 'well_id': 'B19', 'state': 'blank', 'colour': '#7ebfc2'}, 291: {'group': 0, 'well_id': 'C19', 'state': 'blank', 'colour': '#7ebfc2'}, 292: {'group': 0, 'well_id': 'D19', 'state': 'blank', 'colour': '#7ebfc2'}, 293: {'group': 0, 'well_id': 'E19', 'state': 'blank', 'colour': '#7ebfc2'}, 294: {'group': 0, 'well_id': 'F19', 'state': 'blank', 'colour': '#7ebfc2'}, 295: {'group': 0, 'well_id': 'G19', 'state': 'blank', 'colour': '#7ebfc2'}, 296: {'group': 0, 'well_id': 'H19', 'state': 'blank', 'colour': '#7ebfc2'}, 297: {'group': 0, 'well_id': 'I19', 'state': 'blank', 'colour': '#7ebfc2'}, 298: {'group': 0, 'well_id': 'J19', 'state': 'blank', 'colour': '#7ebfc2'}, 299: {'group': 0, 'well_id': 'K19', 'state': 'blank', 'colour': '#7ebfc2'}, 300: {'group': 0, 'well_id': 'L19', 'state': 'blank', 'colour': '#7ebfc2'}, 301: {'group': 0, 'well_id': 'M19', 'state': 'blank', 'colour': '#7ebfc2'}, 302: {'group': 0, 'well_id': 'N19', 'state': 'blank', 'colour': '#7ebfc2'}, 303: {'group': 0, 'well_id': 'O19', 'state': 'blank', 'colour': '#7ebfc2'}, 304: {'group': 0, 'well_id': 'P19', 'state': 'blank', 'colour': '#7ebfc2'}, 305: {'group': 0, 'well_id': 'A20', 'state': 'blank', 'colour': '#7ebfc2'}, 306: {'group': 0, 'well_id': 'B20', 'state': 'blank', 'colour': '#7ebfc2'}, 307: {'group': 0, 'well_id': 'C20', 'state': 'blank', 'colour': '#7ebfc2'}, 308: {'group': 0, 'well_id': 'D20', 'state': 'blank', 'colour': '#7ebfc2'}, 309: {'group': 0, 'well_id': 'E20', 'state': 'blank', 'colour': '#7ebfc2'}, 310: {'group': 0, 'well_id': 'F20', 'state': 'blank', 'colour': '#7ebfc2'}, 311: {'group': 0, 'well_id': 'G20', 'state': 'blank', 'colour': '#7ebfc2'}, 312: {'group': 0, 'well_id': 'H20', 'state': 'blank', 'colour': '#7ebfc2'}, 313: {'group': 0, 'well_id': 'I20', 'state': 'blank', 'colour': '#7ebfc2'}, 314: {'group': 0, 'well_id': 'J20', 'state': 'blank', 'colour': '#7ebfc2'}, 315: {'group': 0, 'well_id': 'K20', 'state': 'blank', 'colour': '#7ebfc2'}, 316: {'group': 0, 'well_id': 'L20', 'state': 'blank', 'colour': '#7ebfc2'}, 317: {'group': 0, 'well_id': 'M20', 'state': 'blank', 'colour': '#7ebfc2'}, 318: {'group': 0, 'well_id': 'N20', 'state': 'blank', 'colour': '#7ebfc2'}, 319: {'group': 0, 'well_id': 'O20', 'state': 'blank', 'colour': '#7ebfc2'}, 320: {'group': 0, 'well_id': 'P20', 'state': 'blank', 'colour': '#7ebfc2'}, 321: {'group': 0, 'well_id': 'A21', 'state': 'blank', 'colour': '#7ebfc2'}, 322: {'group': 0, 'well_id': 'B21', 'state': 'blank', 'colour': '#7ebfc2'}, 323: {'group': 0, 'well_id': 'C21', 'state': 'blank', 'colour': '#7ebfc2'}, 324: {'group': 0, 'well_id': 'D21', 'state': 'blank', 'colour': '#7ebfc2'}, 325: {'group': 0, 'well_id': 'E21', 'state': 'blank', 'colour': '#7ebfc2'}, 326: {'group': '9', 'well_id': 'F21', 'state': 'positive', 'colour': '#09ca0e'}, 327: {'group': 0, 'well_id': 'G21', 'state': 'blank', 'colour': '#7ebfc2'}, 328: {'group': 0, 'well_id': 'H21', 'state': 'blank', 'colour': '#7ebfc2'}, 329: {'group': 0, 'well_id': 'I21', 'state': 'blank', 'colour': '#7ebfc2'}, 330: {'group': 0, 'well_id': 'J21', 'state': 'blank', 'colour': '#7ebfc2'}, 331: {'group': 0, 'well_id': 'K21', 'state': 'blank', 'colour': '#7ebfc2'}, 332: {'group': 0, 'well_id': 'L21', 'state': 'blank', 'colour': '#7ebfc2'}, 333: {'group': 0, 'well_id': 'M21', 'state': 'blank', 'colour': '#7ebfc2'}, 334: {'group': 0, 'well_id': 'N21', 'state': 'blank', 'colour': '#7ebfc2'}, 335: {'group': 0, 'well_id': 'O21', 'state': 'blank', 'colour': '#7ebfc2'}, 336: {'group': 0, 'well_id': 'P21', 'state': 'blank', 'colour': '#7ebfc2'}, 337: {'group': 0, 'well_id': 'A22', 'state': 'blank', 'colour': '#7ebfc2'}, 338: {'group': 0, 'well_id': 'B22', 'state': 'blank', 'colour': '#7ebfc2'}, 339: {'group': 0, 'well_id': 'C22', 'state': 'negative', 'colour': '#d20000'}, 340: {'group': 0, 'well_id': 'D22', 'state': 'blank', 'colour': '#7ebfc2'}, 341: {'group': 0, 'well_id': 'E22', 'state': 'blank', 'colour': '#7ebfc2'}, 342: {'group': 0, 'well_id': 'F22', 'state': 'blank', 'colour': '#7ebfc2'}, 343: {'group': 0, 'well_id': 'G22', 'state': 'blank', 'colour': '#7ebfc2'}, 344: {'group': 0, 'well_id': 'H22', 'state': 'blank', 'colour': '#7ebfc2'}, 345: {'group': 0, 'well_id': 'I22', 'state': 'negative', 'colour': '#d20000'}, 346: {'group': 0, 'well_id': 'J22', 'state': 'blank', 'colour': '#7ebfc2'}, 347: {'group': 0, 'well_id': 'K22', 'state': 'blank', 'colour': '#7ebfc2'}, 348: {'group': 0, 'well_id': 'L22', 'state': 'blank', 'colour': '#7ebfc2'}, 349: {'group': 0, 'well_id': 'M22', 'state': 'blank', 'colour': '#7ebfc2'}, 350: {'group': 0, 'well_id': 'N22', 'state': 'blank', 'colour': '#7ebfc2'}, 351: {'group': 0, 'well_id': 'O22', 'state': 'blank', 'colour': '#7ebfc2'}, 352: {'group': 0, 'well_id': 'P22', 'state': 'blank', 'colour': '#7ebfc2'}, 353: {'group': 0, 'well_id': 'A23', 'state': 'blank', 'colour': '#7ebfc2'}, 354: {'group': 0, 'well_id': 'B23', 'state': 'blank', 'colour': '#7ebfc2'}, 355: {'group': 0, 'well_id': 'C23', 'state': 'blank', 'colour': '#7ebfc2'}, 356: {'group': 0, 'well_id': 'D23', 'state': 'blank', 'colour': '#7ebfc2'}, 357: {'group': 0, 'well_id': 'E23', 'state': 'blank', 'colour': '#7ebfc2'}, 358: {'group': 0, 'well_id': 'F23', 'state': 'blank', 'colour': '#7ebfc2'}, 359: {'group': 0, 'well_id': 'G23', 'state': 'blank', 'colour': '#7ebfc2'}, 360: {'group': 0, 'well_id': 'H23', 'state': 'blank', 'colour': '#7ebfc2'}, 361: {'group': 0, 'well_id': 'I23', 'state': 'blank', 'colour': '#7ebfc2'}, 362: {'group': 0, 'well_id': 'J23', 'state': 'blank', 'colour': '#7ebfc2'}, 363: {'group': 0, 'well_id': 'K23', 'state': 'blank', 'colour': '#7ebfc2'}, 364: {'group': 0, 'well_id': 'L23', 'state': 'blank', 'colour': '#7ebfc2'}, 365: {'group': 0, 'well_id': 'M23', 'state': 'blank', 'colour': '#7ebfc2'}, 366: {'group': 0, 'well_id': 'N23', 'state': 'blank', 'colour': '#7ebfc2'}, 367: {'group': 0, 'well_id': 'O23', 'state': 'blank', 'colour': '#7ebfc2'}, 368: {'group': 0, 'well_id': 'P23', 'state': 'blank', 'colour': '#7ebfc2'}, 369: {'group': 0, 'well_id': 'A24', 'state': 'blank', 'colour': '#7ebfc2'}, 370: {'group': 0, 'well_id': 'B24', 'state': 'blank', 'colour': '#7ebfc2'}, 371: {'group': 0, 'well_id': 'C24', 'state': 'blank', 'colour': '#7ebfc2'}, 372: {'group': 0, 'well_id': 'D24', 'state': 'blank', 'colour': '#7ebfc2'}, 373: {'group': 0, 'well_id': 'E24', 'state': 'blank', 'colour': '#7ebfc2'}, 374: {'group': 0, 'well_id': 'F24', 'state': 'blank', 'colour': '#7ebfc2'}, 375: {'group': 0, 'well_id': 'G24', 'state': 'blank', 'colour': '#7ebfc2'}, 376: {'group': 0, 'well_id': 'H24', 'state': 'blank', 'colour': '#7ebfc2'}, 377: {'group': 0, 'well_id': 'I24', 'state': 'blank', 'colour': '#7ebfc2'}, 378: {'group': 0, 'well_id': 'J24', 'state': 'blank', 'colour': '#7ebfc2'}, 379: {'group': 0, 'well_id': 'K24', 'state': 'blank', 'colour': '#7ebfc2'}, 380: {'group': 0, 'well_id': 'L24', 'state': 'blank', 'colour': '#7ebfc2'}, 381: {'group': '2', 'well_id': 'M24', 'state': 'positive', 'colour': '#09ca0e'}, 382: {'group': 0, 'well_id': 'N24', 'state': 'blank', 'colour': '#7ebfc2'}, 383: {'group': 0, 'well_id': 'O24', 'state': 'blank', 'colour': '#7ebfc2'}, 384: {'group': 0, 'well_id': 'P24', 'state': 'blank', 'colour': '#7ebfc2'}}
    # all_data, well_col_row, well_type, barcode, date = multiscan_data_dict(file, plate_layout)
    # print(f" all_data - {all_data}")
    # print(f" well_col_row - {well_col_row}")
    # print(f" well_type - {well_type}")
    # print(f" barcode - {barcode}")
    # print(f" date - {date}")
    file = r"C:\Users\phch\OneDrive - Danmarks Tekniske Universitet\Mapper\Python_data\procul\procul_001-EPSON Perfection V19-Fn-240403-1548-Avelina calibration (Emilie 21-03-24).txt"
    csv_converter(file)
