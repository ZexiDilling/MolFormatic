import copy
# from math import floor
from pathlib import Path
from openpyxl import Workbook, load_workbook
#
# import PySimpleGUI as sg
#
# from database_handler import DataBaseFunctions
# from gui_layout import GUILayout
#
# from gui_plate_drawing_functions import on_up, save_layout, delete_layout, rename_layout, export_layout, draw_layout, \
#     _update_dose_tool, on_move, dose_colouring, dose_sample_amount, dose_dilution_replicates
# from heatmap import Heatmap
# from gui_functions import get_plate_layout, draw_plate, plate_layout_to_excel, update_database, \
#     delete_records_from_database, rename_record_in_the_database
# from info import clm_to_row_96, clm_to_row_384, clm_to_row_1536, row_to_clm_96, row_to_clm_384, row_to_clm_1536
# from plate_formatting import plate_layout_re_formate
#
#
# class Tester:
#     def __init__(self, config, plate_list):
#         self.config = config
#         self.standard_size = 20
#         self.colour_size = 5
#         self.button_height = 1
#         self.tab_colour = config["GUI"]["tab_colour"]
#         self.sample_style = ["Single Point", "Duplicate", "Triplicate", "Custom", "Dose Response"]
#         self.analyse_style = ["Single", "Dose Response"]
#         self.plate_list = plate_list
#         self.lable_style = "solid"
#         self.show_input_style = "sunken"
#
#
#     def setup_1_plate_layout(self):
#         """
#
#         :return: A layour for the plate-module in the top box
#         :rtype: list
#         """
#         color_select = {}
#         for keys in list(self.config["plate_colouring"].keys()):
#             color_select[keys] = self.config["plate_colouring"][keys]
#
#         plate_type = ["plate_96", "plate_384", "plate_1536"]
#         sample_type = self.sample_style
#
#         col_graph = sg.Frame("Plate Layout", [[
#             sg.Column([
#                 [sg.Graph(canvas_size=(500, 350), graph_bottom_left=(0, 0), graph_top_right=(500, 350),
#                       background_color='grey', key="-RECT_BIO_CANVAS-", enable_events=True, drag_submits=True,
#                       motion_events=True)],
#                 [sg.DropDown(values=plate_type, default_value=plate_type[1], key="-PLATE-"),
#                  sg.B("Draw Plate", key="-DRAW-"),
#                  # sg.B("Add sample layout", key="-DRAW_SAMPLE_LAYOUT-"),
#                  sg.Text(key="-CANVAS_INFO_WELL-"),
#                  sg.Text(key="-CANVAS_INFO_GROUP-"),
#                  sg.Text(key="-CANVAS_INFO_COUNT-")]
#             ])
#         ]])
#
#         coloring_tab = sg.Tab("State", [[
#             sg.Column([
#                 [sg.Radio(f"Select Sample", 1, key="-RECT_SAMPLES-", size=self.standard_size, enable_events=True,
#                           default=True),
#                  sg.T(background_color=self.config["plate_colouring"]["sample"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_SAMPLE-", relief="groove")],
#                 [sg.Radio(f"Select Blank", 1, key="-RECT_BLANK-", size=self.standard_size, enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["blank"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_BLANK-", relief="groove")],
#                 [sg.Radio(f"Select Max Signal", 1, key="-RECT_MAX-", size=self.standard_size, enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["max"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_NAX-", relief="groove")],
#                 [sg.Radio(f"Select Minimum Signal", 1, key="-RECT_MIN-", size=self.standard_size,
#                           enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["minimum"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_MINIMUM-", relief="groove")],
#                 [sg.Radio(f"Select Positive Control", 1, key="-RECT_POS-", size=self.standard_size,
#                           enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["positive"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_POSITIVE-", relief="groove")],
#                 [sg.Radio(f"Select Negative Control", 1, key="-RECT_NEG-", size=self.standard_size,
#                           enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["negative"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_NEGATIVE-", relief="groove")],
#                 [sg.Radio(f"Select Empty", 1, key="-RECT_EMPTY-", size=self.standard_size, enable_events=True),
#                  sg.T(background_color=self.config["plate_colouring"]["empty"], size=self.colour_size,
#                       key="-PLATE_LAYOUT_COLOUR_BOX_EMPTY-", relief="groove")],
#                 [sg.Radio(f"Colour", 1, key="-COLOUR-", enable_events=True, size=self.standard_size),
#                  sg.ColorChooserButton("Colour", key="-PLATE_LAYOUT_COLOUR_CHOSE-",
#                                        target="-PLATE_LAYOUT_COLOUR_CHOSE_TARGET-"),
#                  sg.Input(key="-PLATE_LAYOUT_COLOUR_CHOSE_TARGET-", visible=False, enable_events=True, disabled=True,
#                           default_text="#ffffff")],
#                 # [sg.Radio('Erase', 1, key='-ERASE-', enable_events=True)],
#                 # [sg.Radio('Move Stuff', 1, key='-MOVE-', enable_events=True)],
#
#             ])]])
#
#         grp_tab = sg.Tab("Dose Response", [[
#             sg.Column([
#                 [sg.T("Amount of available sample spots:"),
#                  sg.T("", key="-SAMPLE_SPOTS-")],
#                 [sg.T("Sample Amount", size=self.standard_size),
#                  sg.Input("", key="-DOSE_SAMPLE_AMOUNT-", size=5, enable_events=True)],
#                 [sg.T("Replicate", size=self.standard_size),
#                  sg.Input("", key="-DOSE_REPLICATES_AMOUNT-", size=5, enable_events=True)],
#                 [sg.T("Empty Sample Spots", size=self.standard_size),
#                  sg.Input("", key="-DOSE_EMPTY_SAMPLE_SPOTS-", size=5, enable_events=True)],
#                 [sg.Checkbox("Equal split", key="-EQUAL_SPLIT-", default=True,
#                              tooltip="If this is true, the amount of replicates per sample will be the same")],
#                 [sg.T("", size=self.standard_size)],
#                 [sg.DropDown(values=[1], key="-SAMPLE_CHOOSER_DROPDOWN-", default_value=1, enable_events=True)],
#                 [sg.Radio(f"Horizontal", 2, key="-AUTO_HORIZONTAL-", size=10, enable_events=True,
#                           default=True),
#                  sg.Radio(f"Vertical", 2, key="-AUTO_VERTICAL-", size=10, enable_events=True)],
#                 [sg.ColorChooserButton("First Colour", button_color=self.config["plate_colouring"]["dose_low"],
#                                        target="-DOSE_COLOUR_LOW-", key="-DOSE_COLOUR_BUTTON_LOW-"),
#                  sg.Input(self.config["plate_colouring"]["dose_low"], key="-DOSE_COLOUR_LOW-", size=5,
#                           enable_events=True, visible=False),
#                  sg.ColorChooserButton("Last Colour", button_color=self.config["plate_colouring"]["dose_high"],
#                                        target="-DOSE_COLOUR_LOW-", key="-DOSE_COLOUR_BUTTON_HIGH-"),
#                  sg.Input(self.config["plate_colouring"]["dose_high"], key="-DOSE_COLOUR_HIGH-", size=5,
#                           enable_events=True, visible=False)],
#                 [sg.Radio(f"", 1, key="-RECT_DOSE-", size=15, enable_events=True, visible=False)]
#             ])
#         ]])
#
#         draw_tab_group = [coloring_tab, grp_tab]
#         tab_draw_group = sg.TabGroup([draw_tab_group], selected_title_color=self.tab_colour,
#                                      key="-PLATE_LAYOUT_DRAW_GROUPS-", enable_events=True)
#         draw_options = sg.Frame("Options", [[
#             sg.Column([
#                 [tab_draw_group],
#                 [sg.Checkbox("Use Archive", default=False, key="-ARCHIVE-", size=floor(self.standard_size / 2)),
#                  sg.DropDown(sample_type, key="-RECT_SAMPLE_TYPE-", default_value=sample_type[0], visible=True,
#                              enable_events=True)],
#                 [sg.T("Main", size=5),
#                  sg.DropDown(sorted(self.plate_list), key="-ARCHIVE_PLATES-", size=self.standard_size,
#                              enable_events=True)],
#                 [sg.T("Sub", size=5),
#                  sg.DropDown([], key="-ARCHIVE_PLATES_SUB-", size=self.standard_size)],
#                 [sg.Button("Save", key="-SAVE_LAYOUT-"),
#                  sg.Button("Rename", key="-RENAME_LAYOUT-"),
#                  sg.Button("Delete", key="-DELETE_LAYOUT-"),
#                  sg.Button("Export", key="-EXPORT_LAYOUT-")],
#             ])
#         ]])
#
#         layout = [[col_graph, draw_options]]
#
#         return sg.Window("Samples", layout, finalize=True, resizable=True)
#
#
# def run(config, plate_list, archive_plates_dict):
#
#     gui = GUILayout(config, plate_list)
#     layout = gui.setup_1_plate_layout()
#     window = sg.Window("Samples", layout, finalize=True, resizable=True)
#     graph_bio_exp = well_dict_bio_info = plate_bio_info = None
#     graph_plate = window["-RECT_BIO_CANVAS-"]
#     dragging = False
#     temp_selector = False
#     plate_active = False
#     temp_draw_tool = "sample"
#     color_select = {}
#     for keys in list(config["plate_colouring"].keys()):
#         color_select[keys] = config["plate_colouring"][keys]
#     well_dict = {}
#     start_point = end_point = prior_rect = temp_tool = None
#
#     clm_to_row_converter = {"plate_96": clm_to_row_96, "plate_384": clm_to_row_384, "plate_1536": clm_to_row_1536}
#     row_to_clm_converter = {"plate_96": row_to_clm_96, "plate_384": row_to_clm_384, "plate_1536": row_to_clm_1536}
#     plate_type_count = {"plate_96": 96, "plate_384": 384, "plate_1536": 1536}
#     total_sample_spots = 0
#     temp_sample_amount = 0
#     x = y = min_x = max_x = min_y = max_y = plate_type = None
#     dose_colour_dict = {}
#     draw_tool_dict = {
#         "-RECT_SAMPLES-": "sample",
#         "-RECT_BLANK-": "blank",
#         "-RECT_MAX-": "max",
#         "-RECT_MIN-": "minimum",
#         "-RECT_NEG-": "negative",
#         "-RECT_POS-": "positive",
#         "-RECT_EMPTY-": "empty",
#         "-COLOUR-": "paint",
#         "-RECT_DOSE-": "dose"
#     }
#
#     while True:
#         event, values = window.read()
#
#         if event == sg.WIN_CLOSED:
#             window.close()
#             break
#
#         # Tabs interactions
#         if event == "-PLATE_LAYOUT_DRAW_GROUPS-":
#             if values["-PLATE_LAYOUT_DRAW_GROUPS-"] == "Dose Response":
#                 for tools in draw_tool_dict:
#                     if values[tools]:
#                         temp_draw_tool_tracker = tools
#
#                 window["-RECT_DOSE-"].update(value=True)
#                 total_sample_spots = 0
#                 for wells in well_dict:
#                     if well_dict[wells]["state"] == "sample":
#                         total_sample_spots += 1
#
#                 window["-SAMPLE_SPOTS-"].update(value=total_sample_spots)
#             elif values["-PLATE_LAYOUT_DRAW_GROUPS-"] == "State":
#                 window[temp_draw_tool_tracker].update(value=True)
#
#         # Dose Response Tab
#         if event == "-DOSE_SAMPLE_AMOUNT-" and values["-EQUAL_SPLIT-"]:
#             dose_colour_dict = dose_sample_amount(window, event, values, total_sample_spots, dose_colour_dict)
#
#         if event == "-DOSE_DILUTIONS-" or event == "-DOSE_REPLICATES-":
#             dose_colour_dict = dose_dilution_replicates(window, event, values, total_sample_spots, dose_colour_dict)
#
#         if event == "-DOSE_COLOUR_LOW-":
#             dose_colour_dict = dose_colouring(window, event, values, temp_sample_amount,
#                                               "-DOSE_COLOUR_BUTTON_LOW-", "-DOSE_COLOUR_LOW-")
#
#         if event == "-DOSE_COLOUR_HIGH-":
#             dose_colour_dict = dose_colouring(window, event, values, temp_sample_amount,
#                                               "-DOSE_COLOUR_BUTTON_HIGH-", "-DOSE_COLOUR_HIGH-")
#
#         # State mapping
#         if event == "-PLATE_LAYOUT_COLOUR_CHOSE_TARGET-":
#             if values["-PLATE_LAYOUT_COLOUR_CHOSE_TARGET-"] != "None":
#                 window["-PLATE_LAYOUT_COLOUR_CHOSE-"].update(
#                     button_color=values["-PLATE_LAYOUT_COLOUR_CHOSE_TARGET-"])
#
#         # Plate Layout functions:
#         if event == "-ARCHIVE_PLATES-":
#             print("Needs to update ARCHIVE_PLATES_SUB")
#
#         if event == "-RECT_SAMPLE_TYPE-" and values["-ARCHIVE_PLATES-"]:
#             print("Needs to update ARCHIVE_PLATES_SUB")
#
#         if event == "-DRAW-":
#             well_dict, min_x, min_y, max_x, max_y, plate_active, graph, plate_type, archive_plates, gui_tab, sample_type\
#                 = draw_layout(config, sg, window, event, values, well_dict, graph_plate, archive_plates_dict)
#
#         if event == "-EXPORT_LAYOUT-":
#             print(well_dict)
#             # export_layout(config, sg, window, event, values, well_dict)
#
#         if event == "-SAVE_LAYOUT-":
#             save_layout(config, sg, window, event, values, well_dict, archive_plates_dict)
#
#         if event == "-DELETE_LAYOUT-":
#             delete_layout(config, sg, window, event, values)
#
#         if event == "-RENAME_LAYOUT-":
#             rename_layout(config, sg, window, event, values)
#
#         # Used both for Plate layout and Bio Info
#         # prints coordinate and well under the plate layout
#         try:
#             event.endswith("+MOVE")
#
#         except AttributeError:
#             pass
#
#         else:
#             if event.endswith("+MOVE") and type(event) != tuple:
#                 on_move(config, sg, window, event, values, graph_bio_exp, well_dict_bio_info, plate_bio_info,
#                         graph_plate, well_dict)
#
#         if event == "-RECT_BIO_CANVAS-":
#             x, y = values["-RECT_BIO_CANVAS-"]
#             if not dragging:
#                 start_point = (x, y)
#                 dragging = True
#             else:
#                 end_point = (x, y)
#             if prior_rect:
#                 graph_plate.delete_figure(prior_rect)
#
#             # Choosing which tool to pain the plate with.
#             if None not in (start_point, end_point):
#                 for temp_draw_value in draw_tool_dict:
#                     if values[temp_draw_value]:
#                         temp_draw_tool = draw_tool_dict[temp_draw_value]
#                 temp_selector = True
#                 prior_rect = graph_plate.draw_rectangle(start_point, end_point, fill_color="",
#                                                         line_color="white")
#
#         # it does not always detect this event:
#         try:
#             event.endswith("+UP")
#         except AttributeError:
#             pass
#         else:
#             if event.endswith("+UP"):
#                 start_point, end_point, dragging, prior_rect, temp_selector, temp_draw_tool, well_dict = \
#                     on_up(config, sg, window, event, values, temp_selector, plate_active, start_point, end_point, x, y,
#                           min_x, max_x, min_y, max_y, graph_plate, clm_to_row_converter, plate_type_count, plate_type,
#                           color_select, temp_draw_tool, well_dict, prior_rect, dose_colour_dict)
#

def get_data(text_file, excel_file):

    compound_data = {}
    with open(text_file, "r", newline="\n") as csv_file:
        all_lines = csv_file.readlines()

        for line_index, line in enumerate(all_lines):
            if line_index > 0:
                line = line.strip()
                line = line.split(";")

                if line[-1] != "":
                    compound_id = line[-1]
                    location = line[0]
                    new_loc = ""
                    for data in location:
                        if data != "0":
                            new_loc += data
                compound_data[compound_id] = new_loc

    print(compound_data)

    wb = Workbook()
    ws = wb.active
    indent_col = row_counter = 1
    headlines = ["state", "compound", "concentration", "source_wells", "source_plate", "vol"]

    for headings_index, headings in enumerate(headlines):
        ws.cell(column=indent_col + headings_index, row=row_counter, value=headings)

    converter = {"A": "I", "B": "J",
                 "C": "K", "D": "L",
                 "E": "M", "F": "N",
                 "G": "O", "H": "p"}
    conc = ["10mM", "0.1mM", "0.001mM"]
    sourplace = ["source_1", "source_2", "source_1"]
    vol = 0
    row_counter += 1
    for data_index, data in enumerate(compound_data):
        for count in range(3):
            temp_location = compound_data[data]
            ws.cell(column=indent_col, row=row_counter + count, value="sample")
            ws.cell(column=indent_col + 1, row=row_counter + count, value=data)
            ws.cell(column=indent_col + 2, row=row_counter + count, value=conc[count])
            if count == 2:
                temp_location = f"{converter[temp_location[0]]}{temp_location[1]}"

            ws.cell(column=indent_col + 3, row=row_counter + count, value=temp_location)
            ws.cell(column=indent_col + 4, row=row_counter + count, value=sourplace[count])
            ws.cell(column=indent_col + 5, row=row_counter + count, value=vol)
        row_counter += 3
    wb.save(excel_file)


if __name__ == "__main__":
    # import configparser
    # config = configparser.ConfigParser()
    # config.read("config.ini")
    # plate_list, archive_plates_dict = get_plate_layout(config)
    # # print(archive_plates_dict)
    # run(config, plate_list, archive_plates_dict)

    text_file = Path(r"E:\Alpha_so_dose_response.txt")
    excel_fil = Path(r"C:\Users\phch\Desktop\test\dose_reposnse_layout.xlsx")

    get_data(text_file, excel_fil)




