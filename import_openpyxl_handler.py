from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment

import re

from info import numb2alpha


def border_thin():
    """

    :return: A thin border style
    :rtype: openpyxl.descriptors.MetaSerialisable
    """
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))


def border_thick():
    """

    :return: A thick border style
    :rtype: openpyxl.descriptors.MetaSerialisable
    """
    return Border(left=Side(style='thick'), right=Side(style='thick'),
                  top=Side(style='thick'), bottom=Side(style='thick'))


def border_medium():
    """

    :return: A medium border style
    :rtype: openpyxl.descriptors.MetaSerialisable
    """
    return Border(left=Side(style='medium'), right=Side(style='medium'),
                  top=Side(style='medium'), bottom=Side(style='medium'))


def border_style_red():
    """

    :return: A border style with red double thin lines at the top and bottom and black thin in the side
    :rtype: openpyxl.descriptors.MetaSerialisable
    """
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    return Border(top=double, left=thin, right=thin, bottom=double)


def table_purple(ws, table_name, start_row, start_clm, end_row, end_clm):
    """
    Makes a purple table in the range of the cells.

    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param table_name: The name of the table
    :type table_name: str
    :param start_row: The starting row
    :type start_row: int
    :param start_clm: The starting column
    :type start_clm: int
    :param end_row: The ending row
    :type end_row: int
    :param end_clm: The ending column
    :type end_clm: int
    :return: A purple table for the indicated range
    :rtype: openpyxl.worksheet.table.Table
    """

    ws_name = _get_sheet_name(ws)
    table_name = f"{table_name}_{ws_name}"
    counter = 1

    while [item for item in ws.tables.items() if item[0] == f"{table_name}"]:
        counter += 1
        table_name = f"{table_name}{counter}"

    start_cell = ex_cell(start_row, start_clm)
    end_cell = ex_cell(end_row, end_clm)
    table_range = f"{start_cell}:{end_cell}"
    tab = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    return tab


def merge_cells_single_row(value, ws, row, start_clm,  end_clm, border=False, border_style="thin"):
    """
    Merges cells together that are all on the same row, and apply borders around them, if needed...
    This should be writen out to work for all merge cell groups...

    :param value: What value to write in the merge cells
    :type value: str
    :param ws: the works sheet
    :type ws: "openpyxl.worksheet.worksheet.Worksheet"
    :param row: the row for the merge cells
    :type row: int
    :param start_clm: the starting column for the merge cells
    :type start_clm: int
    :param end_clm: the ending column for the merge cells
    :type end_clm: int
    :param border: If there should be a border around the merged cells or not
    :type border: bool
    :param border_style: The style of the borders. if more is needed they needs to be made
        "thin": border_thin,
        "medium": border_medium,
        "thick": border_thick,
        "red_line": border_style_red
    :type border_style: str
    :return: Merged cells in the excel sheet
    """

    border_styles = {"thin": border_thin,
                     "medium": border_medium,
                     "thick": border_thick,
                     "red_line": border_style_red}
    try:
        ws.merge_cells(start_row=row, start_column=start_clm, end_row=row, end_column=end_clm)
    except ValueError:
        print("To many lines")
    else:
        temp_cell = ws.cell(row=row, column=start_clm)
        temp_cell.value = value
        temp_cell.font = Font(b=True)
        temp_cell.alignment = Alignment(horizontal='center', vertical='center')
        range_merge = end_clm - start_clm + 1
        if border:
            for i in range(range_merge):
                temp_cell = ws.cell(row=row, column=start_clm + i)
                temp_cell.border = border_styles[border_style]()


def _get_sheet_name(ws):
    """
    Remove unnecessary characters around the sheet named pulled from the Workbook

    :param ws: The worksheet for the excel filere where the data is added
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :return: The sheet name stripped of unnecessary characters
    :rtype: str
    """
    worksheet_name = f"{ws}"
    worksheet_name = worksheet_name.split()[-1]
    return re.sub('[^A-Za-z0-9]+', '', worksheet_name)


def next_letter(current_letter, spacer):
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    # Find the index of the current letter
    current_index = letters.index(current_letter.capitalize())

    # Calculate the index of the next letter
    next_index = current_index + spacer

    if next_index < len(letters):
        return letters[next_index]
    else:
        # Calculate the number of times to append "A" to the output
        num_a = next_index // len(letters)

        # Calculate the remainder to find the position of the letter after "Z"
        remainder = next_index % len(letters)

        # Recursively call next_letter to get the letter after "Z"
        next_letter_after_z = next_letter("A", remainder)

        # Concatenate "A" and the letter after "Z"
        return "A" * num_a + next_letter_after_z



def ex_cell(row, col):
    """
    takes row value and column value and translate it to a cell value  ex from (1,1) = (A1)

    :param row: value for the row
    :param col: value for the column
    :return: cell value
    :rtype: str
    """
    return f"{numb2alpha[col]}{row}"


def ws_creator(wb, method, data_type):
    """
    Create different worksheets.

    :param wb: The Workbook / the excel file that are being worked in
    :type wb: openpyxl.workbook.workbook.Workbook
    :param method: The different methods that are being writen data for.
    :type method: str
    :param data_type: The data_type of what is added. (Matrix or List, for now)
    :type data_type: str
    :return: A Worksheet named after the method and the data_type
    :rtype openpyxl.worksheet.worksheet.Worksheet
    """
    return wb.create_sheet(f"{method}_{data_type}")