from openpyxl.reader.excel import load_workbook
from statistics import stdev
from database_handler import DataBaseFunctions


def _get_values_for_calc(compound_data):
    max_values = {}
    min_values = {"start": [], "end": []}
    for state in compound_data["calc"]:

        for compound in compound_data["calc"][state]:
            for compound_group in compound_data["calc"][state][compound]:
                if state == "max" or state == "maximum" or state == "positive":
                    max_values[compound_group] = {"start": [], "end": []}
                for name_counter in compound_data["calc"][state][compound][compound_group]:
                    well_id = compound_data["calc"][state][compound][compound_group][name_counter]["well_id"]
                    data = compound_data["calc"][state][compound][compound_group][name_counter]["data"]
                    if state == "max" or state == "maximum" or state == "positive":
                        max_values[compound_group]["start"].append(data[0])
                        max_values[compound_group]["end"].append(data[1])
                    else:
                        min_values["start"].append(data[0])
                        min_values["end"].append(data[1])

    return max_values, min_values


def _z_prime_calculator(compound_data):
    """
    Calculate Z-prime

    :param compound_data: All the data for the reading, including the state of the well following the plate layout, and the
        results from different calculations is added as they get to it.
    :type compound_data: dict

    :return: Returns the Z-Prime values
    :rtype: dict
    """

    # Calculate the result
    max_values, min_values = _get_values_for_calc(compound_data)
    z_prime_results = {}
    min_stdev_start = stdev(min_values["start"])
    min_stdev_end = stdev(min_values["end"])
    min_avg_start = sum(min_values["start"])/len(min_values["start"])
    min_avg_end = sum(min_values["end"])/len(min_values["end"])

    for compounds in max_values:
        max_stdev_start = stdev(max_values[compounds]["start"])
        max_stdev_end = stdev(max_values[compounds]["end"])
        max_avg_start = sum(max_values[compounds]["start"]) / len(max_values[compounds]["start"])
        max_avg_end = sum(max_values[compounds]["end"]) / len(max_values[compounds]["end"])
        z_prime_results[compounds] = {"start": 1 - ((3 * (max_stdev_start + min_stdev_start)) / abs(max_avg_start + min_avg_start)),
                                      "end": 1 - ((3 * (max_stdev_end + min_stdev_end)) / abs(max_avg_end + min_avg_end)),}

    compound_data["calc"]["z_prime"] = z_prime_results
    # Return the result
    return compound_data


def __plate_counter_to_well(plate_layout):
    plate_layout_counter_conveter = {}
    for counter in plate_layout:
        plate_layout_counter_conveter[plate_layout[counter]["well_id"]] = counter

    return plate_layout_counter_conveter


def _data_collector(config, barcode, all_data, plate_layout, well_type, worklist_dict, translate_dict, include_id):
    plate_layout_counter_conveter = __plate_counter_to_well(plate_layout)

    compound_data = {"plate_data": {},
                     "calc": {}}

    if include_id:
        dbf = DataBaseFunctions(config)
        table_name = "compound_mp"
        barcode_name = "mp_barcode"
        id_name = "mp_well"

    for state in well_type:
        # print(state)

        for wells in well_type[state]:
            try:
                compound_data["plate_data"][state]
            except KeyError:
                compound_data["plate_data"][state] = {}
                if state != "sample":
                    compound_data["calc"][state] = {}

            all_well_data = list(all_data["plates"]["original"]["wells"][wells]["well_data"][0])

            try:
                worklist_dict[barcode][wells]
            except KeyError:
                print(f"missing data for well: {wells}")
                continue
            else:
                source_well = worklist_dict[barcode][wells]["source_well"]
                source_plate = worklist_dict[barcode][wells]["source_plate"]

            # for naming
            name_counter = 0
            run = True

            if state != "sample":

                counter = plate_layout_counter_conveter[source_well]
                compound_group = plate_layout[counter]["group"]
                try:
                    compound = list(translate_dict[source_well].keys())[0]
                except KeyError:
                    print(wells)
                    continue

                try:
                    compound_data["plate_data"][state][compound]

                except KeyError:
                    compound_data["plate_data"][state][compound] = {}
                    compound_data["calc"][state][compound] = {}

                try:
                    compound_data["plate_data"][state][compound][compound_group]

                except KeyError:
                    compound_data["plate_data"][state][compound][compound_group] = {}
                    compound_data["calc"][state][compound][compound_group] = {}

                while run or name_counter > 99999:
                    try:
                        compound_data["plate_data"][state][compound][compound_group][name_counter]
                    except KeyError:
                        compound_data["calc"][state][compound][compound_group][name_counter] = {"well_id": wells,
                                                                                                "data": all_well_data}
                        compound_data["plate_data"][state][compound][compound_group][name_counter] = {"well_id": wells,
                                                                                                  "data": all_well_data}
                        run = False
                    else:
                        name_counter += 1

            else:
                if include_id:

                    sample_row = dbf.find_data_double_lookup(table_name, source_plate, source_well,
                                                             barcode_name, id_name)
                    try:
                        sample_id = sample_row[0][3]
                    except IndexError:
                        sample_id = "Not found"
                else:
                    sample_id = wells

                while run or name_counter > 99999:
                    try:
                        compound_data["plate_data"][state][sample_id][name_counter]
                    except KeyError:
                        compound_data["plate_data"][state][sample_id][name_counter] = {"well_id": wells, "data": all_well_data}
                        run = False
                    else:
                        name_counter += 1

    return compound_data


def _translate_dict_generator(bonus_layout):
    translate_dict = {}
    wb = load_workbook(bonus_layout)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    for row_index, row in enumerate(ws.values):
        if row_index > 0:
            try:
                translate_dict[row[0]]
            except KeyError:
                translate_dict[row[0]] = {row[1]: {row[7]: {"trans_vol": row[6], "type": row[5]}}}
            else:
                translate_dict[row[0]][row[1]][row[7]] = {"trans_vol": row[6], "type": row[5]}

    return translate_dict


def _hit_calculator(compound_data, hit_threshold, data_style):
    compound_data["hits"] = {}
    compound_data["control"] = {}

    for state in compound_data["plate_data"]:
        if state != "sample" and state != "blank":

            for compound in compound_data["plate_data"][state]:
                for compound_group in compound_data["plate_data"][state][compound]:
                    for name_counter in compound_data["plate_data"][state][compound][compound_group]:
                        well_id = compound_data["plate_data"][state][compound][compound_group][name_counter]["well_id"]
                        data = compound_data["plate_data"][state][compound][compound_group][name_counter]["data"]
                        start_value = data[0]
                        end_value = data[-1]
                        if end_value < hit_threshold["pH"]:
                            for data_index, data_points in enumerate(data):
                                if data_points < hit_threshold["pH"]:
                                    time_to_hit = compound_data["time_slots"][data_index]
                                    break
                        else:
                            time_to_hit = None

                        diff = end_value - start_value

                        compound_data["control"][well_id] = {"state": state,
                                                             "compound_name": compound,
                                                             "diff": diff,
                                                             "init": start_value,
                                                             "end": end_value,
                                                             "time_to_hit": time_to_hit}
        elif state == "sample":
            for compound in compound_data["plate_data"][state]:
                for name_counter in compound_data["plate_data"][state][compound]:
                    is_hit = False
                    data = compound_data["plate_data"][state][compound][name_counter]["data"]
                    well_id = compound_data["plate_data"][state][compound][name_counter]["well_id"]
                    start_value = data[0]
                    end_value = data[-1]
                    diff = end_value - start_value

                    if end_value < hit_threshold["pH"]:
                        for data_index, data_points in enumerate(data):
                            if data_points < hit_threshold["pH"]:
                                time_to_hit = compound_data["time_slots"][data_index]
                                break
                    else:
                        time_to_hit = None

                    if time_to_hit and time_to_hit < hit_threshold["time_limit"]:
                        is_hit = True

                    if is_hit and data_style == "org":
                        is_hit = True
                        compound_data["hits"][well_id] = {"state": state,
                                                          "compound_name": compound,
                                                          "compound_group": compound_group,
                                                          "name_counter": name_counter,
                                                          "diff": diff,
                                                          "init": start_value,
                                                          "end": end_value,
                                                          "time_to_hit": time_to_hit}
                    else:
                        is_hit = False
                    compound_data["plate_data"][state][compound][name_counter]["hit"] = is_hit
        else:
            # skipping blank wells
            continue
    return compound_data


def data_controller(config, barcode, data_style, plate_layout, all_data,
                    well_type, worklist_dict, include_id, translate_dict, hit_threshold):

    compound_data = _data_collector(config, barcode, all_data, plate_layout, well_type, worklist_dict, translate_dict,
                                    include_id)

    compound_data["time_slots"] = all_data["time_slots"]
    compound_data = _z_prime_calculator(compound_data)
    compound_data = _hit_calculator(compound_data, hit_threshold, data_style)

    return compound_data





