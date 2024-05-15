from os.path import exists
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook, load_workbook, chart
from openpyxl.chart import LineChart, Reference, series

from openpyxl.chart.axis import DateAxis

from import_openpyxl_handler import ex_cell, next_letter

headline_background_colour = "bb07ed"
colour_red = "ed0707"
colour_yellow = "ede507"
colour_green = "1aed07"

graph_spacer = 10
z_prime_info = {"lower_limit": 0.5,
                "higher_limit": 0.75}



# def _raw_data_clm(wb, compound_data):
#     graph_data = {"plate_data": {},
#                   "compound_data": {},
#                   "control": {},
#                   "hit data": {}}
#     try:
#         ws = wb["Sheet"]
#     except KeyError:
#         ws = wb.create_sheet("raw_data")
#     else:
#         ws.title = "raw_data"
#
#     init_row = 1
#     init_col = 1
#     plate_row = 0
#     index_row = init_row
#     index_col = init_col
#     headlines_raw = ["State", "Compound_ID", "Conc(uM)", "Well_ID"]
#     for barcode in compound_data:
#         try:
#             graph_data["plate_data"][barcode]
#         except KeyError:
#             graph_data["plate_data"][barcode] = {}
#         for style_index, style in enumerate(compound_data[barcode]):
#
#             graph_data["plate_data"][barcode][style] = {"full_plate": {"x_values": {}, "data_points": {}},
#                                                         "controls": {"x_values": {}, "data_points": {}},
#                                                         "hits": {"x_values": {}, "data_points": {}}
#                                                         }
#
#             ws.cell(column=index_col, row=index_row, value=barcode).fill = \
#                 PatternFill(fill_type="solid", fgColor=headline_background_colour)
#             if style:
#                 ws.cell(column=index_col + 1, row=index_row, value=style).fill = \
#                     PatternFill(fill_type="solid", fgColor=headline_background_colour)
#             index_row += 1
#             plate_row = index_row
#             for headline in headlines_raw:
#                 ws.cell(column=index_col, row=index_row, value=headline)
#                 index_row += 1
#
#             for time_index, time_counter in enumerate(compound_data[barcode][style]["time_slots"]):
#                 if time_index == 0:
#                     full_plate_x_values_start_col = index_col
#                     full_plate_x_values_start_row = index_row
#                     full_plate_data_start_row = index_row
#                 time = compound_data[barcode][style]["time_slots"][time_counter]
#                 ws.cell(column=index_col, row=index_row, value=time)
#                 full_plate_x_values_end_col = index_col
#                 full_plate_x_values_end_row = index_row
#
#                 index_row += 1
#
#             index_col += 1
#             index_row = plate_row
#
#             full_plate_data_start_col = index_col
#             for state_index, state in enumerate(compound_data[barcode][style]["plate_data"]):
#                 graph_data[state] = {}
#                 if state != "sample":
#
#                     for compound_name in compound_data[barcode][style]["plate_data"][state]:
#                         for group in compound_data[barcode][style]["plate_data"][state][compound_name]:
#                             for counter_index, counter in enumerate(compound_data[barcode][style]["plate_data"][state][compound_name][group]):
#
#                                 well_id = compound_data[barcode][style]["plate_data"][state][compound_name][group][counter]["well_id"]
#                                 # write down the infor
#                                 ws.cell(column=index_col, row=index_row + 0, value=state)
#                                 ws.cell(column=index_col, row=index_row + 1, value=compound_name)
#                                 ws.cell(column=index_col, row=index_row + 2, value=group)
#                                 ws.cell(column=index_col, row=index_row + 3, value=well_id)
#
#                                 index_row += 4
#
#                                 for readings in compound_data[barcode][style]["plate_data"][state][compound_name][group][counter]["data"]:
#
#                                     ws.cell(column=index_col, row=index_row, value=readings)
#                                     full_plate_data_end_col = index_col
#                                     index_row += 1
#
#                                 max_row = index_row
#                                 index_col += 1
#                                 index_row = plate_row
#                             full_plate_data_end_row = max_row - 1
#             full_plate_x_values_col = [full_plate_x_values_start_col, full_plate_x_values_end_col]
#             full_plate_x_values_row = [full_plate_x_values_start_row, full_plate_x_values_end_row]
#             full_plate_data_col = [full_plate_data_start_col, full_plate_data_end_col]
#             full_plate_data_row = [full_plate_data_start_row, full_plate_data_end_row]
#
#             graph_data["plate_data"][barcode][style]["full_plate"]["data_points"] = \
#                 {"data":
#                      {"col": full_plate_data_col,
#                       "row": full_plate_data_row},
#                  "x_values":
#                      {"col": full_plate_x_values_col,
#                       "row": full_plate_x_values_row},
#                  }
#
#             index_row = max_row + 1
#             index_col = init_col
#         index_row = max_row + 2
#         index_col = init_col
#
#     return graph_data, ws
#


def _raw_data(wb, compound_data):
    graph_data = {"plate_data": {},
                  "compound_data": {},
                  "control": {},
                  "ref": {}}
    try:
        ws = wb["Sheet"]
    except KeyError:
        ws = wb.create_sheet("raw_data")
    else:
        ws.title = "raw_data"

    init_row = 1
    init_col = 1
    index_row = init_row
    index_col = init_col
    headlines_raw = ["State", "Compound_ID", "Conc(uM)", "Well_ID"]
    for barcode in compound_data:
        try:
            graph_data["plate_data"][barcode]
        except KeyError:
            graph_data["plate_data"][barcode] = {}
            graph_data["ref"][barcode] = {}
        for style_index, style in enumerate(compound_data[barcode]):
            graph_data["ref"][barcode][style] = {}
            graph_data["plate_data"][barcode][style] = {"full_plate": {"x_values": {}, "data_points": {}},
                                                        "controls": {"x_values": {}, "data_points": {}},
                                                        "hits": {"x_values": {}, "data_points": {}}
                                                        }

            ws.cell(column=index_col, row=index_row, value=barcode).fill = \
                PatternFill(fill_type="solid", fgColor=headline_background_colour)
            if style:
                ws.cell(column=index_col + 1, row=index_row, value=style).fill = \
                    PatternFill(fill_type="solid", fgColor=headline_background_colour)

            index_row += 1
            index_col = init_col

            # Write Z-prime for the plate:
            prime_row = index_row
            ws.cell(column=index_col, row=prime_row, value="Z-Prime")
            ws.cell(column=index_col, row=prime_row + 1, value="Start")
            ws.cell(column=index_col, row=prime_row + 2, value="End")

            index_col += 1

            for data in compound_data[barcode][style]["calc"]["z_prime"]:
                ws.cell(column=index_col, row=prime_row, value=data)

                z_start = compound_data[barcode][style]["calc"]["z_prime"][data]["start"]

                if z_start < z_prime_info["lower_limit"]:
                    approval_colour = colour_red
                elif z_prime_info["lower_limit"] < z_start < z_prime_info["higher_limit"]:
                    approval_colour = colour_yellow
                else:
                    approval_colour = colour_green

                ws.cell(column=index_col, row=prime_row + 1, value=z_start).fill = \
                    PatternFill(fill_type="solid", fgColor=approval_colour)

                z_end = compound_data[barcode][style]["calc"]["z_prime"][data]["end"]

                if z_end < z_prime_info["lower_limit"]:
                    approval_colour = colour_red
                elif z_prime_info["lower_limit"] < z_end < z_prime_info["higher_limit"]:
                    approval_colour = colour_yellow
                else:
                    approval_colour = colour_green

                ws.cell(column=index_col, row=prime_row + 2, value=z_end).fill = \
                    PatternFill(fill_type="solid", fgColor=approval_colour)

                index_col += 1


            index_row += 3
            index_col = init_col
            plate_row = index_row
            for headline in headlines_raw:
                ws.cell(column=index_col, row=index_row, value=headline)
                index_col += 1

            for time_index, time_counter in enumerate(compound_data[barcode][style]["time_slots"]):
                if time_index == 0:
                    graph_data["ref"][barcode][style]["lables"] = {"row": index_row, "col_start": index_col}
                    full_plate_x_values_start_col = index_col
                    full_plate_x_values_start_row = index_row
                time = compound_data[barcode][style]["time_slots"][time_counter]
                ws.cell(column=index_col, row=index_row, value=time)
                full_plate_x_values_end_col = index_col
                full_plate_x_values_end_row = index_row
                graph_data["ref"][barcode][style]["lables"]["col_end"] = index_col

                index_col += 1

            index_col = init_col
            index_row += 1

            full_plate_data_start_row = index_row
            full_plate_data_start_col = init_col + 4
            for state_index, state in enumerate(compound_data[barcode][style]["plate_data"]):

                if state != "sample":

                    for compound_name in compound_data[barcode][style]["plate_data"][state]:
                        for group in compound_data[barcode][style]["plate_data"][state][compound_name]:
                            for counter_index, counter in enumerate(compound_data[barcode][style]["plate_data"][state][compound_name][group]):

                                well_id = compound_data[barcode][style]["plate_data"][state][compound_name][group][counter]["well_id"]
                                # write down the infor
                                ws.cell(column=index_col + 0, row=index_row, value=state)
                                ws.cell(column=index_col + 1, row=index_row, value=compound_name)
                                ws.cell(column=index_col + 2, row=index_row, value=group)
                                ws.cell(column=index_col + 3, row=index_row, value=well_id)

                                index_col += 4

                                graph_data["ref"][barcode][style][well_id] = {"row": index_row, "col_start": index_col}


                                for readings in compound_data[barcode][style]["plate_data"][state][compound_name][group][counter]["data"]:

                                    ws.cell(column=index_col, row=index_row, value=readings)
                                    full_plate_data_end_col = index_col
                                    index_col += 1
                                graph_data["ref"][barcode][style][well_id]["col_end"] = index_col - 1
                                index_row += 1
                                index_col = init_col
                            max_row = index_row

                            full_plate_data_end_row = max_row - 1
            full_plate_x_values_col = [full_plate_x_values_start_col, full_plate_x_values_end_col]
            full_plate_x_values_row = [full_plate_x_values_start_row, full_plate_x_values_end_row]
            full_plate_data_col = [full_plate_data_start_col, full_plate_data_end_col]
            full_plate_data_row = [full_plate_data_start_row, full_plate_data_end_row]

            graph_data["plate_data"][barcode][style]["full_plate"]["data_points"] = \
                {"data":
                     {"col": full_plate_data_col,
                      "row": full_plate_data_row},
                 "x_values":
                     {"col": full_plate_x_values_col,
                      "row": full_plate_x_values_row},
                 }

            index_row = max_row + 1
            index_col = init_col
        index_row = max_row + 2
        index_col = init_col

    return graph_data, ws


def __write_hits(wb, ws_raw, graph_data, compound_data, hit_data, hit_threshold):
    ws_chart = wb.create_sheet("hits")
    headlines = ["State", "compound_id", "well_id", f"time_to_hit_{hit_threshold['pH']}", "diff", "initial", "end", "ref"]
    init_row = 1
    init_col = 1
    index_col = init_col
    index_row = init_row

    for barcode in compound_data:
        ws_chart.cell(column=index_col, row=index_row, value=barcode)
        index_row += 1

        for headline in headlines:
            ws_chart.cell(column=index_col, row=index_row, value=headline)
            index_col += 1
        index_col = init_col
        index_row += 1

        # Write Hit data
        for well in hit_data[barcode]:
            compound_id = compound_data["hits"][well]["compound_name"]
            diff = compound_data["hits"][well]["diff"]
            init = compound_data["hits"][well]["init"]
            end = compound_data["hits"][well]["end"]
            time_to_hit = compound_data["hits"][well]["time_to_hit"]
            ws_chart.cell(column=index_col + 0, row=index_row, value="Sample")
            ws_chart.cell(column=index_col + 1, row=index_row, value=compound_id)
            ws_chart.cell(column=index_col + 2, row=index_row, value=well)
            ws_chart.cell(column=index_col + 3, row=index_row, value=time_to_hit)
            ws_chart.cell(column=index_col + 4, row=index_row, value=diff)
            ws_chart.cell(column=index_col + 5, row=index_row, value=init)
            ws_chart.cell(column=index_col + 6, row=index_row, value=end)
            index_col += 6
            for style in compound_data[barcode]:
                if style != "org":
                    temp_data = graph_data["ref"][barcode][style][well]
                    temp_init = ws_raw.cell(column=temp_data["col_start"], row=temp_data["row"]).value
                    ws_chart.cell(column=index_col, row=index_row, value=temp_init)
                    index_col += 1
                    temp_end = ws_raw.cell(column=temp_data["col_end"], row=temp_data["row"]).value
                    ws_chart.cell(column=index_col, row=index_row, value=temp_end)
                    index_col += 1
            index_col = init_col
            index_row += 1
        control_row = index_row
        # write reff
        for style in compound_data[barcode]:
            for well in compound_data[barcode][style]["control"]:

                if style == "org":
                    compound_id = compound_data[barcode][style]["control"][well]["compound_name"]
                    state = compound_data[barcode][style]["control"][well]["state"]
                    temp_data = graph_data["ref"][barcode][style][well]
                    init = ws_raw.cell(column=temp_data["col_start"], row=temp_data["row"]).value
                    end = ws_raw.cell(column=temp_data["col_end"], row=temp_data["row"]).value
                    time_to_hit = compound_data[barcode][style]["control"][well]["time_to_hit"]

                    diff = abs(end - init)
                    ws_chart.cell(column=index_col + 0, row=index_row, value=state)
                    ws_chart.cell(column=index_col + 1, row=index_row, value=compound_id)
                    ws_chart.cell(column=index_col + 2, row=index_row, value=well)
                    ws_chart.cell(column=index_col + 3, row=index_row, value=time_to_hit)
                    ws_chart.cell(column=index_col + 4, row=index_row, value=diff)
                    ws_chart.cell(column=index_col + 5, row=index_row, value=init)
                    ws_chart.cell(column=index_col + 6, row=index_row, value=end)

                else:
                    temp_data = graph_data["ref"][barcode][style][well]
                    temp_init = ws_raw.cell(column=temp_data["col_start"], row=temp_data["row"]).value

                    ws_chart.cell(column=index_col + 7, row=index_row, value=temp_init)

                    temp_end = ws_raw.cell(column=temp_data["col_end"], row=temp_data["row"]).value

                    ws_chart.cell(column=index_col + 8, row=index_row, value=temp_end)

                index_col = init_col
                index_row += 1
                end_row = index_row
            index_row = control_row

        # ready for next plate
        index_row = end_row + 1
        index_col = init_col

    return end_row, ws_chart


def __draw_hit_graph(ws_raw, ws_chart, graph_row, graph_data, compound_data, hit_data):
    graph_row += 2
    min_value = 10
    for barcode in compound_data:
        for style in compound_data[barcode]:

            graph_counter = 0
            graph_col = "A"

            if style == "org":
                graph_col = "A"
                graph_counter += 1
            else:
                if graph_counter > 1 and graph_col == "A":
                    graph_col = next_letter(graph_col, graph_spacer * graph_counter)
                else:
                    graph_col = next_letter(graph_col, graph_spacer)
                graph_counter += 1
            current_chart = LineChart()
            current_chart.title = f"{barcode.capitalize()}_{style}_Hits"
            # current_chart.style = 13

            current_chart.x_axis.title = "Time"
            current_chart.y_axis.title = "pH"
            chart_labels_data = graph_data["ref"][barcode][style]["lables"]
            chart_labels = Reference(ws_raw, min_row=chart_labels_data["row"],
                                     min_col=chart_labels_data["col_start"],
                                     max_col=chart_labels_data["col_end"])

            # Hits
            for well in hit_data[barcode]:
                temp_data = graph_data["ref"][barcode][style][well]
                compound_id = compound_data[barcode][style]["hits"][well]["compound_name"]
                data_points = Reference(ws_raw,
                                        min_row=temp_data["row"],
                                        min_col=temp_data["col_start"],
                                        max_col=temp_data["col_end"])
                data_series = chart.Series(values=data_points, title=compound_id)
                current_chart.series.append(data_series)
                temp_min_value = hit_data[barcode][well]["end"]
                if temp_min_value < min_value:
                    min_value = temp_min_value


            # Adding controls

            for well in compound_data[barcode][style]["control"]:

                temp_data = graph_data["ref"][barcode][style][well]
                compound_id = compound_data[barcode][style]["control"][well]["compound_name"]
                data_points = Reference(ws_raw,
                                        min_row=temp_data["row"],
                                        min_col=temp_data["col_start"],
                                        max_col=temp_data["col_end"])
                data_series = chart.Series(values=data_points, title=compound_id)
                current_chart.series.append(data_series)

                temp_min_value = compound_data[barcode][style]["control"][well]["end"]
                if temp_min_value < min_value:
                    min_value = temp_min_value

            current_chart.y_axis.scaling.min = min_value - 1
            ws_chart.add_chart(current_chart, f"{graph_col}{graph_row}")
            current_chart.set_categories(chart_labels)


def __draw_standard_graph(wb, ws_raw, graph_data):

    ws_chart = wb.create_sheet("all_graph_data")

    for plates in graph_data["plate_data"]:
        for style_counter, style in enumerate(graph_data["plate_data"][plates]):
            if style_counter > 0:
                continue

            data = graph_data["plate_data"][plates][style]["full_plate"]["data_points"]["data"]
            x_values_ref = graph_data["plate_data"][plates][style]["full_plate"]["data_points"]["x_values"]

            chart_labels = Reference(ws_raw, min_col=x_values_ref["col"][0], min_row=x_values_ref["row"][0],
                                   max_col=x_values_ref["col"][-1], max_row=x_values_ref["row"][-1])

            current_chart = LineChart()
            current_chart.title = f"{plates}_{style}"
            current_chart.style = 12
            current_chart.x_axis.title = "Time"
            current_chart.y_axis.title = "pH"

            for row_counter in range(data["row"][0], data["row"][-1] + 1):
                if row_counter > 50:
                    continue
                data_points = Reference(ws_raw, min_row=row_counter, min_col=data["col"][0], max_col=data["col"][-1])

                compound_id = ws_raw.cell(row=row_counter, column=data["col"][0]-3).value
                data_series = chart.Series(values=data_points, title=compound_id)
                current_chart.series.append(data_series)

            ws_chart.add_chart(current_chart, "A10")
            current_chart.set_categories(chart_labels)


def _draw_graphs(wb, ws_raw, graph_data, compound_data, hit_data, hit_threshold):
    __draw_standard_graph(wb, ws_raw, graph_data)
    graph_row, ws_chart = __write_hits(wb, ws_raw, graph_data, compound_data, hit_data, hit_threshold)
    __draw_hit_graph(ws_raw, ws_chart, graph_row, graph_data, compound_data, hit_data)


def _grab_hits(compound_data):
    hit_data = {}
    for barcode in compound_data:
        hit_data[barcode] = {}
        for style in compound_data[barcode]:
            for well in compound_data[barcode][style]["hits"]:
                hit_data[barcode][well] = compound_data[barcode][style]["hits"][well]
    return hit_data


def _saving(wb, save_location, report_file_name):
    report_path = f"{save_location}/{report_file_name}.xlsx"
    counter = 1
    while exists(report_path) or counter > 9999:
        report_path = f"{save_location}/{report_file_name}_{counter}.xlsx"
        counter += 1

    wb.save(report_path)

    return report_path


def procul_report_controller(report_file_name, compound_data, save_location, report_setup, hit_threshold):
    wb = Workbook()
    graph_data, ws_raw = _raw_data(wb, compound_data)
    hit_data = _grab_hits(compound_data)
    _draw_graphs(wb, ws_raw, graph_data, compound_data, hit_data, hit_threshold)
    report_path = _saving(wb, save_location, report_file_name)

    return report_path


