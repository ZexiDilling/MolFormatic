from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font


def _cal_writer_final_report(barcode, ws_report, all_data, init_row, init_col, report_output):

    row_counter = init_row

    ws_report.cell(column=-1 + init_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
    row_counter += 1
    for plate_analysed in all_data["calculations"]:
        # Removing other calculations than avg and stdev
        if plate_analysed != "other_data":
            # Checks to see if the overview of avg and stv should be included
            if report_output[plate_analysed]["overview"]:
                # Writes the analysed method in, if the overview is set to true
                ws_report.cell(column=-1 + init_col, row=row_counter, value=plate_analysed).font = Font(b=True)
                # row_counter += 1
                for state in all_data["calculations"][plate_analysed]:
                    if report_output[plate_analysed][state]:
                        ws_report.cell(column=init_col, row=row_counter, value=state).font = Font(b=True)
                        for calc in all_data["calculations"][plate_analysed][state]:
                            # Writes avg and stdev including values
                            ws_report.cell(column=init_col + 1, row=row_counter, value=calc)
                            ws_report.cell(column=init_col + 2, row=row_counter,
                                           value=all_data["calculations"][plate_analysed][state][calc])
                            row_counter += 1
        else:
            if report_output["z_prime"]:
                ws_report.cell(column=init_col, row=row_counter,
                               value="z-Prime").font = Font(b=True)
                try:
                    ws_report.cell(column=init_col + 2, row=row_counter,
                                   value=all_data["calculations"][plate_analysed]["z_prime"])
                except KeyError:
                    ws_report.cell(column=init_col + 2, row=row_counter,
                                   value="Z-Prime is not calculated for the plates")
                row_counter += 1
            row_counter += 1
    return ws_report, row_counter


def _well_writer_final_report(ws, hits, final_report_setup, init_row):
    indent_col = 1
    row_counter = init_row

    for barcode in hits:
        # Writes headline for data inserts to see where the data is coming from
        ws.cell(column=indent_col, row=row_counter, value=barcode).font = Font(b=True, underline="single")
        row_counter += 1

        for method in hits[barcode]:
            if final_report_setup["methods"][method]:
                # writes method
                ws.cell(column=indent_col, row=row_counter, value=method).font = Font(b=True)
                row_counter += 1
                for split in hits[barcode][method]:
                    ws.cell(column=indent_col, row=row_counter, value=split).font = Font(b=True)
                    ws.cell(column=indent_col+1, row=row_counter,
                                   value=final_report_setup["pora_threshold"][split]["min"]).font = \
                        Font(underline="single")
                    ws.cell(column=indent_col+2, row=row_counter,
                                   value=final_report_setup["pora_threshold"][split]["max"]).font = \
                        Font(underline="single")
                    row_counter += 1
                    for well in hits[barcode][method][split]:
                        ws.cell(column=indent_col + 1, row=row_counter, value=well)
                        ws.cell(column=indent_col + 2, row=row_counter,
                                       value=hits[barcode][method][split][well])
                        row_counter += 1
        indent_col += 4
        row_counter = init_row


def _get_data(all_plate_data, final_report_setup):
    data_calc_dict = {}
    temp_hits = {}
    plate_counter = 0
    all_states = []
    all_methods = []

    for barcode in all_plate_data:
        plate_counter += 1
        temp_hits[barcode] = {}
        data_calc_dict[barcode] = {}
        for method in all_plate_data[barcode]["plates"]:
            if method != "other_data":
                if method not in all_methods:
                    all_methods.append(method)
            if final_report_setup["methods"][method]:
                temp_hits[barcode][method] = {"low": {}, "mid": {}, "high": {}}
                for well in all_plate_data[barcode]["plates"][method]["wells"]:
                    if well in all_plate_data[barcode]["plates"][method]["sample"]:
                        for split in final_report_setup["pora_threshold"]:
                            temp_well_value = all_plate_data[barcode]["plates"][method]["wells"][well]
                            if float(final_report_setup["pora_threshold"][split]["min"]) < float(temp_well_value) < \
                                    float(final_report_setup["pora_threshold"][split]["max"]):
                                temp_hits[barcode][method][split][well] = temp_well_value

        for method in all_plate_data[barcode]["calculations"]:
            data_calc_dict[barcode][method] = {}
            if method != "other_data":
                for state in all_plate_data[barcode]["calculations"][method]:
                    if state not in all_states:
                        all_states.append(state)

                    data_calc_dict[barcode][method][state] = {}
                    for calc in all_plate_data[barcode]["calculations"][method][state]:
                        data_calc_dict[barcode][method][state][calc] = \
                            all_plate_data[barcode]["calculations"][method][state][calc]

            else:
                for other_calc in all_plate_data[barcode]["calculations"][method]:
                    data_calc_dict[barcode][method][other_calc] = \
                        all_plate_data[barcode]["calculations"][method][other_calc]

    return temp_hits, data_calc_dict, plate_counter, all_states, all_methods


def _ws_creator(wb, name):

    return wb.create_sheet(f"{name}_Matrix")


def _matrix_writer(ws, data_calc_dict, state, plate_counter, all_methods):
    init_row = 2
    init_col = 2
    spacer = 4

    col_stdev = init_col + plate_counter + spacer
    col_counter = init_col + 1
    row_counter = init_row + 1
    col_stdev_counter = col_stdev + 1
    row_offset = init_row

    for method in all_methods:
        temp_avg_list = []
        temp_stdev_list = []
        mw_col = col_counter
        mw_row = row_counter
        mw_col_stdev = col_stdev_counter

        for barcodes in data_calc_dict:
            # Writes Plate names in row and clm for avg
            ws.cell(column=init_col - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws.cell(column=col_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            # Writes Plate names in row and clm for stdev
            ws.cell(column=col_stdev - 1, row=row_counter, value=barcodes).font = Font(b=True)
            ws.cell(column=col_stdev_counter, row=row_offset - 1, value=barcodes).font = Font(b=True)

            for index_method, _ in enumerate(data_calc_dict[barcodes]):

                if index_method == 0:
                    # Writes method for avg
                    ws.cell(column=init_col, row=row_offset - 1, value=method).font = Font(b=True)
                    # Writes method for stdev
                    ws.cell(column=col_stdev, row=row_offset - 1, value=method).font = Font(b=True)
                    if method != "other_data":
                        for calc in data_calc_dict[barcodes][method][state]:
                            temp_value = data_calc_dict[barcodes][method][state][calc]
                            # gets avg values
                            if calc == "avg":
                                ws.cell(column=init_col, row=row_offset, value=calc).font = Font(b=True)
                                ws.cell(column=init_col, row=row_counter, value=temp_value)
                                ws.cell(column=col_counter, row=row_offset, value=temp_value)
                                temp_avg_list.append(temp_value)
                            elif calc == "stdev":
                                ws.cell(column=col_stdev, row=row_offset, value=calc).font = Font(b=True)
                                ws.cell(column=col_stdev, row=row_counter, value=temp_value)
                                ws.cell(column=col_stdev_counter, row=row_offset, value=temp_value)
                                temp_stdev_list.append(temp_value)
            # Sets offset for next loop, for writing headlines the right place
            col_counter += 1
            row_counter += 1
            col_stdev_counter += 1

        # calculate the % difference between avg for each plate
        _matrix_calculator(ws, mw_row, mw_col, temp_avg_list)
        # calculate the % difference between stdev for each plate
        _matrix_calculator(ws, mw_row, mw_col_stdev, temp_stdev_list)
        # makes sure that next loop is writen below the first method. One method per row, with avg and stdev for each.
        col_stdev = init_col + plate_counter + spacer
        col_counter = init_col + 1
        row_counter += spacer
        col_stdev_counter = col_stdev + 1
        row_offset += (plate_counter + spacer)


def _matrix_calculator(ws, row, col, temp_data_list):
    start_row = row
    start_col = col

    for index_x, _ in enumerate(temp_data_list):
        for index_y, _ in enumerate(temp_data_list):
            try:
                temp_value = (float(temp_data_list[index_x]) / float(temp_data_list[index_y])) * 100
            except ZeroDivisionError:
                temp_value = "Na"
            ws.cell(column=start_col + index_x, row=start_row + index_y, value=temp_value)


def _z_prime(ws, data_calc_dict):

    init_row = 2
    init_col = 2

    col_counter = init_col + 1
    row_counter = init_row + 1

    z_prime_list = []

    for barcodes in data_calc_dict:
        # Writes Plate names
        ws.cell(column=init_col-1, row=row_counter, value=barcodes).font = Font(b=True)
        ws.cell(column=col_counter, row=init_row-1, value=barcodes).font = Font(b=True)
        # Writes values for Z-Prime
        z_prime = data_calc_dict[barcodes]["other_data"]["z_prime"]
        ws.cell(column=init_col, row=row_counter, value=z_prime)
        ws.cell(column=col_counter, row=init_row, value=z_prime)
        col_counter += 1
        row_counter += 1
        z_prime_list.append(z_prime)

    col_counter = init_col + 1
    row_counter = init_row + 1

    for index_x, _ in enumerate(z_prime_list):
        for index_y, _ in enumerate(z_prime_list):
            temp_value = (z_prime_list[index_x] / z_prime_list[index_y]) * 100
            ws.cell(column=col_counter + index_x, row=row_counter + index_y, value=temp_value)


def bio_final_report_controller(analyse_method, all_plate_data, output_file, final_report_setup):
    wb = Workbook()
    ws_report = wb.active
    ws_report.title = "Full report"
    ws_well_info = wb.create_sheet("Well Info")
    ws_z_prime = wb.create_sheet("Z-Prime")
    # ws_minimum = wb.create_sheet("Minimum")
    # ws_maximum = wb.create_sheet("Maximum")

    init_row = 2
    init_col = 2
    row = init_row
    col = init_col
    # calc overview:

    for index, barcode in enumerate(all_plate_data):
        ws, row_counter = _cal_writer_final_report(barcode, ws_report, all_plate_data[barcode], row, col,
                                                   final_report_setup["calc"])
        # Writes 5 plates horizontal, before changing rows.
        col += 5

        if index % 5 == 0 and index > 0:
            row += row_counter
            col = init_col

    # gets data:
    temp_hits, data_calc_dict, plate_counter, all_states, all_methods = _get_data(all_plate_data, final_report_setup)

    # write well data
    _well_writer_final_report(ws_well_info, temp_hits, final_report_setup, init_row)

    # writes Matrix of data:
    # inside guard ! ! ! !
    print(all_states)
    for states in all_states:
        if final_report_setup["full_report_matrix"][states]:
            _matrix_writer(_ws_creator(wb, states), data_calc_dict, states, plate_counter, all_methods)

    # writes Z-prime
    if final_report_setup["full_report_matrix"]["z_prime"]:
        _z_prime(ws_z_prime, data_calc_dict)

    wb.save(output_file)
