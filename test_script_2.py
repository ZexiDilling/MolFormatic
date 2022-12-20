import configparser

from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook, load_workbook
from matplotlib import colors
from bio_data_functions import heatmap

def tester(data_file):
    wb = Workbook()
    ws = wb.active
    row = 1
    col = 1

    with open(data_file) as file:
        for line in file:
            line = line.strip().split()
            print(line)
            for col_counter, data in enumerate(line):
                ws.cell(row=row, column=col + col_counter, value=data)
            row += 1

    save_data = data_file.removesuffix(".txt.txt")
    save_data = f"{save_data}.xlsx"
    wb.save(save_data)

def hex_to_rgb(hex):
    ''' "#FFFFFF" -> [255,255,255] '''
    # Pass 16 to the integer function for change of base
    return [int(hex[i:i + 2], 16) for i in range(1, 6, 2)]

if __name__ == "__main__":
    color = "red"
    color.removeprefix()
    cell_color = colors.cnames[color]
    config = configparser.ConfigParser()
    config.read("config.ini")
    print(hex_to_rgb(config["Settings_bio"]["plate_report_heatmap_colours_low"]))

    #
    # for counter in range(14):
    #     data_file = f"C:/Users/phch/Desktop/Daniels data/alpha_so{counter + 1}.txt.txt"
    #     tester(data_file)

