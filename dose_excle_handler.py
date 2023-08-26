import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

from bio_report_setup import _fetch_smiles_data
from excel_handler import insert_structure


def _write_dose_readings(ws_readings, ws_reading_row, ws_reading_clm, temp_data, add_dose, add_state_data, replicates):

    if replicates:
        temp_sample_counter = 1
        temp_rep_counter = 0
        data_placement = {"dose": {"clm": [], "row": []},
                          "sheet": ws_readings}
    else:
        data_placement = {"dose": {"clm": [], "row": []},
                          "samples": {"clm": [], "row": []},
                          "sheet": ws_readings}

    if add_dose:
        temp_dose_clm = ws_reading_clm
        temp_dose_row = ws_reading_clm
        ws_reading_clm += 1

    if add_state_data:
        for states in temp_data["state_data"]:
            ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=states)
            for well_value in temp_data["state_data"][states]:
                ws_reading_row += 1
                ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=well_value)
            ws_reading_clm += 1
            ws_reading_row = 1

    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data":
            temp_sample_name = samples.split("_")
            temp_sample_name = f"{temp_sample_name[0]}_{temp_sample_name[1]}_{temp_sample_name[2]}_{temp_sample_name[3]}"
            ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=samples)
            ws_reading_row += 1
            for reading_data in temp_data[samples]["reading"]["raw"]:

                ws_readings.cell(column=ws_reading_clm, row=ws_reading_row, value=reading_data)

                if replicates:

                    data_placement_name = f"samples_{temp_sample_name}"
                    try:
                        data_placement[data_placement_name]
                    except KeyError:
                        data_placement[data_placement_name] = {"clm": [], "row": []}

                else:
                    data_placement_name = "samples"

                data_placement[data_placement_name]["row"].append(ws_reading_row)
                data_placement[data_placement_name]["clm"].append(ws_reading_clm)

                ws_reading_row += 1

            if add_dose:
                temp_name = f"Dose({temp_data[samples]['dose']['unit']})"
                ws_readings.cell(column=temp_dose_clm, row=temp_dose_row, value=temp_name)
                temp_dose_row += 1

                for reading_data in temp_data[samples]["dose"]["raw"]:
                    ws_readings.cell(column=temp_dose_clm, row=temp_dose_row, value=reading_data)

                    data_placement["dose"]["clm"].append(1)
                    data_placement["dose"]["row"].append(temp_dose_row)
                    temp_dose_row += 1
                add_dose = False
            if replicates:
                temp_rep_counter += 1
                if temp_rep_counter == replicates:
                    temp_sample_counter += 1
                    temp_rep_counter = 0
            ws_reading_clm += 1
            ws_reading_row = 1
    return data_placement
    # reading - raw
    # reading - normalized
    # reading - min
    # reading - max
    # reading - fitted_normalized
    # reading - fitted
    # dose - raw
    # dose - normalized
    # dose - min
    # dose - max
    # dose - fitted_normalized
    # dose - fitted


def _write_dose_fitted(ws_fitted, ws_fitted_row, ws_fitted_clm, temp_data, replicates):
    add_dose = True
    if replicates:
        temp_sample_counter = 1
        temp_rep_counter = 0
        data_placement = {"dose": {"clm": [], "row": []},
                          "sheet": ws_fitted}
    else:
        data_placement = {"dose": {"clm": [], "row": []},
                          "samples": {"clm": [], "row": []},
                          "sheet": ws_fitted}

    temp_dose_clm = ws_fitted_clm
    temp_dose_row = ws_fitted_row
    ws_fitted_clm += 1

    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data":
            temp_sample_name = samples.split("_")
            temp_sample_name = f"{temp_sample_name[0]}_{temp_sample_name[1]}_{temp_sample_name[2]}_{temp_sample_name[3]}"
            ws_fitted.cell(column=ws_fitted_clm, row=ws_fitted_row, value=samples)
            ws_fitted_row += 1
            for reading_data in temp_data[samples]["reading"]["fitted"]:

                ws_fitted.cell(column=ws_fitted_clm, row=ws_fitted_row, value=reading_data)

                if replicates:

                    data_placement_name = f"samples_{temp_sample_name}"
                    try:
                        data_placement[data_placement_name]
                    except KeyError:
                        data_placement[data_placement_name] = {"clm": [], "row": []}

                else:
                    data_placement_name = "samples"

                data_placement[data_placement_name]["row"].append(ws_fitted_row)
                data_placement[data_placement_name]["clm"].append(ws_fitted_clm)

                ws_fitted_row += 1

            if add_dose:
                temp_name = f"Dose({temp_data[samples]['dose']['unit']})"
                ws_fitted.cell(column=temp_dose_clm, row=temp_dose_row, value=temp_name)
                temp_dose_row += 1

                for reading_data in temp_data[samples]["dose"]["raw"]:
                    ws_fitted.cell(column=temp_dose_clm, row=temp_dose_row, value=reading_data)

                    data_placement["dose"]["clm"].append(1)
                    data_placement["dose"]["row"].append(temp_dose_row)
                    temp_dose_row += 1
                add_dose = False
            if replicates:
                temp_rep_counter += 1
                if temp_rep_counter == replicates:
                    temp_sample_counter += 1
                    temp_rep_counter = 0
            ws_fitted_clm += 1
            ws_fitted_row = 1
    return data_placement


def _write_dose_data(ws_data, ws_data_row, ws_data_clm, temp_data):

    # write_headlines
    check_list = ["EC50", "rsquared", "hillslope", "n_lowdose_datapoints", "std_resp_lowdose_datapoints",
                  "n_highdose_datapoints", "std_resp_highdose_datapoints", "doseconc_stepsize_at_EC50", "saxe_lowdose",
                  "saxe_highdose"]
    for sample_index, samples in enumerate(temp_data):
        if samples != "state_data" and sample_index == 0:
            for data in temp_data[samples]:
                if data in check_list:
                    ws_data.cell(column=ws_data_clm, row=ws_data_row + 1, value=data)
                    ws_data_row += 1

    ws_data_row = 1
    ws_data_clm += 1

    for sample_index, samples in enumerate(temp_data):

        if samples != "state_data":
            ws_data.cell(column=ws_data_clm, row=ws_data_row, value=samples)
            ws_data_row += 1
            for data in temp_data[samples]:
                if data in check_list:
                    temp_value = temp_data[samples][data]["value"]
                    ws_data.cell(column=ws_data_clm, row=ws_data_row, value=temp_value)
                    ws_data_row += 1
        ws_data_row = 1
        ws_data_clm += 1


def _draw_curves(ws_diagram, data_placement, x_title, y_title):

    x_values_min_clm = min(data_placement["dose"]["clm"])
    x_values_min_row = min(data_placement["dose"]["row"])
    x_values_max_row = max(data_placement["dose"]["row"])
    chart_data_sheet = data_placement["sheet"]
    x_values = Reference(chart_data_sheet, min_col=x_values_min_clm, min_row=x_values_min_row, max_row=x_values_max_row)
    chart_list = ["B2", "L2", "V2", "B17", "L17", "V17", "B32", "L32", "V32"]
    chart_counter = 0

    for samples in data_placement:
        if samples != "dose" and samples != "sheet":
            chart = ScatterChart()
            chart.title = f"{samples}"
            chart.style = 15
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            y_value_min_clm = min(data_placement[samples]["clm"])
            y_value_max_clm = max(data_placement[samples]["clm"])
            Y_values_min_row = min(data_placement[samples]["row"])
            Y_values_max_row = max(data_placement[samples]["row"])

            for clm in range(y_value_min_clm, y_value_max_clm + 1):
                values = Reference(chart_data_sheet, min_col=clm, min_row=Y_values_min_row - 1, max_row=Y_values_max_row)
                series = Series(values, x_values, title_from_data=True)
                series.marker = openpyxl.chart.marker.Marker('x')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
            ws_diagram.add_chart(chart, chart_list[chart_counter])
            chart_counter += 1


def _write_overview(config, all_dose_data, ws_overvew, plate_group_to_compound_id, include_structure):
    temp_row = 1
    temp_clm = 1
    if include_structure:
        headlines = ["Temp_name", "Compound_id", "smiles", "structure"]
    else:
        headlines = ["Temp_name", "Compound_id", "smiles"]

    for headline in headlines:
        ws_overvew.cell(column=temp_clm, row=temp_row, value=headline)
        temp_row += 1

    temp_row = 1
    temp_clm += 1

    for plates in all_dose_data:
        for samples in all_dose_data[plates]:
            temp_name = samples
            compound_id = plate_group_to_compound_id[samples]
            smiles = _fetch_smiles_data(config, samples)

            ws_overvew.cell(column=temp_clm, row=temp_row + 0, value=temp_name)
            ws_overvew.cell(column=temp_clm, row=temp_row + 1, value=compound_id)
            ws_overvew.cell(column=temp_clm, row=temp_row + 2, value=smiles)

        temp_clm += 1

    if include_structure:
        insert_structure(ws_overvew)


def dose_excel_controller(config, plate_reader_files, all_dose_data, plate_group_to_compound_id, save_location,
                          include_id, include_structure):
    temp_counter = 0
    for files in plate_reader_files:
        temp_counter += 1
        temp_name = f"{temp_counter} - {files}"
        print(temp_name)

    temp_counter = 0
    for plates in all_dose_data:
        temp_counter += 1
        temp_name = f"{temp_counter} - {plates}"
        print(temp_name)

        wb = Workbook()
        ws_readings = wb.create_sheet("Readings")
        ws_fitted = wb.create_sheet("Fitted")
        ws_data = wb.create_sheet("Data")
        ws_diagram_raw = wb.create_sheet("Diagram_Raw")
        x_title_raw = "Dose"
        y_title_raw = "Signal"
        ws_diagram_fitted = wb.create_sheet("Diagram_Fitted")
        x_title_fitted = "Dose"
        y_title_fitted = "Fitted"
        # Added to make sure that dose is only added once
        add_dose = True
        add_state_data = True
        replicates = 3

        ws_reading_row = ws_reading_clm = ws_fitted_row = ws_fitted_clm = ws_data_row = ws_data_clm = \
            ws_diagram_row = ws_diagram_clm = 1

        temp_data = all_dose_data[plates]
        data_placement_raw = _write_dose_readings(ws_readings, ws_reading_row, ws_reading_clm, temp_data, add_dose,
                                                  add_state_data, replicates)
        data_placement_fitted = _write_dose_fitted(ws_fitted, ws_fitted_row, ws_fitted_clm, temp_data, replicates)
        _write_dose_data(ws_data, ws_data_row, ws_data_clm, temp_data)
        _draw_curves(ws_diagram_raw, data_placement_raw, x_title_raw, y_title_raw)
        _draw_curves(ws_diagram_fitted, data_placement_fitted, x_title_fitted, y_title_fitted)
        if include_id:
            ws_overview = wb.create_sheet("Overview")
            _write_overview(config, all_dose_data, ws_overview, plate_group_to_compound_id, include_structure)

        name = save_location/f"{plates}_dose_response.xlsx"
        wb.save(name)


if __name__ == "__main__":
    pass